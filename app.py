# !/usr/bin/env python3
# -*- coding: utf-8 -*-

import json
import math
import os
import queue
import re
import shutil
import subprocess
import threading
import zipfile
import html
import io
import redis
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from uuid import uuid4

import pandas as pd
import requests
from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker
from werkzeug.utils import secure_filename
from functools import wraps

try:
    from dotenv import load_dotenv
except ImportError:
    load_dotenv = None

BASE_DIR = os.path.dirname(__file__)
if load_dotenv:
    load_dotenv(os.path.join(BASE_DIR, '.env'))

app = Flask(__name__)
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', os.urandom(32).hex())
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(__file__), 'uploads')
app.config['UC_SUBMISSION_FOLDER'] = os.path.join(app.config['UPLOAD_FOLDER'], 'uc_submissions')
app.config['BAIDU_SUBMISSION_FOLDER'] = os.path.join(app.config['UPLOAD_FOLDER'], 'baidu_submissions')
app.config['TASK_RESULT_FOLDER'] = os.path.join(os.path.dirname(__file__), 'task_results')
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0

DB_HOST = os.getenv('DB_HOST', '127.0.0.1')
DB_PORT = os.getenv('DB_PORT', '3306')
DB_NAME = os.getenv('DB_NAME', 'complaint_form')
DB_USER = os.getenv('DB_USER', 'navicat')
DB_PASSWORD = os.getenv('DB_PASSWORD', 'navicat123')
DATABASE_URL = os.getenv(
    'DATABASE_URL',
    f'mysql+pymysql://{DB_USER}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/{DB_NAME}?charset=utf8mb4'
)
REDIS_URL = os.getenv('REDIS_URL', 'redis://127.0.0.1:6379/0')
UC_QUEUE_NAME = os.getenv('UC_QUEUE_NAME', 'uc_complaint_queue')
UNIFIED_QUEUE_NAME = os.getenv('UNIFIED_QUEUE_NAME', 'complaint_queue')
UC_WORKER_LOCK_KEY = os.getenv('UC_WORKER_LOCK_KEY', 'uc_complaint_worker_lock')
UC_WORKER_LOCK_TTL = int(os.getenv('UC_WORKER_LOCK_TTL', '15'))
UC_COMPLAIN_LIST_API = 'https://ipp.uc.cn/api/complain/accuse'

PRINCIPAL_UPDATE_UID_WHITELIST = {'1717602'}

BAIDU_QUEUE_NAME = os.getenv('BAIDU_QUEUE_NAME', 'baidu_complaint_queue')
BAIDU_WORKER_LOCK_KEY = os.getenv('BAIDU_WORKER_LOCK_KEY', 'baidu_complaint_worker_lock')
BAIDU_WORKER_LOCK_TTL = int(os.getenv('BAIDU_WORKER_LOCK_TTL', '15'))
BAIDU_API_BASE = 'https://newcopyright.baidu.com'
BAIDU_COMPLAINT_TYPE_MAP = {
    '百度搜索': 1401,
    '百度网盘': 1402,
    '好看视频': 1408,
    '百家号': 1410,
    '百度APP': 1407,
    '百度知道': 1405,
    '百度文库': 1403,
    '百度贴吧': 1404,
    '百度图片': 1406,
    '度小视': 1409,
    '百度手机浏览器': 1412,
}
BAIDU_INFRINGE_TYPE_MAP = {
    '影视版权': 1, '综艺版权': 2, '动漫动画版权': 3, '音乐版权': 4,
    '游戏版权': 5, '体育赛事版权': 6, '新闻媒体版权': 7, '自媒体版权': 8,
}
BAIDU_WORKS_CATEGORY_MAP = {
    1: '文字', 2: '图片', 3: '音乐', 4: '软件',
    5: '视听作品(影视)', 6: '视听作品(综艺)', 7: '视听作品(动漫)',
    8: '视听作品(其他)', 9: '其他作品', 11: '软件(游戏)',
    12: '软件(社交)', 13: '软件(工具)', 14: '软件(其它)',
    15: '视听作品(短剧)',
}
BAIDU_OWNER_TYPE_MAP = {1: '权利人', 2: '代理人'}

engine = create_engine(DATABASE_URL, pool_pre_ping=True, future=True)
SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)

# 任务状态存储（生产环境建议用数据库）
tasks = {}

# ==================== 登录系统 ====================
import time
from auth_client import login as auth_login, verify_token


def get_client_ip():
    return request.headers.get('X-Forwarded-For', request.remote_addr).split(',')[0].strip()


def get_current_user():
    return session.get('username', '')


def can_update_principal_authorization():
    return str(session.get('uid') or '') in PRINCIPAL_UPDATE_UID_WHITELIST


def get_current_operator_name():
    return get_current_user()


# 登录有效期（秒），12小时
LOGIN_EXPIRE_SECONDS = 43200


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = session.get('token')
        if not token:
            if request.is_json or request.path.startswith('/api/'):
                return jsonify({'success': False, 'error': '未登录', 'login_required': True}), 401
            return redirect(url_for('login_page', next=request.path))
        login_time = session.get('login_time', 0)
        if time.time() - login_time > LOGIN_EXPIRE_SECONDS:
            session.clear()
            if request.is_json or request.path.startswith('/api/'):
                return jsonify({'success': False, 'error': '登录已过期，请重新登录', 'login_required': True}), 401
            return redirect(url_for('login_page', next=request.path))
        v = verify_token(token, get_client_ip())
        if not v['valid']:
            session.clear()
            if request.is_json or request.path.startswith('/api/'):
                return jsonify({'success': False, 'error': v['error'], 'login_required': True}), 401
            return redirect(url_for('login_page', next=request.path))
        return f(*args, **kwargs)
    return decorated


@app.route('/login')
def login_page():
    if session.get('token'):
        v = verify_token(session['token'], get_client_ip())
        if v['valid']:
            return redirect('/')
    return render_template('login.html')


@app.route('/api/login', methods=['POST'])
def api_login():
    data = request.get_json()
    username = (data or {}).get('username', '').strip()
    password = (data or {}).get('password', '').strip()
    if not username or not password:
        return jsonify({'success': False, 'error': '用户名和密码不能为空'})

    client_ip = get_client_ip()
    result = auth_login(username, password, client_ip)

    if not result['success']:
        return jsonify({'success': False, 'error': result['error']})

    session['token'] = result['token']
    session['username'] = result['user_info']['username']
    session['uid'] = result['user_info'].get('uid')
    session['level'] = result['user_info'].get('level')
    session['login_time'] = time.time()

    next_url = (data or {}).get('next') or request.args.get('next') or '/'
    return jsonify({'success': True, 'redirect': next_url})


@app.route('/api/logout', methods=['POST'])
def api_logout():
    session.clear()
    return jsonify({'success': True})


# ==================== END 登录系统 ====================


@app.context_processor
def inject_user():
    return {'current_user': session.get('username', ''), 'current_token': session.get('token', '')}


@app.after_request
def no_cache(response):
    response.headers['Cache-Control'] = 'no-store'
    return response

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['UC_SUBMISSION_FOLDER'], exist_ok=True)
os.makedirs(app.config['BAIDU_SUBMISSION_FOLDER'], exist_ok=True)
os.makedirs(app.config['TASK_RESULT_FOLDER'], exist_ok=True)


def ensure_dir(path):
    os.makedirs(path, exist_ok=True)


def extract_xtstk_from_cookie(cookie_str):
    m = re.search(r'cmptstk=([^;]+)', cookie_str or '')
    if not m:
        raise RuntimeError('cookie 中找不到 cmptstk，无法构造 xtstk 请求头')
    return m.group(1).strip()


def save_uploaded_file(file_storage, target_dir, prefix=None):
    if not file_storage or not file_storage.filename:
        return None

    original_name = Path(file_storage.filename).name
    suffix = Path(original_name).suffix.lower()
    stem = secure_filename(Path(original_name).stem)

    if not suffix:
        suffix = Path(original_name).suffix

    if prefix:
        filename = f"{prefix}{suffix}" if suffix else prefix
    else:
        filename = f"{stem}{suffix}" if stem else None

    if not filename:
        return None

    save_path = os.path.join(target_dir, filename)
    file_storage.save(save_path)
    return filename




def create_submission_dir():
    submission_id = datetime.now().strftime('%Y%m%d_%H%M%S_') + uuid4().hex[:8]
    submission_dir = os.path.join(app.config['UC_SUBMISSION_FOLDER'], submission_id)
    ensure_dir(submission_dir)
    return submission_id, submission_dir


def split_excel_into_batches(df, batch_dir, batch_size=200):
    ensure_dir(batch_dir)
    batches = []

    for index, start in enumerate(range(0, len(df), batch_size), start=1):
        end = min(start + batch_size, len(df))
        batch_df = df.iloc[start:end].copy()
        batch_filename = f'part_{index:03d}.xlsx'
        batch_path = os.path.join(batch_dir, batch_filename)
        batch_df.to_excel(batch_path, index=False)
        batches.append({
            'batch_no': index,
            'filename': batch_filename,
            'path': batch_path,
            'start_row': start + 1,
            'end_row': end,
            'rows': len(batch_df),
        })

    return batches


def load_task_result(task_id):
    result_path = os.path.join(app.config['TASK_RESULT_FOLDER'], f'{task_id}.json')
    if not os.path.exists(result_path):
        return None

    with open(result_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def append_task_log(task_id, message):
    log_path = os.path.join(app.config['TASK_RESULT_FOLDER'], f'{task_id}.log')
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with open(log_path, 'a', encoding='utf-8') as f:
        f.write(f'[{timestamp}] {message}\n')


def read_task_log_file(task_id):
    log_path = os.path.join(app.config['TASK_RESULT_FOLDER'], f'{task_id}.log')
    if not os.path.exists(log_path):
        return None
    with open(log_path, 'r', encoding='utf-8') as f:
        return f.read()


def cleanup_old_task_logs(max_age_days=5):
    cutoff = datetime.now().timestamp() - max_age_days * 24 * 3600
    task_results_dir = Path(app.config['TASK_RESULT_FOLDER'])
    if not task_results_dir.exists():
        return
    for path in task_results_dir.iterdir():
        if not path.is_file():
            continue
        try:
            if path.stat().st_mtime < cutoff:
                path.unlink()
        except FileNotFoundError:
            continue


def upsert_task_execution_log(task_id, submission_id, status, log_text):
    pass


def sync_task_log_to_db(task_id, submission_id, status):
    pass


def get_task_execution_log(task_id):
    return None


def has_available_task_log(task_id):
    return bool(read_task_log_file(task_id))


def get_redis_client():
    return redis.Redis.from_url(REDIS_URL, decode_responses=True)


def enqueue_uc_task(task_payload):
    client = get_redis_client()
    task_payload['platform'] = 'uc'
    client.lpush(UNIFIED_QUEUE_NAME, json.dumps(task_payload, ensure_ascii=False))


def dequeue_uc_task(timeout=0):
    client = get_redis_client()
    item = client.brpop(UC_QUEUE_NAME, timeout=timeout)
    if not item:
        return None
    _, payload = item
    return json.loads(payload)


def acquire_worker_lock(ttl_seconds=None):
    client = get_redis_client()
    token = uuid4().hex
    ttl = ttl_seconds or UC_WORKER_LOCK_TTL
    acquired = client.set(UC_WORKER_LOCK_KEY, token, nx=True, ex=ttl)
    return token if acquired else None


def refresh_worker_lock(token, ttl_seconds=None):
    client = get_redis_client()
    ttl = ttl_seconds or UC_WORKER_LOCK_TTL
    current = client.get(UC_WORKER_LOCK_KEY)
    if current != token:
        return False
    client.expire(UC_WORKER_LOCK_KEY, ttl)
    return True


def release_worker_lock(token):
    client = get_redis_client()
    current = client.get(UC_WORKER_LOCK_KEY)
    if current == token:
        client.delete(UC_WORKER_LOCK_KEY)


def enqueue_baidu_task(task_payload):
    client = get_redis_client()
    task_payload['platform'] = 'baidu'
    client.lpush(UNIFIED_QUEUE_NAME, json.dumps(task_payload, ensure_ascii=False))


def dequeue_baidu_task(timeout=0):
    client = get_redis_client()
    item = client.brpop(BAIDU_QUEUE_NAME, timeout=timeout)
    if not item:
        return None
    _, payload = item
    return json.loads(payload)


def dequeue_unified_task(timeout=0):
    client = get_redis_client()
    item = client.brpop(UNIFIED_QUEUE_NAME, timeout=timeout)
    if not item:
        return None
    _, payload = item
    return json.loads(payload)


@app.route('/')
@login_required
def index():
    return render_template('index.html', is_index=True, login_time_marker=str(session.get('login_time') or ''))


@app.route('/works')
@login_required
def works():
    return render_template('works.html')


# 平台映射
def get_db_session():
    return SessionLocal()


def _load_platform_map():
    try:
        with get_db_session() as session:
            rows = session.execute(text(
                "SELECT platform_code, platform_name FROM platforms"
            )).fetchall()
        return {row[0]: {'platform_name': row[1], 'pingtai': row[1]} for row in rows}
    except Exception:
        return {
            'uc': {'platform_name': 'UC', 'pingtai': 'UC'},
            'quark': {'platform_name': '夸克', 'pingtai': '夸克'},
            'baidu': {'platform_name': '百度', 'pingtai': '百度'},
        }


PLATFORM_MAP = _load_platform_map()


def get_platforms_list():
    with get_db_session() as session:
        rows = session.execute(text("""
            SELECT platform_code, platform_name, status
            FROM platforms
            WHERE status = 'active'
            ORDER BY sort_order ASC
        """)).mappings().all()
    return [dict(row) for row in rows]


def normalize_datetime(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value)


def row_to_account_dict(row):
    return {
        'id': row.account_id,
        'platform_code': row.platform_code,
        'platform_name': row.platform_name,
        'pingtai': row.platform_label,
        'used_company': row.used_company,
        'user': row.account_user,
        'cookie': row.cookie_text,
        'account_purpose': row.account_purpose,
        'status': row.status,
        'created_at': normalize_datetime(row.created_at),
        'updated_at': normalize_datetime(row.updated_at),
    }


def guess_mime_type(filename):
    suffix = Path(filename).suffix.lower()
    mapping = {
        '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        '.xls': 'application/vnd.ms-excel',
        '.png': 'image/png',
        '.jpg': 'image/jpeg',
        '.jpeg': 'image/jpeg',
        '.pdf': 'application/pdf',
        '.bmp': 'image/bmp',
    }
    return mapping.get(suffix, 'application/octet-stream')


def build_file_asset_row(business_type, business_id, category, local_path, saved_name, original_name=None):
    path_obj = Path(local_path)
    return {
        'asset_id': uuid4().hex[:12],
        'business_type': business_type,
        'business_id': business_id,
        'category': category,
        'storage_type': 'local',
        'bucket_name': None,
        'object_key': None,
        'local_path': str(path_obj),
        'original_name': original_name or saved_name,
        'saved_name': saved_name,
        'mime_type': guess_mime_type(saved_name),
        'file_size': path_obj.stat().st_size if path_obj.exists() else 0,
        'file_hash': None,
        'created_at': datetime.now(),
    }


def insert_file_asset(asset_row):
    pass



def register_submission_files(*args, **kwargs):
    pass



def migrate_submission_file_assets_if_needed():
    pass



def load_accounts():
    with get_db_session() as session:
        rows = session.execute(text("""
            SELECT account_id, platform_code, platform_name, platform_label,
                   used_company, account_user, cookie_text, account_purpose, status,
                   created_at, updated_at
            FROM accounts
            ORDER BY id ASC
        """)).mappings().all()
        return [row_to_account_dict(row) for row in rows]


def load_principals_map():
    with get_db_session() as session:
        rows = session.execute(text("""
            SELECT id, platform_code, used_company, account_user,
                   principal_name, created_at, updated_at
            FROM principals
            ORDER BY id ASC
        """)).mappings().all()

    principals_map = {}
    for row in rows:
        key = f"{row.platform_code}:{row.account_user}"
        entry = principals_map.setdefault(key, {
            'group_id': str(row.id),
            'platform_code': row.platform_code,
            'platform_name': PLATFORM_MAP.get(row.platform_code, {}).get('platform_name', row.platform_code),
            'account_user': row.account_user,
            'principals': [],
            'created_at': normalize_datetime(row.created_at),
            'updated_at': normalize_datetime(row.updated_at),
        })
        if row.principal_name and row.principal_name not in entry['principals']:
            entry['principals'].append(row.principal_name)
    return principals_map


def normalize_company_name(value):
    return (value or '').strip().replace('（', '(').replace('）', ')')


def normalize_filename_part(value):
    import re
    normalized = (value or '').strip().replace('（', '(').replace('）', ')')
    return re.sub(r'[/:*?"<>|\\：＊？＂＜＞｜＼／]', '', normalized)


# 真 emoji 的 Unicode 区段（窄匹配，只动真 emoji，不碰正常标点/CJK/英数）。
# 百度 /upload 的 link_name 含 emoji 会直接返回 code=500「提交失败」。
_EMOJI_PATTERN = re.compile(
    "["
    "\U0001F300-\U0001FAFF"  # 表情/图形/交通/补充符号等（😂🥵🥰😭…）
    "\U0001F1E6-\U0001F1FF"  # 区域指示符（国旗）
    "\U00002600-\U000026FF"  # 杂项符号（☀☂…）
    "\U00002700-\U000027BF"  # 装饰符 Dingbats（✂✅…）
    "\U00002B00-\U00002BFF"  # 箭头/星标（⭐⬆…）
    "\U0000FE00-\U0000FE0F"  # 变体选择符（emoji 变体）
    "\U0000200D"             # 零宽连接符 ZWJ（组合 emoji）
    "\U000020E3"             # keycap 组合符
    "]+",
    flags=re.UNICODE,
)


def strip_emoji(value):
    """移除字符串中的 emoji（窄匹配，仅真 emoji）。
    返回 (清洗后文本, 是否有改动)。清洗后合并多余空格并去首尾空白。"""
    if not value or not isinstance(value, str):
        return value, False
    cleaned = _EMOJI_PATTERN.sub('', value)
    if cleaned == value:
        return value, False
    cleaned = re.sub(r'\s{2,}', ' ', cleaned).strip()
    return cleaned, True


# 链接地址(Sheet3 C列)非法字符：汉字 + 全角标点。含这些说明链接本身有误，
# 不能自动改，须用户删除后重传。percent 编码(%E7%99%BD…)是纯 ASCII，不受影响。
_URL_ILLEGAL_PATTERN = re.compile(
    "["
    "一-鿿"   # CJK 汉字
    "㐀-䶿"   # CJK 扩展A（生僻字）
    "　-〿"   # CJK 标点（。、《》「」 等，含全角空格）
    "＀-￯"   # 全角 ASCII/标点（，！？（）：； 等）
    "]"
)


def find_illegal_url_chars(url):
    """返回链接地址中出现的汉字/全角标点（去重保序），无则返回空字符串。"""
    if not url:
        return ''
    found = _URL_ILLEGAL_PATTERN.findall(url)
    return ''.join(dict.fromkeys(found))


def get_principal_document_record(platform_code, used_company, principal_name):
    normalized_company = normalize_company_name(used_company)
    normalized_principal = normalize_company_name(principal_name)
    with get_db_session() as session:
        row = session.execute(text("""
            SELECT business_license_filename, authorization_filename, authorization_expires_on
            FROM principals
            WHERE principal_name = :principal_name
              AND platform_code = :platform_code
              AND used_company = :used_company
            ORDER BY updated_at DESC
            LIMIT 1
        """), {
            'principal_name': normalized_principal,
            'platform_code': platform_code,
            'used_company': normalized_company,
        }).mappings().first()

    if not row:
        return None

    data = {
        'principal_name': normalized_principal,
        'used_company': normalized_company,
        'business_license_filename': row.get('business_license_filename') or None,
        'authorization_filename': row.get('authorization_filename') or None,
        'authorization_expires_on': normalize_datetime(row.get('authorization_expires_on')) if row.get('authorization_expires_on') else None,
        'business_license_locked': bool(row.get('business_license_filename')),
        'authorization_locked': bool(row.get('authorization_filename')),
    }
    return data


def get_authorization_expiry_alerts(days_threshold=30):
    with get_db_session() as session:
        rows = session.execute(text("""
            SELECT DISTINCT principal_name, authorization_expires_on
            FROM principals
            WHERE authorization_filename IS NOT NULL
              AND authorization_filename <> ''
              AND authorization_expires_on IS NOT NULL
              AND principal_name IS NOT NULL
              AND principal_name <> ''
            ORDER BY authorization_expires_on ASC, principal_name ASC
        """)).mappings().all()

    today = datetime.now().date()
    alerts = []
    for row in rows:
        principal_name = normalize_company_name(row.get('principal_name'))
        expires_value = row.get('authorization_expires_on')
        if not principal_name or not expires_value:
            continue

        try:
            expires_date = expires_value.date() if isinstance(expires_value, datetime) else datetime.fromisoformat(str(expires_value)).date()
        except Exception:
            continue

        delta_days = (expires_date - today).days
        if delta_days < 0:
            alerts.append({
                'principal_name': principal_name,
                'expires_on': expires_date.isoformat(),
                'days': abs(delta_days),
                'status': 'expired',
                'message': f'被代理人信息管理页面中，{principal_name}的授权委托书已过期{abs(delta_days)}天，请及时更新',
            })
        elif delta_days <= days_threshold:
            alerts.append({
                'principal_name': principal_name,
                'expires_on': expires_date.isoformat(),
                'days': delta_days,
                'status': 'expiring',
                'message': f'被代理人信息管理页面中，{principal_name}的授权委托书将于{delta_days}天到期，请及时更新',
            })

    return alerts


def check_principal_authorization_blocked(principal_name, platform_code='uc', account_user=''):
    """检查所选被代理人的授权是否已过期 / 缺少截止日期。
    返回错误提示文案（需拦截）或 None（可正常投诉）。"""
    normalized_principal = normalize_company_name(principal_name)
    if not normalized_principal:
        return None

    params = {
        'principal_name': normalized_principal,
        'platform_code': platform_code,
    }
    account_filter = ''
    if account_user:
        account_filter = 'AND account_user = :account_user'
        params['account_user'] = account_user

    with get_db_session() as session:
        row = session.execute(text(f"""
            SELECT authorization_expires_on
            FROM principals
            WHERE principal_name = :principal_name
              AND platform_code = :platform_code
              {account_filter}
            ORDER BY updated_at DESC
            LIMIT 1
        """), params).mappings().first()

    if not row:
        return None

    expires_value = row.get('authorization_expires_on')
    if not expires_value:
        return (f'所选被代理人「{normalized_principal}」未设置授权截止日期，无法确认授权是否有效，'
                f'暂不能发起投诉。请先到「被代理人信息管理」补充授权委托书及截止日期后再试。')

    try:
        expires_date = expires_value.date() if isinstance(expires_value, datetime) else datetime.fromisoformat(str(expires_value)).date()
    except Exception:
        return (f'所选被代理人「{normalized_principal}」未设置授权截止日期，无法确认授权是否有效，'
                f'暂不能发起投诉。请先到「被代理人信息管理」补充授权委托书及截止日期后再试。')

    delta_days = (expires_date - datetime.now().date()).days
    if delta_days < 0:
        return (f'所选被代理人「{normalized_principal}」的授权委托书已于 {expires_date.isoformat()} 过期'
                f'（已过期 {abs(delta_days)} 天），无法发起投诉。请先到「被代理人信息管理」更新授权委托书后再试。')

    return None


def save_named_upload(file_storage, target_dir, target_name_without_ext):
    if not file_storage or not file_storage.filename:
        return None
    suffix = Path(file_storage.filename).suffix.lower()
    filename = f"{target_name_without_ext}{suffix}"
    save_path = os.path.join(target_dir, filename)
    file_storage.save(save_path)
    return filename


PRINCIPAL_UPLOAD_MAX_SIZE = 5 * 1024 * 1024


def validate_principal_upload_filenames(principal_name, used_company, authorization_expires_on,
                                        business_license_file=None, authorization_file=None):
    normalized_principal = normalize_company_name(principal_name)
    normalized_used_company = normalize_company_name(used_company)
    expires_yyyymmdd = (authorization_expires_on or '').replace('-', '')

    if business_license_file and business_license_file.filename:
        business_stem = normalize_company_name(Path(business_license_file.filename).stem)
        expected_business_stem = f'营业执照_{normalized_principal}'
        if business_stem != expected_business_stem:
            return f'被代理人营业执照文件名不符合要求，请上传命名为“{expected_business_stem}.文件后缀”的文件'

    if authorization_file and authorization_file.filename:
        authorization_stem = normalize_company_name(Path(authorization_file.filename).stem)
        expected_authorization_stem = f'授权委托书_{normalized_principal}_{normalized_used_company}_截止日期{expires_yyyymmdd}'
        if authorization_stem != expected_authorization_stem:
            return f'授权委托书文件名不符合要求，请上传命名为“{expected_authorization_stem}.文件后缀”的文件'

    return None


def validate_principal_upload_file_sizes(business_license_file=None, authorization_file=None):
    files_to_check = []
    if business_license_file and business_license_file.filename:
        files_to_check.append(('被代理人营业执照文件', business_license_file))
    if authorization_file and authorization_file.filename:
        files_to_check.append(('授权委托书文件', authorization_file))

    for label, file_storage in files_to_check:
        current_pos = file_storage.stream.tell()
        file_storage.stream.seek(0, os.SEEK_END)
        size = file_storage.stream.tell()
        file_storage.stream.seek(current_pos)
        if size > PRINCIPAL_UPLOAD_MAX_SIZE:
            return f'{label}不能超过5MB'

    return None


def upsert_principal_documents(platform_code, used_company, account_user, principal_name,
                              business_license_filename, authorization_filename, authorization_expires_on):
    normalized_company = normalize_company_name(used_company)
    normalized_principal = normalize_company_name(principal_name)
    with get_db_session() as session:
        existing = session.execute(text("""
            SELECT id
            FROM principals
            WHERE platform_code = :platform_code
              AND account_user = :account_user
              AND principal_name = :principal_name
            LIMIT 1
        """), {
            'platform_code': platform_code,
            'account_user': account_user,
            'principal_name': normalized_principal,
        }).mappings().first()
        if existing:
            session.execute(text("""
                UPDATE principals
                SET used_company = :used_company,
                    business_license_filename = :business_license_filename,
                    authorization_filename = :authorization_filename,
                    authorization_expires_on = :authorization_expires_on,
                    updated_at = NOW()
                WHERE id = :id
            """), {
                'used_company': normalized_company,
                'business_license_filename': business_license_filename,
                'authorization_filename': authorization_filename,
                'authorization_expires_on': authorization_expires_on,
                'id': existing['id'],
            })
        else:
            session.execute(text("""
                INSERT INTO principals (
                    platform_code, used_company, account_user, principal_name,
                    business_license_filename, authorization_filename, authorization_expires_on
                ) VALUES (
                    :platform_code, :used_company, :account_user, :principal_name,
                    :business_license_filename, :authorization_filename, :authorization_expires_on
                )
            """), {
                'platform_code': platform_code,
                'used_company': normalized_company,
                'account_user': account_user,
                'principal_name': normalized_principal,
                'business_license_filename': business_license_filename,
                'authorization_filename': authorization_filename,
                'authorization_expires_on': authorization_expires_on,
            })
        session.commit()


def get_principal_options_by_used_company(used_company):
    if not used_company:
        return []
    with get_db_session() as session:
        rows = session.execute(text("""
            SELECT DISTINCT principal_name
            FROM principals
            WHERE used_company = :used_company
              AND principal_name IS NOT NULL
              AND principal_name <> ''
            ORDER BY principal_name ASC
        """), {'used_company': used_company}).mappings().all()
    return [row['principal_name'] for row in rows]


def validate_work_name_format(work_name):
    # 剧名只做两端去空格，中间空格保留（例如 "hello kitty" 原样保存）。
    # 这里不再校验格式，保留函数与调用点以便将来扩展其它规则。
    return None


WORK_ASSET_MAX_SIZE = 5 * 1024 * 1024


def validate_work_asset_filenames(work_name, proof_file=None, other_proof_files=None):
    normalized_work_name = normalize_filename_part(work_name)
    display_work_name = normalize_company_name(work_name)

    if proof_file and proof_file.filename:
        proof_stem = normalize_filename_part(Path(proof_file.filename).stem)
        expected_prefix = f'证明文件_{normalized_work_name}'
        if not proof_stem.startswith(expected_prefix):
            return f'作品权属文件名不符合要求，请上传以”证明文件_{display_work_name}”开头的文件'

    for file_storage in (other_proof_files or []):
        if not file_storage or not file_storage.filename:
            continue
        other_stem = normalize_filename_part(Path(file_storage.filename).stem)
        pattern = re.compile(rf'^其他证明_{re.escape(normalized_work_name)}_[0-9]+$')
        if not pattern.match(other_stem):
            return f'其他证明文件名不符合要求，请上传命名为”其他证明_{display_work_name}_序号.文件后缀”的文件'

    return None


def validate_work_asset_file_sizes(proof_file=None, other_proof_files=None):
    files_to_check = []
    if proof_file and proof_file.filename:
        files_to_check.append(('作品权属文件', proof_file))

    for idx, file_storage in enumerate(other_proof_files or [], start=1):
        if file_storage and file_storage.filename:
            files_to_check.append((f'其他证明文件#{idx}', file_storage))

    for label, file_storage in files_to_check:
        current_pos = file_storage.stream.tell()
        file_storage.stream.seek(0, os.SEEK_END)
        size = file_storage.stream.tell()
        file_storage.stream.seek(current_pos)
        if size > WORK_ASSET_MAX_SIZE:
            return f'{label}不能超过5MB'

    return None


def migrate_works_principal_name_if_needed():
    with get_db_session() as session:
        columns = {row[0] for row in session.execute(text("SHOW COLUMNS FROM works")).all()}
        schema_changed = False
        if 'principal_name' not in columns:
            session.execute(text("ALTER TABLE works ADD COLUMN principal_name varchar(255) NULL AFTER used_company"))
            schema_changed = True
        if 'operator_name' not in columns:
            session.execute(text("ALTER TABLE works ADD COLUMN operator_name varchar(255) NULL AFTER principal_name"))
            schema_changed = True
        if schema_changed:
            session.commit()


try:
    migrate_works_principal_name_if_needed()
except Exception as exc:
    print(f'works schema migration skipped: {exc}')


def get_work_content_types():
    with get_db_session() as session:
        rows = session.execute(text("""
            SELECT dict_code AS id, dict_name AS name
            FROM dictionaries WHERE dict_type = 'content_type'
            ORDER BY sort_order ASC, id ASC
        """)).mappings().all()
        return [dict(row) for row in rows]


def get_work_complaint_types():
    with get_db_session() as session:
        rows = session.execute(text("""
            SELECT dict_code AS id, dict_name AS name
            FROM dictionaries WHERE dict_type = 'complaint_type'
            ORDER BY sort_order ASC, id ASC
        """)).mappings().all()
        return [dict(row) for row in rows]


def normalize_work_path_part(value):
    return (value or '').strip().replace('/', '_').replace('\\', '_')


def save_work_asset_file(file_storage, target_dir, filename_prefix):
    if not file_storage or not file_storage.filename:
        return None, None
    original_name = Path(file_storage.filename).name
    suffix = Path(original_name).suffix.lower()
    random_suffix = uuid4().hex[:8]
    filename = f"{filename_prefix}_{random_suffix}{suffix}"
    save_path = os.path.join(target_dir, filename)
    file_storage.save(save_path)
    return filename, save_path


def create_work_with_assets(work_name, used_company, principal_name, content_type_id, complaint_type_id, proof_file, other_proof_files, operator_name=''):
    normalized_work_name = normalize_work_path_part(work_name)
    with get_db_session() as session:
        exists = session.execute(text("""
            SELECT 1 FROM works
            WHERE work_name = :work_name
              AND used_company = :used_company
              AND principal_name = :principal_name
              AND content_type_id = :content_type_id
              AND complaint_type_id = :complaint_type_id
            LIMIT 1
        """), {
            'work_name': work_name,
            'used_company': used_company,
            'principal_name': principal_name,
            'content_type_id': content_type_id,
            'complaint_type_id': complaint_type_id,
        }).first()
        if exists:
            return None, '该作品在当前代理主体、被代理人、所属类型和投诉类型下已存在'

        content_type = session.execute(text("SELECT dict_name AS name FROM dictionaries WHERE dict_type = 'content_type' AND dict_code = :id LIMIT 1"), {'id': str(content_type_id)}).mappings().first()
        complaint_type = session.execute(text("SELECT dict_name AS name FROM dictionaries WHERE dict_type = 'complaint_type' AND dict_code = :id LIMIT 1"), {'id': str(complaint_type_id)}).mappings().first()
        if not content_type or not complaint_type:
            return None, '内容类型或投诉类型无效'

        content_type_name = normalize_work_path_part(content_type['name'])
        complaint_type_name = normalize_work_path_part(complaint_type['name'])
        work_dir_name = f"{normalized_work_name}_{normalize_work_path_part(used_company)}_{content_type_name}_{complaint_type_name}"
        work_dir = os.path.join(os.path.dirname(__file__), 'static', 'imgs', '剧名', work_dir_name)
        ensure_dir(work_dir)

        proof_filename, proof_path = save_work_asset_file(proof_file, work_dir, f'证明文件_{normalized_work_name}')
        if not proof_filename:
            return None, '请上传作品权属文件'

        other_saved = []
        for idx, file_storage in enumerate(other_proof_files[:2], start=1):
            saved_name, saved_path = save_work_asset_file(file_storage, work_dir, f'其他证明_{normalized_work_name}_{idx}')
            if saved_name:
                other_saved.append((saved_name, saved_path))

        now = datetime.now()
        session.execute(text("""
            INSERT INTO works (
                work_name, used_company, principal_name, operator_name, content_type_id, complaint_type_id, created_at, updated_at
            ) VALUES (
                :work_name, :used_company, :principal_name, :operator_name, :content_type_id, :complaint_type_id, :created_at, :updated_at
            )
        """), {
            'work_name': work_name,
            'used_company': used_company,
            'principal_name': principal_name,
            'operator_name': operator_name,
            'content_type_id': content_type_id,
            'complaint_type_id': complaint_type_id,
            'created_at': now,
            'updated_at': now,
        })
        work_id = session.execute(text('SELECT LAST_INSERT_ID()')).scalar_one()

        # 更新证明文件路径到 works 表
        other_proof_list = [saved_name for saved_name, saved_path in other_saved]
        session.execute(text("""
            UPDATE works SET proof_file = :proof_file, other_proof_files = :other_proof_files
            WHERE id = :work_id
        """), {
            'proof_file': proof_filename,
            'other_proof_files': json.dumps(other_proof_list, ensure_ascii=False) if other_proof_list else None,
            'work_id': work_id,
        })
        session.commit()
        return {
            'work_id': work_id,
            'work_name': work_name,
            'used_company': used_company,
            'principal_name': principal_name,
            'operator_name': operator_name,
            'content_type': content_type['name'],
            'complaint_type': complaint_type['name'],
            'proof_file': proof_filename,
            'other_proof_count': len(other_saved),
        }, None


def serialize_complaint_numbers(value):
    if value is None:
        return json.dumps([])
    if isinstance(value, str):
        return value
    return json.dumps(value, ensure_ascii=False)


def deserialize_complaint_numbers(value):
    if not value:
        return []
    if isinstance(value, (list, tuple)):
        return list(value)
    try:
        return json.loads(value)
    except (TypeError, json.JSONDecodeError):
        return []


def map_task_status_label(status):
    if status == 'running':
        return '执行中'
    if status == 'completed':
        return '已完成'
    if status == 'failed':
        return '失败'
    if status == 'pending' or status == 'queued':
        return '等待中'
    if status == 'partial_failed':
        return '部分失败'
    return status or '未知'


def insert_complaint(complaint_id, task_id, platform_code, payload, rights_holder, operator='', upload_filename=''):
    submitted_at = datetime.fromisoformat(payload['submitted_at'])
    work_name = payload['form'].get('作品名称') or ''
    with get_db_session() as session:
        session.execute(text("""
            INSERT INTO complaints (
                complaint_id, task_id, platform_code, collect_account, cookie_snapshot,
                identity_type, agent_name, principal_name,
                complaint_category, complaint_type, module_name, content_type,
                description_text, work_name, total_links, batch_size, batch_count,
                status, submitted_at, operator, upload_filename
            ) VALUES (
                :complaint_id, :task_id, :platform_code, :collect_account, :cookie_snapshot,
                :identity_type, :agent_name, :principal_name,
                :complaint_category, :complaint_type, :module_name, :content_type,
                :description_text, :work_name, :total_links, :batch_size, :batch_count,
                :status, :submitted_at, :operator, :upload_filename
            )
        """), {
            'complaint_id': complaint_id,
            'task_id': task_id,
            'platform_code': platform_code,
            'collect_account': payload['form'].get('collect_account', ''),
            'cookie_snapshot': (payload['form'].get('cookie', '') or '')[:100] + '...',
            'identity_type': payload['form'].get('identity', ''),
            'agent_name': payload['form'].get('agent', ''),
            'principal_name': payload['form'].get('principal') or rights_holder or '',
            'complaint_category': payload['form'].get('complaint_category', ''),
            'complaint_type': payload['form'].get('complaint_type', ''),
            'module_name': payload['form'].get('module', ''),
            'content_type': payload['form'].get('content_type', ''),
            'description_text': payload['form'].get('description', ''),
            'work_name': work_name,
            'total_links': payload.get('excel_rows', 0),
            'batch_size': payload.get('batch_size', 200),
            'batch_count': payload.get('batch_count', 0),
            'status': 'queued',
            'submitted_at': submitted_at,
            'operator': operator,
            'upload_filename': upload_filename,
        })
        session.commit()


def insert_complaint_task(task_id, submission_id, submitted_at, batch_count, excel_rows):
    """兼容旧调用，内部不再使用"""
    pass


def insert_complaint_batches(complaint_id, batches):
    with get_db_session() as session:
        for batch in batches:
            session.execute(text("""
                INSERT INTO complaint_batches (
                    batch_id, complaint_id, batch_no, work_name, batch_filename,
                    start_row, end_row, row_count, status
                ) VALUES (
                    :batch_id, :complaint_id, :batch_no, :work_name, :batch_filename,
                    :start_row, :end_row, :row_count, 'pending'
                )
            """), {
                'batch_id': uuid4().hex[:12],
                'complaint_id': complaint_id,
                'batch_no': batch['batch_no'],
                'work_name': batch.get('work_name', ''),
                'batch_filename': batch.get('filename', ''),
                'start_row': batch.get('start_row', 0),
                'end_row': batch.get('end_row', 0),
                'row_count': batch.get('rows', 0),
            })
        session.commit()


def update_complaint_task(task_id, **fields):
    if not fields:
        return
    allowed = {
        'status', 'current_batch', 'batch_count', 'completed_batches', 'failed_batches',
        'complaint_numbers_json', 'error_message', 'worker_name', 'redis_job_id',
        'submitted_at', 'queued_at', 'started_at', 'completed_at'
    }
    updates = {k: v for k, v in fields.items() if k in allowed}
    if 'complaint_numbers_json' in updates:
        updates['complaint_numbers_json'] = serialize_complaint_numbers(updates['complaint_numbers_json'])
    if not updates:
        return
    updates['updated_at'] = datetime.now()
    set_clause = ', '.join(f"{key} = :{key}" for key in updates.keys())
    updates['task_id'] = task_id
    with get_db_session() as session:
        session.execute(text(f"UPDATE complaints SET {set_clause} WHERE task_id = :task_id"), updates)
        session.commit()


def update_complaint_batch(submission_id, batch_no, **fields):
    if not fields:
        return
    allowed = {'status', 'complaint_number', 'error_message'}
    updates = {k: v for k, v in fields.items() if k in allowed}
    if not updates:
        return
    updates['updated_at'] = datetime.now()
    updates['complaint_id'] = submission_id
    updates['batch_no'] = batch_no
    set_clause = ', '.join(f"{key} = :{key}" for key in updates.keys() if key not in {'complaint_id', 'batch_no'})
    with get_db_session() as session:
        session.execute(text(f"""
            UPDATE complaint_batches
            SET {set_clause}
            WHERE complaint_id = :complaint_id AND batch_no = :batch_no
        """), updates)
        session.commit()


def get_complaint_task(task_id):
    with get_db_session() as session:
        task = session.execute(text("""
            SELECT task_id, submission_id, status, current_batch, batch_count,
                   completed_batches, failed_batches, complaint_numbers_json,
                   error_message, submitted_at, queued_at, started_at, completed_at
            FROM complaints
            WHERE task_id = :task_id
            LIMIT 1
        """), {'task_id': task_id}).mappings().first()
        if not task:
            return None
        batches = session.execute(text("""
            SELECT batch_no, row_count, start_row, end_row, batch_filename,
                   status, complaint_number, error_message
            FROM complaint_batches
            WHERE complaint_id = :complaint_id
            ORDER BY batch_no ASC
        """), {'submission_id': task['submission_id']}).mappings().all()

    complaint_numbers = deserialize_complaint_numbers(task.get('complaint_numbers_json'))
    batch_items = []
    for batch in batches:
        batch_items.append({
            'batch_no': batch['batch_no'],
            'rows': batch['row_count'],
            'start_row': batch['start_row'],
            'end_row': batch['end_row'],
            'filename': batch.get('batch_filename'),
            'status': batch['status'],
            'error': batch.get('error_message'),
            'complaint_number': batch.get('complaint_number'),
        })
    complaint_number = complaint_numbers[0] if complaint_numbers else None
    log_file_path = os.path.join(app.config['TASK_RESULT_FOLDER'], f"{task_id}.log")
    return {
        'task_id': task['task_id'],
        'submission_id': task['submission_id'],
        'status': task['status'],
        'complaint_number': complaint_number,
        'complaint_numbers': complaint_numbers,
        'batch_count': task['batch_count'],
        'completed_batches': task['completed_batches'],
        'failed_batches': task['failed_batches'],
        'current_batch': task['current_batch'],
        'batches': batch_items,
        'error': task.get('error_message'),
        'submitted_at': normalize_datetime(task.get('submitted_at')),
        'started_at': normalize_datetime(task.get('started_at')),
        'completed_at': normalize_datetime(task.get('completed_at')),
        'log_file_path': log_file_path,
    }


def get_submission_status_list():
    with get_db_session() as session:
        rows = session.execute(text("""
            SELECT complaint_id, submitted_at, collect_account, work_name,
                   total_links, batch_count, status, complaint_numbers_json, operator
            FROM complaints
            WHERE platform_code = 'uc'
            ORDER BY submitted_at DESC
        """)).mappings().all()

    items = []
    for row in rows:
        task_id = f"uc_{row['complaint_id']}"
        items.append({
            'submission_id': row['complaint_id'],
            'submitted_at': normalize_datetime(row.get('submitted_at')),
            'collect_account': row.get('collect_account') or '',
            'work_name': row.get('work_name') or '',
            'excel_rows': row.get('total_links') or 0,
            'batch_count': row.get('batch_count') or 0,
            'status': map_task_status_label(row.get('status')),
            'complaint_numbers': deserialize_complaint_numbers(row.get('complaint_numbers_json')),
            'log_available': has_available_task_log(task_id),
            'operator': row.get('operator') or '',
        })
    return items


def migrate_submission_and_task_data_if_needed():
    pass


cleanup_old_task_logs()


def _schedule_daily_log_cleanup():
    import time as _time
    while True:
        _time.sleep(86400)
        try:
            cleanup_old_task_logs()
        except Exception:
            pass


threading.Thread(target=_schedule_daily_log_cleanup, daemon=True, name='daily-log-cleanup').start()


@app.route('/accounts')
@login_required
def accounts():
    return render_template('accounts.html')


@app.route('/principals')
@login_required
def principals():
    return render_template('principals.html')


@app.route('/api/platforms')
@login_required
def api_platforms():
    return jsonify({'success': True, 'data': get_platforms_list()})


@app.route('/api/accounts/list')
@login_required
def accounts_list():
    platform_code = request.args.get('platform_code')
    accounts = load_accounts()
    if platform_code:
        accounts = [a for a in accounts if a.get('platform_code') == platform_code]
    return jsonify({'success': True, 'data': accounts})


@app.route('/api/accounts/add', methods=['POST'])
@login_required
def accounts_add():
    data = request.get_json()
    platform_code = data.get('platform_code', '').strip()
    used_company = data.get('used_company', '').strip()
    user = data.get('user', '').strip()
    cookie = data.get('cookie', '').strip()
    if not platform_code or not used_company or not user or not cookie:
        return jsonify({'success': False, 'error': '使用的公司、平台名称、投诉账号、Cookie都不能为空'}), 400
    if used_company not in {'和晞科技', '柏蒙文化', '中惠信科'}:
        return jsonify({'success': False, 'error': '使用的公司无效'}), 400
    if platform_code not in PLATFORM_MAP:
        return jsonify({'success': False, 'error': '平台编码无效'}), 400

    with get_db_session() as session:
        exists = session.execute(text("""
            SELECT 1 FROM accounts
            WHERE platform_code = :platform_code AND account_user = :account_user
            LIMIT 1
        """), {'platform_code': platform_code, 'account_user': user}).first()
        if exists:
            return jsonify({'success': False, 'error': f'该平台下投诉账号「{user}」已存在'}), 400

        new_id = uuid4().hex[:12]
        now = datetime.now()
        session.execute(text("""
            INSERT INTO accounts (
                account_id, platform_code, platform_name, platform_label,
                used_company, account_user, cookie_text, account_purpose, status,
                created_at, updated_at
            ) VALUES (
                :account_id, :platform_code, :platform_name, :platform_label,
                :used_company, :account_user, :cookie_text, :account_purpose, :status,
                :created_at, :updated_at
            )
        """), {
            'account_id': new_id,
            'platform_code': platform_code,
            'platform_name': PLATFORM_MAP[platform_code]['platform_name'],
            'platform_label': PLATFORM_MAP[platform_code]['pingtai'],
            'used_company': used_company,
            'account_user': user,
            'cookie_text': cookie,
            'account_purpose': data.get('account_purpose', '').strip() or None,
            'status': 'active',
            'created_at': now,
            'updated_at': now,
        })
        session.commit()

        row = session.execute(text("""
            SELECT account_id, platform_code, platform_name, platform_label,
                   used_company, account_user, cookie_text, account_purpose, status,
                   created_at, updated_at
            FROM accounts
            WHERE account_id = :account_id
            LIMIT 1
        """), {'account_id': new_id}).mappings().one()

    return jsonify({'success': True, 'data': row_to_account_dict(row)})


@app.route('/api/accounts/update_cookie', methods=['POST'])
@login_required
def accounts_update_cookie():
    data = request.get_json()
    acc_id = data.get('id')
    cookie = data.get('cookie', '').strip()
    if not cookie:
        return jsonify({'success': False, 'error': 'Cookie不能为空'}), 400

    with get_db_session() as session:
        result = session.execute(text("""
            UPDATE accounts
            SET cookie_text = :cookie_text, updated_at = :updated_at
            WHERE account_id = :account_id
        """), {
            'cookie_text': cookie,
            'updated_at': datetime.now(),
            'account_id': acc_id,
        })
        session.commit()
        if result.rowcount == 0:
            return jsonify({'success': False, 'error': '账号不存在'}), 404

    return jsonify({'success': True})


@app.route('/api/principals/list')
@login_required
def principals_list():
    """返回所有账号及其被代理人信息，每行一个被代理人"""
    platform_code_filter = request.args.get('platform_code', '').strip()
    account_user_filter = request.args.get('account_user', '').strip()
    used_company_filter = request.args.get('used_company', '').strip()
    can_update_authorization = can_update_principal_authorization()
    principals_data = load_principals_map()
    accounts = load_accounts()
    if platform_code_filter:
        accounts = [acc for acc in accounts if acc.get('platform_code') == platform_code_filter]
    if account_user_filter:
        accounts = [acc for acc in accounts if acc.get('user') == account_user_filter]
    if used_company_filter:
        accounts = [acc for acc in accounts if acc.get('used_company') == used_company_filter]
    results = []
    for acc in accounts:
        key = f"{acc['platform_code']}:{acc['user']}"
        entry = principals_data.get(key, {})
        principals = entry.get('principals', [])
        count = len(principals) if principals else 1
        if principals:
            for i, name in enumerate(principals):
                doc_record = get_principal_document_record(acc['platform_code'], acc.get('used_company', ''), name)
                results.append({
                    'platform_code': acc['platform_code'],
                    'platform_name': acc.get('platform_name', ''),
                    'used_company': acc.get('used_company', ''),
                    'account_user': acc['user'],
                    'account_purpose': acc.get('account_purpose', ''),
                    'principal_name': name,
                    'business_license_filename': doc_record.get('business_license_filename') if doc_record else None,
                    'authorization_filename': doc_record.get('authorization_filename') if doc_record else None,
                    'authorization_expires_on': doc_record.get('authorization_expires_on') if doc_record else None,
                    'can_update': can_update_authorization,
                    'rowspan': count if i == 0 else 0,
                })
        else:
            results.append({
                'platform_code': acc['platform_code'],
                'platform_name': acc.get('platform_name', ''),
                'used_company': acc.get('used_company', ''),
                'account_user': acc['user'],
                'account_purpose': acc.get('account_purpose', ''),
                'principal_name': '-',
                'business_license_filename': None,
                'authorization_filename': None,
                'authorization_expires_on': None,
                'can_update': False,
                'rowspan': 1,
            })
    return jsonify({'success': True, 'data': results})


@app.route('/api/principals/authorization_alerts')
@login_required
def principal_authorization_alerts():
    return jsonify({'success': True, 'data': get_authorization_expiry_alerts()})


@app.route('/api/principals/document', methods=['GET'])
@login_required
def principal_document_detail():
    platform_code = request.args.get('platform_code', '').strip()
    used_company = request.args.get('used_company', '').strip()
    principal_name = request.args.get('principal_name', '').strip()
    if not platform_code or not used_company or not principal_name:
        return jsonify({'success': True, 'data': None})

    record = get_principal_document_record(platform_code, used_company, principal_name)
    if not record:
        return jsonify({'success': True, 'data': None})

    data = {
        'principal_name': record['principal_name'],
        'used_company': record['used_company'],
        'authorization_expires_on': record['authorization_expires_on'],
        'business_license_locked': record['business_license_locked'],
        'authorization_locked': record['authorization_locked'],
        'business_license_filename': record['business_license_filename'],
        'authorization_filename': record['authorization_filename'],
        'business_license_path': f"营业执照/{record['business_license_filename']}" if record['business_license_filename'] else None,
        'authorization_path': f"授权委托书/{record['authorization_filename']}" if record['authorization_filename'] else None,
    }
    return jsonify({'success': True, 'data': data})


@app.route('/api/principals/add', methods=['POST'])
@login_required
def principals_add():
    """添加被代理人信息，按 (platform_code + account_user) 分组"""
    if request.is_json:
        data = request.get_json()
        platform_code = data.get('platform_code', '').strip()
        account_user = data.get('account_user', '').strip()
        principal_name = data.get('principal_name', '').strip()
        used_company = ''
        business_license_file = None
        authorization_file = None
        authorization_expires_on = ''
    else:
        platform_code = request.form.get('platform_code', '').strip()
        account_user = request.form.get('account_user', '').strip()
        principal_name = request.form.get('principal_name', '').strip()
        used_company = request.form.get('used_company', '').strip()
        business_license_file = request.files.get('business_license_file')
        authorization_file = request.files.get('authorization_file')
        authorization_expires_on = request.form.get('authorization_expires_on', '').strip()

    if not platform_code or not account_user or not principal_name:
        return jsonify({'success': False, 'error': '平台名称、投诉账号、被代理人信息都不能为空'}), 400
    if platform_code not in PLATFORM_MAP:
        return jsonify({'success': False, 'error': '平台编码无效'}), 400

    normalized_principal_name = normalize_company_name(principal_name)
    normalized_used_company = normalize_company_name(used_company)

    with get_db_session() as session:
        account_exists = session.execute(text("""
            SELECT platform_name, used_company FROM accounts
            WHERE platform_code = :platform_code AND account_user = :account_user
            LIMIT 1
        """), {
            'platform_code': platform_code,
            'account_user': account_user,
        }).mappings().first()
        if not account_exists:
            return jsonify({'success': False, 'error': '投诉账号不存在'}), 400

        if not request.is_json:
            if not normalized_used_company:
                return jsonify({'success': False, 'error': '使用的公司不能为空'}), 400
            if account_exists.get('used_company') != normalized_used_company:
                return jsonify({'success': False, 'error': '所选投诉账号与使用的公司不匹配'}), 400

            existing_docs = get_principal_document_record(platform_code, normalized_used_company, normalized_principal_name)
            if not existing_docs or not existing_docs.get('business_license_locked'):
                if not business_license_file or not business_license_file.filename:
                    return jsonify({'success': False, 'error': '请上传被代理人营业执照'}), 400
            else:
                business_license_file = None

            if not existing_docs or not existing_docs.get('authorization_locked'):
                if not authorization_file or not authorization_file.filename:
                    return jsonify({'success': False, 'error': '请上传授权委托书'}), 400
                if not authorization_expires_on:
                    return jsonify({'success': False, 'error': '请填写授权期限截止日期'}), 400
            else:
                authorization_file = None

            size_error = validate_principal_upload_file_sizes(
                business_license_file=business_license_file,
                authorization_file=authorization_file,
            )
            if size_error:
                return jsonify({'success': False, 'error': size_error}), 400

            filename_error = validate_principal_upload_filenames(
                normalized_principal_name,
                normalized_used_company,
                authorization_expires_on,
                business_license_file=business_license_file,
                authorization_file=authorization_file,
            )
            if filename_error:
                return jsonify({'success': False, 'error': filename_error}), 400

        # 检查是否已存在
        exists = session.execute(text("""
            SELECT 1 FROM principals
            WHERE platform_code = :platform_code
              AND account_user = :account_user
              AND principal_name = :principal_name
            LIMIT 1
        """), {
            'platform_code': platform_code,
            'account_user': account_user,
            'principal_name': normalized_principal_name,
        }).first()
        if exists:
            return jsonify({'success': False, 'error': '该被代理人信息已存在'}), 400

        session.execute(text("""
            INSERT INTO principals (
                platform_code, used_company, account_user, principal_name
            ) VALUES (
                :platform_code, :used_company, :account_user, :principal_name
            )
        """), {
            'platform_code': platform_code,
            'used_company': normalized_used_company,
            'account_user': account_user,
            'principal_name': normalized_principal_name,
        })
        session.commit()

    if not request.is_json:
        business_license_dir = os.path.join(os.path.dirname(__file__), 'static', 'imgs', '营业执照')
        auth_dir = os.path.join(os.path.dirname(__file__), 'static', 'imgs', '授权委托书')
        ensure_dir(business_license_dir)
        ensure_dir(auth_dir)
        existing_docs = get_principal_document_record(platform_code, normalized_used_company, normalized_principal_name) or {}
        expires_yyyymmdd = authorization_expires_on.replace('-', '') if authorization_expires_on else None
        business_license_filename = existing_docs.get('business_license_filename')
        authorization_filename = existing_docs.get('authorization_filename')
        authorization_expires_value = existing_docs.get('authorization_expires_on')
        if business_license_file:
            business_license_filename = save_named_upload(business_license_file, business_license_dir, f'营业执照_{normalized_principal_name}')
        if authorization_file:
            authorization_filename = save_named_upload(authorization_file, auth_dir, f'授权委托书_{normalized_principal_name}_{normalized_used_company}_截止日期{expires_yyyymmdd}')
            authorization_expires_value = authorization_expires_on
        upsert_principal_documents(
            platform_code,
            normalized_used_company,
            account_user,
            normalized_principal_name,
            business_license_filename,
            authorization_filename,
            authorization_expires_value,
        )

    return jsonify({'success': True, 'data': {
        'platform_code': platform_code,
        'platform_name': PLATFORM_MAP[platform_code]['platform_name'],
        'used_company': account_exists.get('used_company', normalized_used_company),
        'account_user': account_user,
        'principal_name': normalized_principal_name,
    }})


@app.route('/api/principals/update', methods=['POST'])
@login_required
def principals_update():
    if not can_update_principal_authorization():
        return jsonify({'success': False, 'error': '当前账号无权限更新授权委托书'}), 403

    platform_code = request.form.get('platform_code', '').strip()
    account_user = request.form.get('account_user', '').strip()
    principal_name = request.form.get('principal_name', '').strip()
    used_company = request.form.get('used_company', '').strip()
    authorization_file = request.files.get('authorization_file')
    authorization_expires_on = request.form.get('authorization_expires_on', '').strip()

    if not platform_code or not account_user or not principal_name or not used_company:
        return jsonify({'success': False, 'error': '平台名称、投诉账号、代理主体(司内)、被代理人信息都不能为空'}), 400
    if platform_code not in PLATFORM_MAP:
        return jsonify({'success': False, 'error': '平台编码无效'}), 400
    if not authorization_file or not authorization_file.filename:
        return jsonify({'success': False, 'error': '请上传授权委托书'}), 400
    if not authorization_expires_on:
        return jsonify({'success': False, 'error': '请填写授权期限截止日期'}), 400

    normalized_principal_name = normalize_company_name(principal_name)
    normalized_used_company = normalize_company_name(used_company)
    if not normalized_principal_name or normalized_principal_name == '-':
        return jsonify({'success': False, 'error': '被代理人信息无效，无法更新授权委托书'}), 400

    with get_db_session() as session:
        account_exists = session.execute(text("""
            SELECT platform_name, used_company FROM accounts
            WHERE platform_code = :platform_code AND account_user = :account_user
            LIMIT 1
        """), {
            'platform_code': platform_code,
            'account_user': account_user,
        }).mappings().first()
        if not account_exists:
            return jsonify({'success': False, 'error': '投诉账号不存在'}), 400
        if account_exists.get('used_company') != normalized_used_company:
            return jsonify({'success': False, 'error': '所选投诉账号与使用的公司不匹配'}), 400

        principal_exists = session.execute(text("""
            SELECT 1 FROM principals
            WHERE platform_code = :platform_code
              AND account_user = :account_user
              AND principal_name = :principal_name
            LIMIT 1
        """), {
            'platform_code': platform_code,
            'account_user': account_user,
            'principal_name': normalized_principal_name,
        }).first()
        if not principal_exists:
            return jsonify({'success': False, 'error': '被代理人信息不存在'}), 404

    size_error = validate_principal_upload_file_sizes(
        authorization_file=authorization_file,
    )
    if size_error:
        return jsonify({'success': False, 'error': size_error}), 400

    filename_error = validate_principal_upload_filenames(
        normalized_principal_name,
        normalized_used_company,
        authorization_expires_on,
        authorization_file=authorization_file,
    )
    if filename_error:
        return jsonify({'success': False, 'error': filename_error}), 400

    business_license_dir = os.path.join(os.path.dirname(__file__), 'static', 'imgs', '营业执照')
    auth_dir = os.path.join(os.path.dirname(__file__), 'static', 'imgs', '授权委托书')
    ensure_dir(business_license_dir)
    ensure_dir(auth_dir)

    existing_docs = get_principal_document_record(platform_code, normalized_used_company, normalized_principal_name)
    if not existing_docs:
        return jsonify({'success': False, 'error': '被代理人资料不存在'}), 404

    expires_yyyymmdd = authorization_expires_on.replace('-', '')
    authorization_filename = save_named_upload(
        authorization_file,
        auth_dir,
        f'授权委托书_{normalized_principal_name}_{normalized_used_company}_截止日期{expires_yyyymmdd}'
    )
    upsert_principal_documents(
        platform_code,
        normalized_used_company,
        account_user,
        normalized_principal_name,
        existing_docs.get('business_license_filename'),
        authorization_filename,
        authorization_expires_on,
    )

    return jsonify({'success': True, 'data': {
        'platform_code': platform_code,
        'platform_name': PLATFORM_MAP[platform_code]['platform_name'],
        'used_company': normalized_used_company,
        'account_user': account_user,
        'principal_name': normalized_principal_name,
        'authorization_filename': authorization_filename,
        'authorization_expires_on': authorization_expires_on,
    }})


@app.route('/api/works/content_types')
@login_required
def works_content_types():
    return jsonify({'success': True, 'data': get_work_content_types()})


@app.route('/api/works/principal_options')
@login_required
def works_principal_options():
    used_company = request.args.get('used_company', '').strip()
    return jsonify({'success': True, 'data': get_principal_options_by_used_company(used_company)})


@app.route('/api/works/complaint_types')
@login_required
def works_complaint_types():
    return jsonify({'success': True, 'data': get_work_complaint_types()})


@app.route('/api/works/list')
@login_required
def works_list():
    can_update = can_update_principal_authorization()
    with get_db_session() as session:
        rows = session.execute(text("""
            SELECT w.id, w.work_name, w.used_company, w.principal_name, w.operator_name,
                   w.proof_file, w.other_proof_files,
                   ct.dict_name AS content_type, cpt.dict_name AS complaint_type
            FROM works w
            JOIN dictionaries ct ON ct.dict_type = 'content_type' AND ct.dict_code = CAST(w.content_type_id AS CHAR)
            JOIN dictionaries cpt ON cpt.dict_type = 'complaint_type' AND cpt.dict_code = CAST(w.complaint_type_id AS CHAR)
            ORDER BY w.updated_at DESC, w.id DESC
        """)).mappings().all()
        results = []
        for row in rows:
            proof_file = row.get('proof_file') or None
            other_files = []
            if row.get('other_proof_files'):
                try:
                    other_files = json.loads(row['other_proof_files']) if isinstance(row['other_proof_files'], str) else row['other_proof_files']
                except:
                    pass
            results.append({
                'id': row['id'],
                'work_name': row['work_name'],
                'used_company': row['used_company'],
                'principal_name': row.get('principal_name') or '',
                'operator_name': row.get('operator_name') or '',
                'content_type': row['content_type'],
                'complaint_type': row['complaint_type'],
                'proof_file': proof_file,
                'other_proof_files': other_files,
                'other_proof_count': len(other_files),
                'can_update': can_update,
            })
    return jsonify({'success': True, 'data': results})


@app.route('/api/works/add', methods=['POST'])
@login_required
def works_add():
    work_name = request.form.get('work_name', '').strip()
    used_company = request.form.get('used_company', '').strip()
    principal_name = request.form.get('principal_name', '').strip()
    content_type_id = request.form.get('content_type_id', '').strip()
    complaint_type_id = request.form.get('complaint_type_id', '').strip()
    proof_file = request.files.get('proof_file')
    other_files = [f for f in request.files.getlist('other_proof_file') if f and f.filename]

    if not work_name:
        return jsonify({'success': False, 'error': '剧名、代理主体(司内)、被代理人信息、内容类型、投诉类型都不能为空'}), 400

    work_name_error = validate_work_name_format(work_name)
    if work_name_error:
        return jsonify({'success': False, 'error': work_name_error}), 400

    if not used_company or not principal_name or not content_type_id or not complaint_type_id:
        return jsonify({'success': False, 'error': '剧名、代理主体(司内)、被代理人信息、内容类型、投诉类型都不能为空'}), 400

    if not proof_file or not proof_file.filename:
        return jsonify({'success': False, 'error': '请上传作品权属文件'}), 400
    if len(other_files) > 2:
        return jsonify({'success': False, 'error': '其他证明文件最多上传2个'}), 400

    size_error = validate_work_asset_file_sizes(proof_file=proof_file, other_proof_files=other_files)
    if size_error:
        return jsonify({'success': False, 'error': size_error}), 400

    filename_error = validate_work_asset_filenames(work_name, proof_file=proof_file, other_proof_files=other_files)
    if filename_error:
        return jsonify({'success': False, 'error': filename_error}), 400

    data, error = create_work_with_assets(
        work_name,
        used_company,
        principal_name,
        int(content_type_id),
        int(complaint_type_id),
        proof_file,
        other_files,
        operator_name=get_current_operator_name(),
    )
    if error:
        return jsonify({'success': False, 'error': error}), 400
    return jsonify({'success': True, 'data': data})


@app.route('/kuake')
@login_required
def kuake():
    return render_template('kuake.html')


@app.route('/uc')
@login_required
def uc():
    return render_template('uc.html')


@app.route('/api/check_excel', methods=['POST'])
@login_required
def check_excel():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '未上传文件'})
    f = request.files['file']
    if not f.filename:
        return jsonify({'success': False, 'error': '文件为空'})

    filename = secure_filename(f.filename)
    save_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    f.save(save_path)

    try:
        df = pd.read_excel(save_path)
        rows = len(df)
        if rows <= 0:
            return jsonify({'success': False, 'error': 'Excel 文件没有可提交的数据行'}), 400
        batch_size = 200
        batch_count = math.ceil(rows / batch_size)
        return jsonify({
            'success': True,
            'rows': rows,
            'filename': filename,
            'batch_size': batch_size,
            'batch_count': batch_count,
        })
    except Exception as e:
        return jsonify({'success': False, 'error': f'文件解析失败：{str(e)}'})
    finally:
        if os.path.exists(save_path):
            os.remove(save_path)


@app.route('/api/works/check_active/<work_name>')
@login_required
def works_check_active(work_name):
    with get_db_session() as session:
        row = session.execute(text("""
            SELECT task_id FROM complaints
            WHERE work_name LIKE :pattern AND status IN ('running', 'queued')
            LIMIT 1
        """), {'pattern': f'%{work_name}%'}).fetchone()
    return jsonify({'active': bool(row), 'task_id': row[0] if row else None})


@app.route('/api/works/update_proof', methods=['POST'])
@login_required
def works_update_proof():
    if not can_update_principal_authorization():
        return jsonify({'success': False, 'error': '当前账号无权限更新作品权属文件'}), 403

    work_id = request.form.get('work_id', '').strip()
    if not work_id:
        return jsonify({'success': False, 'error': 'work_id 不能为空'}), 400

    proof_file = request.files.get('proof_file')
    other_files = [f for f in request.files.getlist('other_proof_file') if f and f.filename]

    if not proof_file or not proof_file.filename:
        return jsonify({'success': False, 'error': '证明文件不能为空'}), 400

    with get_db_session() as session:
        row = session.execute(text("""
            SELECT w.id, w.work_name, w.used_company, ct.dict_name AS content_type_name, cpt.dict_name AS complaint_type_name
            FROM works w
            JOIN dictionaries ct ON ct.dict_type='content_type' AND ct.dict_code=CAST(w.content_type_id AS CHAR)
            JOIN dictionaries cpt ON cpt.dict_type='complaint_type' AND cpt.dict_code=CAST(w.complaint_type_id AS CHAR)
            WHERE w.id=:wid
        """), {'wid': work_id}).mappings().first()
        if not row:
            return jsonify({'success': False, 'error': '作品不存在'}), 404

        normalized_work_name = normalize_work_path_part(row['work_name'])
        work_dir_name = f"{normalized_work_name}_{normalize_work_path_part(row['used_company'])}_{normalize_work_path_part(row['content_type_name'])}_{normalize_work_path_part(row['complaint_type_name'])}"
        work_dir = os.path.join(os.path.dirname(__file__), 'static', 'imgs', '剧名', work_dir_name)
        ensure_dir(work_dir)

        proof_filename, _ = save_work_asset_file(proof_file, work_dir, f'证明文件_{normalized_work_name}')
        other_saved = []
        for idx, f in enumerate(other_files[:2], start=1):
            name, _ = save_work_asset_file(f, work_dir, f'其他证明_{normalized_work_name}_{idx}')
            if name:
                other_saved.append(name)

        session.execute(text("""
            UPDATE works SET proof_file=:pf, other_proof_files=:opf, operator_name=:op, updated_at=NOW()
            WHERE id=:wid
        """), {
            'pf': proof_filename,
            'opf': json.dumps(other_saved, ensure_ascii=False) if other_saved else None,
            'op': get_current_user(),
            'wid': work_id,
        })
        session.commit()

    return jsonify({'success': True, 'proof_file': proof_filename, 'other_proof_files': other_saved})


@app.route('/api/uc/submit', methods=['POST'])
@login_required
def submit_uc_form():
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': '请求数据为空'}), 400

    required_fields = {
        'collect_account': '采集账号',
        'cookie': 'Cookie',
        'identity': '您的身份',
        'agent': '代理人/权利人',
        'complaint_category': '投诉大类',
        'complaint_type': '投诉类型',
        'module': '功能模块',
        'content_type': '内容类型',
        'description': '投诉内容描述',
    }

    missing_fields = [label for key, label in required_fields.items() if not data.get(key, '').strip()]
    identity = data.get('identity', '').strip()
    if identity == '代理人' and not data.get('principal', '').strip():
        missing_fields.append('被代理人（权利人）信息')

    works_config = data.get('works', [])
    skipped_works = data.get('skipped_works', [])
    upload_filename = data.get('upload_filename', '').strip()
    if not works_config and not skipped_works:
        missing_fields.append('作品列表')

    if missing_fields:
        return jsonify({'success': False, 'error': '缺少必填项：' + '、'.join(missing_fields)}), 400

    # 授权拦截：所选被代理人授权已过期 / 缺少截止日期时，禁止投诉
    auth_block_error = check_principal_authorization_blocked(
        data.get('principal', '').strip(),
        platform_code='uc',
        account_user=data.get('collect_account', '').strip(),
    )
    if auth_block_error:
        return jsonify({'success': False, 'error': auth_block_error}), 400

    submission_id, submission_dir = create_submission_dir()

    try:
        static_imgs_dir = os.path.join(os.path.dirname(__file__), 'static', 'imgs')

        # 按作品生成 Excel 分片
        all_batches = []
        works_payload = []
        batch_no_global = 0

        for work_idx, work in enumerate(works_config):
            work_name = work['work_name']
            work_links = work.get('excel_rows', [])
            proof_file_rel = work.get('proof_file', '')
            other_proof_files_rel = work.get('other_proof_files', [])

            # 将链接转为 DataFrame 并按 200 条分片
            df_rows = []
            for link in work_links:
                df_rows.append({
                    '侵权链接': link.get('侵权链接', ''),
                    '对应原创链接/对应访问码': link.get('对应原创链接/对应访问码', ''),
                    '作品名称': work_name,
                })
            df = pd.DataFrame(df_rows)

            batch_dir = os.path.join(submission_dir, 'batches')
            os.makedirs(batch_dir, exist_ok=True)

            link_count = len(df)
            work_batch_files = []
            work_batch_metadata = []

            for chunk_start in range(0, link_count, 200):
                batch_no_global += 1
                chunk_end = min(chunk_start + 200, link_count)
                chunk_df = df.iloc[chunk_start:chunk_end]

                batch_filename = f'part_{batch_no_global:03d}.xlsx'
                batch_path = os.path.join(batch_dir, batch_filename)
                chunk_df.to_excel(batch_path, index=False)

                work_batch_files.append(batch_path)
                batch_meta = {
                    'batch_no': batch_no_global,
                    'work_name': work_name,
                    'filename': batch_filename,
                    'start_row': chunk_start + 1,
                    'end_row': chunk_end,
                    'rows': chunk_end - chunk_start,
                }
                work_batch_metadata.append(batch_meta)
                all_batches.append(batch_meta)

            # 解析证明文件绝对路径
            proof_file_abs = os.path.join(static_imgs_dir, proof_file_rel) if proof_file_rel else ''
            other_proof_abs = [os.path.join(static_imgs_dir, p) for p in other_proof_files_rel if p]

            works_payload.append({
                'work_name': work_name,
                'excel_files': work_batch_files,
                'proof_file': proof_file_abs,
                'other_proof_files': other_proof_abs,
                'batch_count': len(work_batch_files),
                'link_count': link_count,
            })

        total_links = sum(w['link_count'] for w in works_payload)
        total_batches = batch_no_global
        all_work_names = [w['work_name'] for w in works_payload] + [w['work_name'] for w in skipped_works]
        work_names_str = ', '.join(all_work_names)
        rights_holder = data.get('principal', '').strip() if identity == '代理人' else data.get('agent', '').strip()

        # 保存 submission.json
        payload = {
            'submission_id': submission_id,
            'submitted_at': datetime.now().isoformat(),
            'form': {
                'collect_account': data.get('collect_account', '').strip(),
                'cookie': data.get('cookie', '').strip(),
                'identity': identity,
                'agent': data.get('agent', '').strip(),
                'principal': data.get('principal', '').strip(),
                'complaint_category': data.get('complaint_category', '').strip(),
                'complaint_type': data.get('complaint_type', '').strip(),
                'module': data.get('module', '').strip(),
                'content_type': data.get('content_type', '').strip(),
                'description': data.get('description', '').strip(),
                '作品名称': work_names_str,
            },
            'works_config': works_payload,
            'excel_rows': total_links,
            'batch_size': 200,
            'batch_count': total_batches,
            'batches': all_batches,
        }

        metadata_path = os.path.join(submission_dir, 'submission.json')
        with open(metadata_path, 'w', encoding='utf-8') as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)

        task_id = f"uc_{submission_id}"
        task_state = {
            'status': 'pending',
            'submission_id': submission_id,
            'submitted_at': payload['submitted_at'],
            'excel_rows': total_links,
            'batch_count': total_batches,
            'completed_batches': 0,
            'failed_batches': 0,
            'current_batch': 0,
            'complaint_numbers': [],
        }
        tasks[task_id] = task_state

        # 防重复：同账号+同文件名已有非失败记录则拒绝
        if upload_filename:
            _sess = get_db_session()
            try:
                dup = _sess.execute(text("""
                    SELECT task_id FROM complaints
                    WHERE collect_account=:acc AND upload_filename=:fn
                      AND platform_code='uc' AND status NOT IN ('failed')
                    LIMIT 1
                """), {'acc': data.get('collect_account', ''), 'fn': upload_filename}).fetchone()
            finally:
                _sess.close()
            if dup:
                return jsonify({'success': False, 'error': f'文件「{upload_filename}」已投诉过（任务 {dup[0]}），请勿重复提交'}), 400

        insert_complaint(submission_id, task_id, 'uc', payload, rights_holder, operator=get_current_user(), upload_filename=upload_filename)
        insert_complaint_task(task_id, submission_id, payload['submitted_at'], total_batches, total_links)
        insert_complaint_batches(submission_id, all_batches)

        complaint_category = data.get('complaint_category', '').strip()
        complaint_type = data.get('complaint_type', '').strip()
        copyright_type = complaint_type if complaint_category == '知识产权' else ''

        task_payload = {
            'task_id': task_id,
            'cookie': data.get('cookie', '').strip(),
            'identity': identity,
            'agent': data.get('agent', '').strip(),
            'rights_holder': rights_holder,
            'complaint_category': complaint_category,
            'copyright_type': copyright_type,
            'module': data.get('module', '').strip(),
            'content_type': data.get('content_type', '').strip(),
            'description': data.get('description', '').strip(),
            'works_config': works_payload,
            'total_batches': total_batches,
            'skipped_works': skipped_works,
            'operator': get_current_user(),
        }

        if works_payload:
            enqueue_uc_task(task_payload)
            update_complaint_task(task_id, status='queued')
            tasks[task_id]['status'] = 'queued'
        else:
            # 没有可投诉的作品，直接标记完成
            skipped_numbers = [f"{sw['work_name']}：{sw.get('reason', '作品覆盖列表中未匹配到或证明文件不齐全')}" for sw in skipped_works]
            update_complaint_task(task_id, status='completed',
                                 complaint_numbers_json=skipped_numbers,
                                 completed_at=datetime.now())
            tasks[task_id]['status'] = 'completed'

        return jsonify({
            'success': True,
            'task_id': task_id,
            'message': '任务已创建，正在排队执行投诉',
            'total_works': len(works_payload),
            'skipped_works': len(skipped_works),
            'total_links': total_links,
            'total_batches': total_batches,
        })
    except Exception as e:
        return jsonify({'success': False, 'error': f'提交失败：{str(e)}'}), 500


# 自定义模板上传后的临时文件目录
CUSTOM_TEMPLATE_FOLDER = os.path.join(os.path.dirname(__file__), 'uploads', 'custom_templates')


def cleanup_old_template_files(max_age_hours=24):
    """清理超过指定时间的临时模板目录"""
    import time
    if not os.path.exists(CUSTOM_TEMPLATE_FOLDER):
        return
    now = time.time()
    for item in os.listdir(CUSTOM_TEMPLATE_FOLDER):
        item_path = os.path.join(CUSTOM_TEMPLATE_FOLDER, item)
        if os.path.isdir(item_path):
            # 检查目录修改时间
            mtime = os.path.getmtime(item_path)
            age_hours = (now - mtime) / 3600
            if age_hours > max_age_hours:
                shutil.rmtree(item_path, ignore_errors=True)


def extract_zip_with_correct_encoding(zip_file_storage, extract_dir):
    """使用unzip命令解压ZIP文件以保留正确的中文文件名"""
    import tempfile
    import subprocess

    # 先保存上传的ZIP到临时文件
    with tempfile.NamedTemporaryFile(suffix='.zip', delete=False) as tmp_zip:
        zip_path = tmp_zip.name
        zip_file_storage.save(zip_path)

    # 使用unzip命令解压
    try:
        subprocess.run(['unzip', '-o', '-q', zip_path, '-d', extract_dir],
                      check=True, capture_output=True)
    finally:
        os.unlink(zip_path)  # 删除临时ZIP文件


@app.route('/api/download_custom_template', methods=['GET'])
@login_required
def download_custom_template():
    """下载自定义模板Excel（3个Sheet）"""
    try:
        # Sheet1: 表单内容（除采集账号和Cookie外，删除委托代理文件）
        sheet1_data = {
            '字段': [
                '您的身份', '代理人/权利人', '被代理人（权利人）信息', '投诉大类',
                '投诉类型', '功能模块', '内容类型', '投诉内容描述'
            ],
            '值': [
                '代理人', '北京和晞科技有限公司', '', '',
                '', '', '', ''
            ],
            '可选值(备注)': [
                '权利人、代理人', '北京和晞科技有限公司',
                '', '知识产权、人身权',
                '著作权（含视频、图文、图集等）、商标、专利、其他知识产权',
                '头条内容、大鱼号账号、UC网盘、神马搜索',
                '影视剧集、其他视频、小说、漫画、图片、文章、软件/游戏、其他',
                '可在描述中写 ${work_title}，投诉每部作品时会自动换成该作品名。例如写"链接涉及作品${work_title}侵权"，投诉《仙逆》时实际填写的是"链接涉及作品仙逆侵权"'
            ]
        }
        df_sheet1 = pd.DataFrame(sheet1_data)

        # Sheet2: 批量导入的Excel表格
        sheet2_headers = ['侵权链接', '对应原创链接/对应访问码', '作品名称']
        sheet2_data = [
            ['', '', ''],
            ['', '', ''],
            ['', '', ''],
        ]
        df_sheet2 = pd.DataFrame(sheet2_data, columns=sheet2_headers)

        # Sheet3: 填写要求说明
        sheet3_lines = [
            ['Sheet1 表单填写说明'],
            [''],
            ['字段名', '填写说明'],
            ['您的身份', '必填，选择「权利人」或「代理人」'],
            ['代理人/权利人', '必填，选择代理人或权利人名称'],
            ['被代理人（权利人）信息', '代理人身份时必填，选择被代理人，名称必须与投诉平台选项中的值保持一致'],
            ['投诉大类', '必填，选择「知识产权」或「人身权」'],
            ['投诉类型', '必填，根据投诉大类选择具体类型'],
            ['功能模块', '必填，选择功能模块'],
            ['内容类型', '必填，选择内容类型'],
            ['投诉内容描述', '必填，根据公司要求填写'],
            [''],
            ['Sheet2 批量导入Excel填写说明'],
            [''],
            ['字段名', '填写说明'],
            ['侵权链接', '必填，填写需要投诉的侵权内容链接'],
            ['对应原创链接/对应访问码', '必填，填写原创内容链接或访问码'],
            ['作品名称', '必填，填写原创作品名称，支持多个作品混合填写，系统按作品名称自动分组'],
            [''],
            ['证明文件说明'],
            [''],
            ['上传模版后，系统会根据Sheet2中的作品名称逐个匹配证明文件：'],
            [''],
            ['证明文件', '根据「作品名称」在 static/imgs/剧名/ 下查找以”剧名_所属公司_内容类型_投诉类型”命名的目录，并在该目录内匹配「证明文件_*」文件'],
            ['其他证明[1]', '根据「被代理人」和「使用的公司」在 static/imgs/授权委托书/ 目录下查找「授权委托书_被代理人_使用的公司_截止日期YYYYMMDD」文件'],
            ['其他证明[2]', '根据「被代理人」在 static/imgs/营业执照/ 目录下查找「营业执照_被代理人」文件'],
            ['其他证明[3]', '根据「代理人」在 static/imgs/营业执照/ 目录下查找「营业执照_代理人」文件'],
            [''],
            ['注意事项'],
            [''],
            ['1. 上传自定义模板时，只需上传Excel文件（.xlsx或.xls）'],
            ['2. Sheet2支持填写多个作品的链接，系统按作品名称自动分组，每个作品独立投诉'],
            ['3. 每个作品的链接超过200条时，系统自动按200条分片'],
            ['4. 作品名称必须与 static/imgs/剧名/ 下的目录名中的剧名部分一致'],
            ['5. 文件格式支持：jpg、png、jpeg、bmp、pdf'],
        ]
        df_sheet3 = pd.DataFrame(sheet3_lines)

        # 创建Excel文件
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_sheet1.to_excel(writer, sheet_name='表单内容', index=False)
            df_sheet2.to_excel(writer, sheet_name='批量导入Excel', index=False)
            df_sheet3.to_excel(writer, sheet_name='填写说明', index=False, header=False)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_sheet1.to_excel(writer, sheet_name='表单内容', index=False)
            df_sheet2.to_excel(writer, sheet_name='批量导入Excel', index=False)
            df_sheet3.to_excel(writer, sheet_name='填写说明', index=False, header=False)

        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='custom_template.xlsx'
        )
    except Exception as e:
        return jsonify({'success': False, 'error': f'生成模板失败：{str(e)}'}), 500


@app.route('/api/upload_custom_template', methods=['POST'])
@login_required
def upload_custom_template():
    """上传自定义模板Excel，自动匹配证明文件"""
    import glob

    selected_current_principal = request.form.get('current_principal', '').strip()
    collect_account = request.form.get('collect_account', '').strip()

    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '未上传文件'}), 400

    excel_file = request.files['file']
    if not excel_file.filename:
        return jsonify({'success': False, 'error': '文件为空'}), 400

    filename = excel_file.filename.lower()
    if not (filename.endswith('.xlsx') or filename.endswith('.xls')):
        return jsonify({'success': False, 'error': '请上传Excel格式文件（.xlsx或.xls）'}), 400

    try:
        # 保存上传的Excel到临时位置
        ensure_dir(CUSTOM_TEMPLATE_FOLDER)
        template_id = datetime.now().strftime('%Y%m%d_%H%M%S_') + uuid4().hex[:8]
        template_dir = os.path.join(CUSTOM_TEMPLATE_FOLDER, template_id)
        ensure_dir(template_dir)

        excel_path = os.path.join(template_dir, 'template.xlsx')
        excel_file.save(excel_path)

        # 读取Excel
        try:
            xls = pd.ExcelFile(excel_path)
            sheet1_data = pd.read_excel(xls, sheet_name='表单内容')
            sheet2_data = pd.read_excel(xls, sheet_name='批量导入Excel')
        except Exception as e:
            shutil.rmtree(template_dir, ignore_errors=True)
            return jsonify({'success': False, 'error': f'Excel解析失败：{str(e)}'}), 400

        # 检测Sheet2是否有空行（通过openpyxl读取真实行数，与pandas读取结果对比）
        try:
            wb_check = load_workbook(excel_path)
            ws_sheet2 = wb_check['批量导入Excel']
            actual_data_rows = ws_sheet2.max_row - 1  # 减去标题行
            pandas_data_rows = len(sheet2_data)
            if actual_data_rows > pandas_data_rows:
                empty_count = actual_data_rows - pandas_data_rows
                shutil.rmtree(template_dir, ignore_errors=True)
                return jsonify({'success': False, 'error': f'Sheet2中存在 {empty_count} 行空行，请删除空行后再上传'}), 400
        except Exception:
            pass  # 忽略openpyxl读取错误，以pandas结果为准

        # 辅助函数：标准化括号（全角转半角）
        def normalize_paren(s):
            return s.replace('（', '(').replace('）', ')')

        def normalize_principal_value(s):
            return normalize_paren((s or '').strip())

        def parse_work_folder_name(folder_name):
            parts = folder_name.split('_')
            if len(parts) < 4:
                return None
            work_name = '_'.join(parts[:-3]).strip()
            used_company = parts[-3].strip()
            content_type = parts[-2].strip()
            complaint_type = parts[-1].strip()
            return {
                'folder_name': folder_name,
                'work_name': work_name,
                'used_company': used_company,
                'content_type': content_type,
                'complaint_type': complaint_type,
            }

        def map_agent_to_used_company(agent_name):
            mapping = {
                '北京和晞科技有限公司': '和晞科技',
                '北京柏蒙文化传媒有限公司': '柏蒙文化',
                '北京中惠信科科技有限公司': '中惠信科',
            }
            return mapping.get((agent_name or '').strip(), '')

        def normalize_template_complaint_type(value):
            mapping = {
                '著作权（含视频、图文、图集等）': '著作权',
            }
            normalized = (value or '').strip()
            return mapping.get(normalized, normalized)

        # 辅助函数：检查公司名是否匹配（支持括号中英文模糊匹配）
        def company_match(principal, filename):
            # 标准化后精确匹配
            return normalize_paren(principal) in normalize_paren(filename)

        # 解析Sheet1表单数据
        form_data = {}
        try:
            for _, row in sheet1_data.iterrows():
                field = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''
                value = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ''
                if field and field != 'nan':
                    form_data[field] = value
        except Exception as e:
            shutil.rmtree(template_dir, ignore_errors=True)
            return jsonify({'success': False, 'error': f'Sheet1解析失败：{str(e)}'}), 400

        # Sheet1 必填字段校验（去掉了作品名称）
        required_fields_sheet1 = {
            '您的身份': '您的身份',
            '代理人/权利人': '代理人/权利人',
            '被代理人（权利人）信息': '被代理人（权利人）信息',
            '投诉大类': '投诉大类',
            '投诉类型': '投诉类型',
            '功能模块': '功能模块',
            '内容类型': '内容类型',
            '投诉内容描述': '投诉内容描述',
        }
        missing_fields = [label for field, label in required_fields_sheet1.items() if not form_data.get(field, '').strip()]
        if missing_fields:
            shutil.rmtree(template_dir, ignore_errors=True)
            return jsonify({'success': False, 'error': '以下必填项未填写：' + '、'.join(missing_fields)}), 400

        # 解析Sheet2批量导入数据
        excel_rows = []
        try:
            for i, row in sheet2_data.iterrows():
                # 检测空行：侵权链接列为空或仅空白字符
                if not (pd.notna(row.iloc[0]) and str(row.iloc[0]).strip()):
                    continue  # 空行已被前面的 openpyxl 检测捕获，这里静默跳过
                excel_rows.append({
                    '侵权链接': str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else '',
                    '对应原创链接/对应访问码': str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else '',
                    '作品名称': str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ''
                })
        except Exception as e:
            shutil.rmtree(template_dir, ignore_errors=True)
            return jsonify({'success': False, 'error': f'Sheet2解析失败：{str(e)}'}), 400

        if not excel_rows:
            shutil.rmtree(template_dir, ignore_errors=True)
            return jsonify({'success': False, 'error': 'Sheet2为空，请填写后重新上传'}), 400

        # 校验Sheet2 A列（侵权链接）格式：每行必须以 http 或 https 开头
        invalid_rows = []
        for i, row in enumerate(excel_rows, start=2):  # start=2 因为第1行是标题
            link = row.get('侵权链接', '')
            if link and not link.startswith('http://') and not link.startswith('https://'):
                invalid_rows.append(f'第{i}行')
        if invalid_rows:
            shutil.rmtree(template_dir, ignore_errors=True)
            return jsonify({'success': False, 'error': 'Sheet2 A列（侵权链接）格式错误：' + '、'.join(invalid_rows) + '，必须以 http:// 或 https:// 开头'}), 400

        # 校验链接地址是否重复（全局判重，只要链接地址相同即视为重复，与作品名称无关）
        uc_link_positions = {}
        for i, row in enumerate(excel_rows, start=2):
            key = row.get('侵权链接', '')
            if key:
                if key not in uc_link_positions:
                    uc_link_positions[key] = []
                uc_link_positions[key].append(i)
        uc_duplicate_errors = []
        for url, rows in uc_link_positions.items():
            if len(rows) > 1:
                uc_duplicate_errors.append(f'第{rows[0]}行与第{rows[1]}行链接地址重复')
        if uc_duplicate_errors:
            shutil.rmtree(template_dir, ignore_errors=True)
            return jsonify({'success': False, 'error': '侵权链接中存在重复，请删除后重新上传：\n' + '\n'.join(uc_duplicate_errors[:5])}), 400

        # 获取基本信息
        principal = form_data.get('被代理人（权利人）信息', '')  # 如 "北京uc"
        agent = form_data.get('代理人/权利人', '')  # 如 "北京和晞科技有限公司"

        normalized_template_principal = normalize_principal_value(principal)
        normalized_selected_current_principal = normalize_principal_value(selected_current_principal)
        if not normalized_template_principal:
            shutil.rmtree(template_dir, ignore_errors=True)
            return jsonify({
                'success': False,
                'error': '自定义模板中的“被代理人（权利人）信息”不能为空'
            }), 400

        if normalized_selected_current_principal and normalized_template_principal != normalized_selected_current_principal:
            shutil.rmtree(template_dir, ignore_errors=True)
            return jsonify({
                'success': False,
                'error': f'自定义模板中的“被代理人（权利人）信息”与当前页面选择的“本次投诉使用的被代理人信息”不一致。页面选择：{selected_current_principal}；模板内容：{principal}'
            }), 400

        principal = normalized_template_principal

        # 定义静态文件目录
        static_imgs_dir = os.path.join(os.path.dirname(__file__), 'static', 'imgs')

        # 优先从模板中的代理人/权利人反查所属公司，匹配不到时再回退到当前账号
        used_company = map_agent_to_used_company(form_data.get('代理人/权利人', ''))
        if not used_company and collect_account:
            with get_db_session() as session:
                account_row = session.execute(text("""
                    SELECT used_company FROM accounts
                    WHERE platform_code = 'uc' AND account_user = :account_user
                    LIMIT 1
                """), {'account_user': collect_account}).mappings().first()
                if account_row:
                    used_company = account_row.get('used_company', '').strip()

        # 1. 从 Sheet2 提取所有不重复的作品名称，按出现顺序排列
        work_names_ordered = []
        for row in excel_rows:
            wn = row.get('作品名称', '').strip()
            if wn and wn not in work_names_ordered:
                work_names_ordered.append(wn)

        if not work_names_ordered:
            shutil.rmtree(template_dir, ignore_errors=True)
            return jsonify({'success': False, 'error': 'Sheet2中没有有效的作品名称'}), 400

        # 按作品名称分组链接
        links_by_work = {}
        for row in excel_rows:
            wn = row.get('作品名称', '').strip()
            if wn:
                if wn not in links_by_work:
                    links_by_work[wn] = []
                links_by_work[wn].append(row)

        content_type_name = form_data.get('内容类型', '').strip()
        complaint_type_name = normalize_template_complaint_type(form_data.get('投诉类型', '').strip())

        works_base_dir = os.path.join(static_imgs_dir, '剧名')
        if not os.path.isdir(works_base_dir):
            shutil.rmtree(template_dir, ignore_errors=True)
            return jsonify({'success': False, 'error': '作品资料库目录不存在，请先建立作品资料'}), 400

        # 2. 逐个作品匹配证明文件
        works_config = []
        match_errors = []

        for work_name in work_names_ordered:
            work_links = links_by_work.get(work_name, [])
            link_count = len(work_links)
            batch_count = math.ceil(link_count / 200)

            # 查询 works 表
            with get_db_session() as session:
                work_rows = session.execute(text("""
                    SELECT w.id, w.work_name, w.used_company, w.principal_name, ct.dict_name AS content_type, cpt.dict_name AS complaint_type
                    FROM works w
                    JOIN dictionaries ct ON ct.dict_type = 'content_type' AND ct.dict_code = CAST(w.content_type_id AS CHAR)
                    JOIN dictionaries cpt ON cpt.dict_type = 'complaint_type' AND cpt.dict_code = CAST(w.complaint_type_id AS CHAR)
                    WHERE w.work_name = :work_name
                """), {'work_name': work_name}).mappings().all()

            if not work_rows:
                match_errors.append(f'「{work_name}」在作品覆盖列表中不存在')
                continue

            principal_matched_rows = [row for row in work_rows if normalize_company_name(row.get('principal_name') or '') == principal]
            if not principal_matched_rows:
                match_errors.append(f'「{work_name}」的被代理人信息与模板不一致')
                continue

            candidate_rows = [row for row in principal_matched_rows if row.get('used_company') == used_company]
            if not candidate_rows:
                match_errors.append(f'「{work_name}」在当前代理主体下没有匹配到')
                continue

            narrowed_rows = [row for row in candidate_rows if row.get('content_type') == content_type_name]
            if not narrowed_rows:
                match_errors.append(f'「{work_name}」内容类型不匹配')
                continue

            if len(narrowed_rows) > 1:
                narrowed_rows = [row for row in narrowed_rows if row.get('complaint_type') == complaint_type_name]
            if not narrowed_rows:
                match_errors.append(f'「{work_name}」投诉类型不匹配')
                continue

            matched_row = narrowed_rows[0]
            work_dir_name = f"{normalize_work_path_part(matched_row['work_name'])}_{normalize_work_path_part(matched_row['used_company'])}_{normalize_work_path_part(matched_row['content_type'])}_{normalize_work_path_part(matched_row['complaint_type'])}"
            drama_dir = os.path.join(works_base_dir, work_dir_name)
            if not os.path.isdir(drama_dir):
                match_errors.append(f'「{work_name}」作品目录不存在')
                continue

            work_rel_dir = os.path.join('剧名', work_dir_name)
            proof_file = None
            work_other_proofs = []

            for f in os.listdir(drama_dir):
                if f.startswith('证明文件_') and not f.startswith('._'):
                    proof_file = os.path.join(work_rel_dir, f)
                    break

            for f in os.listdir(drama_dir):
                if f.startswith('其他证明_') and not f.startswith('._'):
                    work_other_proofs.append(os.path.join(work_rel_dir, f))

            # 授权委托书
            proxy_file = None
            if principal:
                auth_dir = os.path.join(static_imgs_dir, '授权委托书')
                if os.path.isdir(auth_dir):
                    for f in os.listdir(auth_dir):
                        if f.startswith('授权委托书_') and not f.startswith('._'):
                            if company_match(principal, f):
                                proxy_file = os.path.join('授权委托书', f)
                                break

            # 营业执照(被代理人)
            biz_license_principal = None
            if principal:
                biz_dir = os.path.join(static_imgs_dir, '营业执照')
                if os.path.isdir(biz_dir):
                    for f in os.listdir(biz_dir):
                        if f.startswith('营业执照_') and not f.startswith('._'):
                            if company_match(principal, f):
                                biz_license_principal = os.path.join('营业执照', f)
                                break

            # 营业执照(代理人)
            biz_license_agent = None
            if agent:
                biz_dir = os.path.join(static_imgs_dir, '营业执照')
                if os.path.isdir(biz_dir):
                    for f in os.listdir(biz_dir):
                        if f.startswith('营业执照_') and not f.startswith('._'):
                            if company_match(agent, f):
                                biz_license_agent = os.path.join('营业执照', f)
                                break

            if proxy_file:
                work_other_proofs.append(proxy_file)
            if biz_license_principal:
                work_other_proofs.append(biz_license_principal)
            if biz_license_agent:
                work_other_proofs.append(biz_license_agent)

            # 校验必须匹配到的证明文件
            missing_proofs = []
            if not proof_file:
                missing_proofs.append('证明文件')
            if not proxy_file:
                missing_proofs.append('授权委托书')
            if not biz_license_principal:
                missing_proofs.append('营业执照(被代理人)')
            if not biz_license_agent:
                missing_proofs.append('营业执照(代理人)')

            if missing_proofs:
                match_errors.append(f'「{work_name}」缺少：{", ".join(missing_proofs)}')
                continue

            works_config.append({
                'work_name': work_name,
                'link_count': link_count,
                'batch_count': batch_count,
                'proof_file': proof_file,
                'other_proof_files': work_other_proofs,
                'excel_rows': work_links,
            })

        if match_errors and not works_config:
            shutil.rmtree(template_dir, ignore_errors=True)
            return jsonify({'success': False, 'error': '所有作品匹配失败：\n' + '\n'.join(match_errors)}), 400

        total_links = sum(w['link_count'] for w in works_config)
        total_batches = sum(w['batch_count'] for w in works_config)

        # 准备返回数据
        result = {
            'success': True,
            'template_id': template_id,
            'upload_filename': Path(excel_file.filename).name,
            'form_data': form_data,
            'works': [{
                'work_name': w['work_name'],
                'link_count': w['link_count'],
                'batch_count': w['batch_count'],
                'proof_file': w['proof_file'],
                'other_proof_files': w['other_proof_files'],
            } for w in works_config],
            'excel_rows': excel_rows,
            'total_works': len(works_config),
            'total_links': total_links,
            'total_batches': total_batches,
            'match_errors': match_errors,
        }

        return jsonify(result)

    except Exception as e:
        if 'template_dir' in dir() and template_dir:
            shutil.rmtree(template_dir, ignore_errors=True)
        return jsonify({'success': False, 'error': f'处理失败：{str(e)}'}), 500


@app.route('/api/proof_file/<path:filename>', methods=['GET'])
@login_required
def serve_proof_file(filename):
    """服务证明文件（从static/imgs目录）"""
    # 安全检查：防止路径遍历
    static_dir = os.path.join(os.path.dirname(__file__), 'static', 'imgs')
    file_path = os.path.normpath(os.path.join(static_dir, filename))

    # 确保文件仍在static_dir内
    if not file_path.startswith(os.path.abspath(static_dir) + os.sep):
        return jsonify({'success': False, 'error': '无效的文件路径'}), 400

    if not os.path.exists(file_path):
        return jsonify({'success': False, 'error': '文件不存在'}), 404

    return send_file(file_path)


@app.route('/api/custom_template_file/<template_id>/<path:filename>', methods=['GET'])
@login_required
def serve_custom_template_file(template_id, filename):
    """服务自定义模板的临时文件"""
    # 安全检查：防止路径遍历
    template_dir = os.path.join(CUSTOM_TEMPLATE_FOLDER, template_id)
    file_path = os.path.normpath(os.path.join(template_dir, filename))

    # 确保文件仍在template_dir内
    if not file_path.startswith(os.path.abspath(template_dir) + os.sep):
        return jsonify({'success': False, 'error': '无效的文件路径'}), 400

    if not os.path.exists(file_path):
        return jsonify({'success': False, 'error': '文件不存在'}), 404

    return send_file(file_path)


def run_complaint_script(task_payload):
    """在后台线程中执行UC投诉自动化脚本"""
    import sys

    task_id = task_payload['task_id']
    script_path = os.path.join(os.path.dirname(__file__), 'uc_complaint_from_backend.py')
    submission_id = task_id[3:] if task_id.startswith('uc_') else task_id

    works_config = task_payload.get('works_config', [])
    total_batches = task_payload.get('total_batches', 0)
    operator = task_payload.get('operator', '')

    cmd = [
        sys.executable,
        script_path,
        '--task-id', task_id,
        '--cookie', task_payload.get('cookie', ''),
        '--identity', task_payload.get('identity', ''),
        '--agent', task_payload.get('agent', ''),
        '--rights-holder', task_payload.get('rights_holder', ''),
        '--module', task_payload.get('module', ''),
        '--content-type', task_payload.get('content_type', ''),
        '--description', task_payload.get('description', ''),
        '--works-config', json.dumps(works_config, ensure_ascii=False),
    ]

    complaint_category = task_payload.get('complaint_category', '')
    copyright_type = task_payload.get('copyright_type', '')
    if complaint_category == '知识产权' and copyright_type:
        cmd.extend(['--complaint-type', complaint_category, '--copyright-type', copyright_type])

    print(f"[{task_id}] 执行UC多作品投诉，作品数: {len(works_config)}，总批次: {total_batches}")
    append_task_log(task_id, f"操作人: {operator}, 执行UC多作品投诉，作品数: {len(works_config)}，总批次: {total_batches}")

    try:
        started_at = datetime.now().isoformat()
        append_task_log(task_id, f"任务开始执行，started_at={started_at}")
        if task_id in tasks:
            tasks[task_id]['status'] = 'running'
            tasks[task_id]['started_at'] = started_at
        update_complaint_task(task_id, status='running', started_at=datetime.fromisoformat(started_at))

        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            # 每批现在要 reload + 完整重填表单 + 重传证明文件，实测约 5 分钟/批，
            # 旧的 300s/批毫无余量必超时。给 360s/批 + 180s 基础（覆盖初始打开和
            # 最后的单号查询），下限 900s。超时会导致已成功的投诉拿不到单号。
            timeout=max(900, 180 + total_batches * 360)
        )

        print(f"[{task_id}] stdout: {result.stdout}")
        print(f"[{task_id}] stderr: {result.stderr}")
        append_task_log(task_id, 'stdout:\n' + (result.stdout or ''))
        append_task_log(task_id, 'stderr:\n' + (result.stderr or ''))

        task_result = None
        try:
            start_idx = result.stdout.find('JSON_RESULT_START')
            end_idx = result.stdout.find('JSON_RESULT_END')
            if start_idx != -1 and end_idx != -1:
                json_str = result.stdout[start_idx + 17:end_idx].strip()
                task_result = json.loads(json_str)
        except Exception:
            pass

        if task_result:
            append_task_log(task_id, '解析到 JSON_RESULT，准备更新任务状态')
            if task_id in tasks:
                tasks[task_id].update(task_result)

            # 追加 skipped_works 信息到 complaint_numbers
            complaint_numbers = task_result.get('complaint_numbers', [])
            skipped_in_payload = task_payload.get('skipped_works', [])
            for sw in skipped_in_payload:
                complaint_numbers.append(f"{sw['work_name']}：{sw.get('reason', '作品覆盖列表中未匹配到或证明文件不齐全')}")

            update_fields = {
                'status': task_result.get('status'),
                'current_batch': task_result.get('current_batch'),
                'completed_batches': task_result.get('completed_batches'),
                'failed_batches': task_result.get('failed_batches'),
                'complaint_numbers_json': complaint_numbers,
                'error_message': task_result.get('error'),
            }
            if task_result.get('started_at'):
                update_fields['started_at'] = datetime.fromisoformat(task_result['started_at'])
            if task_result.get('completed_at'):
                update_fields['completed_at'] = datetime.fromisoformat(task_result['completed_at'])
            update_complaint_task(task_id, **update_fields)
            for batch in task_result.get('batches', []):
                update_complaint_batch(
                    submission_id,
                    batch['batch_no'],
                    status=batch.get('status'),
                    complaint_number=batch.get('complaint_number'),
                    error_message=batch.get('error')
                )
            append_task_log(task_id, f"任务执行完成，status={task_result.get('status')}, complaint_numbers={task_result.get('complaint_numbers', [])}")
            sync_task_log_to_db(task_id, submission_id, task_result.get('status'))
        else:
            if task_id in tasks:
                tasks[task_id]['status'] = 'failed'
                tasks[task_id]['error'] = result.stderr or '执行失败'
            update_complaint_task(
                task_id,
                status='failed',
                error_message=result.stderr or '执行失败',
                completed_at=datetime.now()
            )
            append_task_log(task_id, '未解析到 JSON_RESULT，任务标记为 failed')
            sync_task_log_to_db(task_id, submission_id, 'failed')
    except subprocess.TimeoutExpired:
        # 超时被杀时，脚本没机会输出 JSON_RESULT。但它每批确认成功后已把单号
        # 逐批写进 task_results/<task_id>.json，这里读回来，避免“已成功投诉的单号全丢”。
        recovered = load_task_result(task_id)
        recovered_numbers = (recovered or {}).get('complaint_numbers') or []
        rec_completed = (recovered or {}).get('completed_batches')
        rec_failed = (recovered or {}).get('failed_batches')
        if task_id in tasks:
            tasks[task_id]['status'] = 'partial_failed' if recovered_numbers else 'failed'
            tasks[task_id]['error'] = '执行超时'
            if recovered_numbers:
                tasks[task_id]['complaint_numbers'] = recovered_numbers
        timeout_fields = {
            'status': 'partial_failed' if recovered_numbers else 'failed',
            'error_message': '执行超时',
            'completed_at': datetime.now(),
        }
        if recovered_numbers:
            timeout_fields['complaint_numbers_json'] = recovered_numbers
        if rec_completed is not None:
            timeout_fields['completed_batches'] = rec_completed
        if rec_failed is not None:
            timeout_fields['failed_batches'] = rec_failed
        update_complaint_task(task_id, **timeout_fields)
        # 回填已完成批次的单号到批次子表
        for batch in (recovered or {}).get('batches', []):
            if batch.get('status') == 'completed':
                update_complaint_batch(
                    submission_id, batch['batch_no'],
                    status='completed', complaint_number=batch.get('complaint_number')
                )
        msg = '任务执行超时' + (f'（已回收 {len(recovered_numbers)} 个已成功批次单号）' if recovered_numbers else '')
        append_task_log(task_id, msg)
        sync_task_log_to_db(task_id, submission_id, timeout_fields['status'])
    except Exception as e:
        if task_id in tasks:
            tasks[task_id]['status'] = 'failed'
            tasks[task_id]['error'] = str(e)
        update_complaint_task(task_id, status='failed', error_message=str(e), completed_at=datetime.now())
        append_task_log(task_id, f'任务执行异常: {str(e)}')
        sync_task_log_to_db(task_id, submission_id, 'failed')


@app.route('/api/uc/task/<task_id>', methods=['GET'])
@login_required
def get_task_status(task_id):
    """查询任务状态"""
    task = get_complaint_task(task_id)
    if not task:
        task = tasks.get(task_id)
    if not task:
        task = load_task_result(task_id)

    if not task:
        return jsonify({'success': False, 'error': '任务不存在'}), 404

    return jsonify({
        'success': True,
        'task_id': task_id,
        'status': task.get('status'),
        'complaint_number': task.get('complaint_number'),
        'complaint_numbers': task.get('complaint_numbers', []),
        'batch_count': task.get('batch_count'),
        'completed_batches': task.get('completed_batches'),
        'current_batch': task.get('current_batch'),
        'batches': task.get('batches', []),
        'error': task.get('error'),
        'submitted_at': task.get('submitted_at'),
        'started_at': task.get('started_at'),
        'completed_at': task.get('completed_at'),
        'log_file_path': task.get('log_file_path'),
    })


@app.route('/uc/task/<task_id>/log', methods=['GET'])
@login_required
def view_task_log(task_id):
    task_log = get_task_execution_log(task_id)
    if not task_log:
        log_text = read_task_log_file(task_id)
        if not log_text:
            return f"""
            <!DOCTYPE html>
            <html lang='zh-CN'>
            <head>
                <meta charset='UTF-8'>
                <title>任务日志不存在</title>
                <style>
                    body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; background: #f7f7f9; margin: 0; padding: 24px; color: #222; }}
                    .box {{ max-width: 760px; margin: 40px auto; background: #fff; border: 1px solid #ddd; border-radius: 8px; padding: 24px; }}
                    .title {{ font-size: 20px; font-weight: 600; margin-bottom: 12px; }}
                    .desc {{ color: #666; line-height: 1.7; }}
                    code {{ background: #f1f3f5; padding: 2px 6px; border-radius: 4px; }}
                </style>
            </head>
            <body>
                <div class='box'>
                    <div class='title'>这条任务目前没有可查看的日志</div>
                    <div class='desc'>
                        <p><strong>任务ID：</strong><code>{html.escape(task_id)}</code></p>
                        <p>可能原因：</p>
                        <ul>
                            <li>这是较早的历史任务，当时还没有启用详细日志保存功能。</li>
                            <li>本地日志文件已经被清理（当前本地日志只保留最近 5 天）。</li>
                            <li>数据库里也没有同步到这条任务的日志内容。</li>
                        </ul>
                    </div>
                </div>
            </body>
            </html>
            """, 404
        task_log = {
            'task_id': task_id,
            'submission_id': task_id[3:] if task_id.startswith('uc_') else task_id,
            'status': 'unknown',
            'log_text': log_text,
            'updated_at': datetime.now(),
        }

    title = f"任务日志 - {task_id}"
    safe_log = html.escape(task_log.get('log_text') or '')
    safe_status = html.escape(str(task_log.get('status') or 'unknown'))
    safe_submission = html.escape(str(task_log.get('submission_id') or ''))
    safe_updated = html.escape(normalize_datetime(task_log.get('updated_at')) or '-')
    return f"""
    <!DOCTYPE html>
    <html lang='zh-CN'>
    <head>
        <meta charset='UTF-8'>
        <title>{title}</title>
        <style>
            body {{ font-family: Menlo, Monaco, Consolas, monospace; background: #f7f7f9; margin: 0; padding: 24px; color: #222; }}
            .meta {{ background: #fff; border: 1px solid #ddd; border-radius: 8px; padding: 16px; margin-bottom: 16px; }}
            .meta div {{ margin-bottom: 6px; }}
            pre {{ background: #fff; border: 1px solid #ddd; border-radius: 8px; padding: 16px; overflow-x: auto; white-space: pre-wrap; word-break: break-word; line-height: 1.5; }}
        </style>
    </head>
    <body>
        <div class='meta'>
            <div><strong>任务ID：</strong>{html.escape(task_id)}</div>
            <div><strong>Submission ID：</strong>{safe_submission}</div>
            <div><strong>状态：</strong>{safe_status}</div>
            <div><strong>最后更新时间：</strong>{safe_updated}</div>
        </div>
        <pre>{safe_log}</pre>
    </body>
    </html>
    """


@app.route('/api/worker/queue_status', methods=['GET'])
@login_required
def worker_queue_status():
    session = get_db_session()
    try:
        rows = session.execute(text("""
            SELECT task_id, platform_code, batch_count, status, submitted_at
            FROM complaints
            WHERE status IN ('queued', 'running')
            ORDER BY submitted_at ASC
        """)).fetchall()

        queue = []
        for row in rows:
            queue.append({
                'task_id': row.task_id,
                'platform_code': row.platform_code,
                'batch_count': row.batch_count or 1,
                'status': row.status,
                'submitted_at': normalize_datetime(row.submitted_at),
            })
        return jsonify({'success': True, 'queue': queue})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        session.close()


@app.route('/api/uc/status_list', methods=['GET'])
@login_required
def get_uc_status_list():
    """获取UC投诉状态列表"""
    submissions = get_submission_status_list()
    if submissions:
        return jsonify({'success': True, 'data': submissions})

    submissions = []
    uc_submissions_path = app.config['UC_SUBMISSION_FOLDER']

    if not os.path.exists(uc_submissions_path):
        return jsonify({'success': True, 'data': []})

    for item in sorted(os.listdir(uc_submissions_path), reverse=True):
        item_path = os.path.join(uc_submissions_path, item)
        if not os.path.isdir(item_path):
            continue

        submission_file = os.path.join(item_path, 'submission.json')
        if not os.path.exists(submission_file):
            continue

        try:
            with open(submission_file, 'r', encoding='utf-8') as f:
                data = json.load(f)

            task_id = f"uc_{data.get('submission_id', item)}"
            task_info = tasks.get(task_id)
            if not task_info:
                task_info = load_task_result(task_id)

            status = '未知'
            if task_info:
                status = map_task_status_label(task_info.get('status', '未知'))

            complaint_numbers = []
            if task_info and task_info.get('complaint_numbers'):
                complaint_numbers = task_info.get('complaint_numbers', [])
            elif task_info and task_info.get('complaint_number'):
                complaint_numbers = [task_info.get('complaint_number')]

            submissions.append({
                'submission_id': data.get('submission_id', item),
                'submitted_at': data.get('submitted_at', ''),
                'collect_account': data.get('form', {}).get('collect_account', ''),
                'work_name': data.get('form', {}).get('作品名称', ''),
                'excel_rows': data.get('excel_rows', 0),
                'batch_count': data.get('batch_count', 0),
                'status': status,
                'complaint_numbers': complaint_numbers,
            })
        except Exception:
            continue

    return jsonify({'success': True, 'data': submissions})


@app.route('/api/uc/export_excel/<submission_id>', methods=['GET'])
@login_required
def uc_export_excel(submission_id):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    session = get_db_session()
    try:
        sub = session.execute(text("""
            SELECT complaint_id, collect_account, work_name, submitted_at,
                   complaint_numbers_json
            FROM complaints
            WHERE complaint_id = :sid AND platform_code = 'uc'
        """), {'sid': submission_id}).fetchone()
        if not sub:
            return jsonify({'success': False, 'error': '记录不存在'}), 404

        # 解析投诉单号（按批次顺序：单号是按每200条一片产生的，一部作品可能占多个单号）
        complaint_numbers = []
        if sub.complaint_numbers_json:
            try:
                complaint_numbers = json.loads(sub.complaint_numbers_json)
            except:
                pass

        # 重建「每个单号属于哪部作品」：用 submission.json 的 works_config，
        # 它每条都有 work_name + batch_count（新老记录都在）。把每部作品名按 batch_count
        # 展开，拼成与 complaint_numbers 等长的作品名序列，再逐行配对。
        # 例：宦妃天下(3批)+开局一座山(2批) → [宦妃,宦妃,宦妃,开局,开局] zip 5个单号。
        expanded_work_names = []
        try:
            submission_file = os.path.join(
                app.config['UC_SUBMISSION_FOLDER'], submission_id, 'submission.json'
            )
            if os.path.exists(submission_file):
                with open(submission_file, 'r', encoding='utf-8') as f:
                    sub_meta = json.load(f)
                for w in sub_meta.get('works_config', []):
                    wn = (w.get('work_name') or '').strip()
                    bc = w.get('batch_count') or len(w.get('excel_files', [])) or 1
                    expanded_work_names.extend([wn] * int(bc))
        except Exception:
            expanded_work_names = []

        # 兜底：读不到 works_config 时，退回按逗号拆分的作品名（旧逻辑）
        if not expanded_work_names:
            work_name_str = sub.work_name or ''
            if ',' in work_name_str:
                expanded_work_names = [w.strip() for w in work_name_str.split(',') if w.strip()]
            elif '，' in work_name_str:
                expanded_work_names = [w.strip() for w in work_name_str.split('，') if w.strip()]
            else:
                expanded_work_names = [work_name_str] if work_name_str else []

        # 提交时间格式化
        submitted_at = ''
        if sub.submitted_at:
            submitted_at = sub.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if hasattr(sub.submitted_at, 'strftime') else str(sub.submitted_at)

        # 生成 Excel
        wb = Workbook()
        ws = wb.active
        ws.title = '投诉结果'
        ws.append(['采集时间', '采集账号', '作品名称', '投诉单号'])
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # complaint_numbers 末尾可能追加了 skipped_works 的占位串（形如「作品名：原因」），
        # 这些超出 expanded_work_names 长度的行，作品名留空、单号列原样写占位串。
        max_rows = max(len(expanded_work_names), len(complaint_numbers))
        for i in range(max_rows):
            wn = expanded_work_names[i] if i < len(expanded_work_names) else ''
            fn = complaint_numbers[i] if i < len(complaint_numbers) else ''
            ws.append([submitted_at, sub.collect_account, wn, str(fn)])

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 30

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f'uc_result_{submission_id}.xlsx'
        return send_file(buf, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    finally:
        session.close()


@app.route('/api/uc/verify_cookie', methods=['POST'])
@login_required
def verify_cookie():
    """验证Cookie是否有效"""
    data = request.get_json()
    cookie = data.get('cookie', '').strip()

    if not cookie:
        return jsonify({'success': False, 'error': 'Cookie不能为空'}), 400

    from playwright.sync_api import sync_playwright

    def cookie_to_context(context, cookie_value):
        if cookie_value.startswith('[') or cookie_value.startswith('{'):
            cookies = json.loads(cookie_value) if isinstance(cookie_value, str) else cookie_value
            context.add_cookies(cookies)
            return

        for pair in cookie_value.split(';'):
            pair = pair.strip()
            if '=' in pair:
                key, value = pair.split('=', 1)
                context.add_cookies([{
                    'name': key,
                    'value': value,
                    'domain': '.uc.cn',
                    'path': '/'
                }])

    def verify_cookie_by_api(cookie_value):
        xtstk = extract_xtstk_from_cookie(cookie_value)
        headers = {
            'accept': '*/*',
            'accept-language': 'zh-CN,zh;q=0.9',
            'referer': 'https://ipp.uc.cn/',
            'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 '
                          '(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'xtstk': xtstk,
            'cookie': cookie_value,
        }
        resp = requests.get(
            UC_COMPLAIN_LIST_API,
            params={'pageNo': 1, 'pageSize': 1, 'platform': 'uc'},
            headers=headers,
            timeout=15,
        )
        if resp.status_code != 200:
            return False, f'接口返回状态码 {resp.status_code}'

        body = resp.json()
        if body.get('code') != 200:
            return False, f'接口返回异常：{body}'

        return True, 'Cookie有效（接口校验通过）'

    try:
        api_valid = False
        api_message = ''
        try:
            api_valid, api_message = verify_cookie_by_api(cookie)
        except Exception as api_error:
            api_message = str(api_error)

        if api_valid:
            return jsonify({'success': True, 'message': api_message})

        with sync_playwright() as p:
            chromium_path = os.getenv('PLAYWRIGHT_CHROMIUM_PATH', '').strip()
            launch_kwargs = {
                'headless': True,
                'args': [
                    "--disable-blink-features=AutomationControlled",
                    "--no-sandbox",
                    "--disable-setuid-sandbox",
                    "--lang=zh-CN,en-US",
                ],
            }
            if chromium_path:
                launch_kwargs['executable_path'] = chromium_path

            browser = p.chromium.launch(**launch_kwargs)
            context = browser.new_context(
                user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                viewport={"width": 1920, "height": 1080},
            )

            cookie_to_context(context, cookie)

            page = context.new_page()
            page.goto("https://ipp.uc.cn/#/home", wait_until="load", timeout=15000)
            page.wait_for_timeout(2000)

            login_dialog = page.locator("text=UC账号登录").first
            if login_dialog.count() > 0 and login_dialog.is_visible():
                browser.close()
                return jsonify({'success': False, 'error': f'Cookie已过期，请重新登录（接口校验提示：{api_message or "未通过"}）'}), 401

            browser.close()
            return jsonify({'success': True, 'message': 'Cookie有效（页面兜底校验通过）'})

    except Exception as e:
        return jsonify({'success': False, 'error': f'验证失败：{str(e)}'}), 500


# ============================================================
# 百度投诉平台路由
# ============================================================

@app.route('/baidu')
@login_required
def baidu_page():
    return render_template('baidu.html')


@app.route('/quark')
@login_required
def quark_page():
    return render_template('quark.html')


@app.route('/api/baidu/verify_cookie', methods=['POST'])
@login_required
def baidu_verify_cookie():
    data = request.get_json() or {}
    cookie = data.get('cookie', '').strip()
    if not cookie:
        return jsonify({'success': False, 'error': 'Cookie不能为空'}), 400
    try:
        resp = requests.get(
            f'{BAIDU_API_BASE}/login/check',
            headers={'Cookie': cookie, 'User-Agent': 'Mozilla/5.0'},
            timeout=10
        )
        result = resp.json()
        if result.get('code') == 200 and result.get('data', {}).get('uid'):
            return jsonify({'success': True, 'user': result['data']})
        return jsonify({'success': False, 'error': 'Cookie无效或已过期'}), 401
    except Exception as e:
        return jsonify({'success': False, 'error': f'验证失败：{str(e)}'}), 500


@app.route('/api/baidu/pre_check', methods=['POST'])
@login_required
def baidu_pre_check():
    data = request.get_json() or {}
    cookie = data.get('cookie', '').strip()
    work_names = data.get('work_names', [])

    if not cookie:
        return jsonify({'success': False, 'error': 'Cookie不能为空'}), 400
    if not work_names:
        return jsonify({'success': False, 'error': '作品列表为空'}), 400

    can_complain = []
    cannot_complain = []

    try:
        for work_name in work_names:
            resp = requests.post(
                f'{BAIDU_API_BASE}/ownership/keyword',
                headers={
                    'Cookie': cookie,
                    'Content-Type': 'application/json',
                    'User-Agent': 'Mozilla/5.0',
                },
                json={'page': 1, 'size': 50, 'lastPageNo': 0, 'key_word': work_name, 'owner_type': 0},
                timeout=15,
            )
            result = resp.json()
            found_passed = False
            found_but_not_passed = False

            if result.get('code') == 200:
                records = result.get('data', {}).get('records', [])
                for record in records:
                    if record.get('works_name', '') == work_name:
                        if record.get('ownership_status') == 2:
                            found_passed = True
                            break
                        else:
                            found_but_not_passed = True

            if found_passed:
                can_complain.append(work_name)
            elif found_but_not_passed:
                cannot_complain.append({
                    'work_name': work_name,
                    'reason': '权属状态未通过，请在百度投诉原平台进行投诉',
                })
            else:
                cannot_complain.append({
                    'work_name': work_name,
                    'reason': '未找到已通过审核的权属记录，请在百度投诉原平台进行投诉',
                })

        return jsonify({
            'success': True,
            'can_complain': can_complain,
            'cannot_complain': cannot_complain,
        })
    except Exception as e:
        return jsonify({'success': False, 'error': f'预检失败：{str(e)}'}), 500


@app.route('/api/baidu/search_ownership', methods=['POST'])
@login_required
def baidu_search_ownership():
    data = request.get_json() or {}
    cookie = data.get('cookie', '').strip()
    key_word = data.get('key_word', '').strip()
    if not cookie:
        return jsonify({'success': False, 'error': 'Cookie不能为空'}), 400
    try:
        resp = requests.post(
            f'{BAIDU_API_BASE}/ownership/keyword',
            headers={
                'Cookie': cookie,
                'Content-Type': 'application/json',
                'User-Agent': 'Mozilla/5.0',
            },
            json={
                'page': 1,
                'size': 50,
                'lastPageNo': 0,
                'key_word': key_word,
                'owner_type': 0,
            },
            timeout=15
        )
        result = resp.json()
        if result.get('code') == 200:
            return jsonify({'success': True, 'data': result.get('data', {})})
        return jsonify({'success': False, 'error': result.get('message', '查询失败')}), 400
    except Exception as e:
        return jsonify({'success': False, 'error': f'查询失败：{str(e)}'}), 500


@app.route('/api/baidu/download_template', methods=['GET'])
@login_required
def baidu_download_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()

    # Sheet1: 投诉配置
    ws1 = wb.active
    ws1.title = '投诉配置'
    ws1.append(['字段', '值', '可选值'])
    ws1.append(['投诉产品', '', '百度网盘 / 百度搜索 / 百度APP / 百家号 / 百度贴吧 / 好看视频'])
    ws1.append(['侵权类型', '', '影视版权 / 综艺版权 / 动漫动画版权 / 音乐版权 / 游戏版权 / 体育赛事版权 / 新闻媒体版权 / 自媒体版权（仅投诉产品为"好看视频"时需填写）'])
    header_font = Font(bold=True)
    for cell in ws1[1]:
        cell.font = header_font

    # Sheet2: 作品列表
    ws2 = wb.create_sheet('作品列表')
    ws2.append(['作品名称', '投诉问题描述', '原版链接标题', '原版链接地址'])
    ws2.append(['示例作品名', '链接涉及上传分享传播独家作品存在侵权行为 请尽快处理', '示例作品名', 'https://www.example.com/original'])
    for cell in ws2[1]:
        cell.font = header_font
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 50
    ws2.column_dimensions['C'].width = 25
    ws2.column_dimensions['D'].width = 45

    # Sheet3: 侵权链接
    ws3 = wb.create_sheet('侵权链接')
    ws3.append(['序号', '链接名称', '链接地址', '作品名称'])
    ws3.append([1, '示例作品名', 'https://pan.baidu.com/s/example1', '示例作品名'])
    ws3.append([2, '示例作品名', 'https://pan.baidu.com/s/example2', '示例作品名'])
    for cell in ws3[1]:
        cell.font = header_font
    ws3.column_dimensions['A'].width = 8
    ws3.column_dimensions['B'].width = 30
    ws3.column_dimensions['C'].width = 55
    ws3.column_dimensions['D'].width = 25

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='baidu_custom_template.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/baidu/upload_template', methods=['POST'])
@login_required
def baidu_upload_template():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '未上传文件'}), 400
    file = request.files['file']
    if not file.filename:
        return jsonify({'success': False, 'error': '文件名为空'}), 400
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ('.xlsx', '.xls'):
        return jsonify({'success': False, 'error': '仅支持 .xlsx / .xls 格式'}), 400

    cookie = request.form.get('cookie', '').strip()
    if not cookie:
        return jsonify({'success': False, 'error': '请先选择账号'}), 400

    try:
        wb = load_workbook(file, data_only=True)
    except Exception as e:
        return jsonify({'success': False, 'error': f'文件解析失败：{str(e)}'}), 400

    sheet_names = wb.sheetnames
    if '投诉配置' not in sheet_names:
        return jsonify({'success': False, 'error': '缺少"投诉配置"工作表'}), 400
    if '作品列表' not in sheet_names:
        return jsonify({'success': False, 'error': '缺少"作品列表"工作表'}), 400
    if '侵权链接' not in sheet_names:
        return jsonify({'success': False, 'error': '缺少"侵权链接"工作表'}), 400

    # 解析 Sheet1: 投诉配置
    ws_config = wb['投诉配置']
    config = {}
    for row in ws_config.iter_rows(min_row=2, max_col=2, values_only=True):
        if row[0] and row[1]:
            config[str(row[0]).strip()] = str(row[1]).strip()

    complaint_product = config.get('投诉产品', '').strip()
    if not complaint_product:
        return jsonify({'success': False, 'error': '投诉配置中"投诉产品"不能为空'}), 400
    if complaint_product not in BAIDU_COMPLAINT_TYPE_MAP:
        return jsonify({'success': False, 'error': f'不支持的投诉产品：{complaint_product}，可选：{", ".join(BAIDU_COMPLAINT_TYPE_MAP.keys())}'}), 400

    infringe_type = None
    if complaint_product == '好看视频':
        infringe_type_str = config.get('侵权类型', '').strip()
        if not infringe_type_str:
            return jsonify({'success': False, 'error': '投诉产品为"好看视频"时，"侵权类型"不能为空'}), 400
        if infringe_type_str not in BAIDU_INFRINGE_TYPE_MAP:
            return jsonify({'success': False, 'error': f'不支持的侵权类型：{infringe_type_str}，可选：{", ".join(BAIDU_INFRINGE_TYPE_MAP.keys())}'}), 400
        infringe_type = BAIDU_INFRINGE_TYPE_MAP[infringe_type_str]

    # 解析 Sheet2: 作品列表
    ws_works = wb['作品列表']
    works_list = []
    for row in ws_works.iter_rows(min_row=2, max_col=4, values_only=True):
        if not row[0]:
            continue
        work_name = str(row[0]).strip()
        description = str(row[1]).strip() if row[1] else ''
        actual_name = str(row[2]).strip() if row[2] else ''
        actual_url = str(row[3]).strip() if row[3] else ''
        if not description:
            return jsonify({'success': False, 'error': f'作品"{work_name}"的投诉问题描述不能为空'}), 400
        if not actual_name or not actual_url:
            return jsonify({'success': False, 'error': f'作品"{work_name}"的原版链接标题和地址不能为空'}), 400
        works_list.append({
            'work_name': work_name,
            'description': description,
            'actual_name': actual_name,
            'actual_url': actual_url,
        })

    if not works_list:
        return jsonify({'success': False, 'error': '"作品列表"中没有有效数据'}), 400

    # 校验 Sheet2 作品名称不允许重复
    seen_work_names = []
    duplicate_work_names = []
    for w in works_list:
        if w['work_name'] in seen_work_names:
            if w['work_name'] not in duplicate_work_names:
                duplicate_work_names.append(w['work_name'])
        else:
            seen_work_names.append(w['work_name'])
    if duplicate_work_names:
        return jsonify({'success': False, 'error': f'"作品列表"中存在重复的作品名称：{", ".join(duplicate_work_names)}'}), 400

    # 解析 Sheet3: 侵权链接
    ws_links = wb['侵权链接']
    all_links = []
    emoji_cleaned = []  # 链接名称含 emoji 被自动清洗的行：{row, original, cleaned}
    illegal_url_rows = []  # 链接地址含汉字/全角标点的行：{row, chars}
    for row_idx, row in enumerate(ws_links.iter_rows(min_row=2, max_col=4, values_only=True), start=2):
        if not row[2]:
            continue
        link_name = str(row[1]).strip() if row[1] else ''
        link_url = str(row[2]).strip()
        work_name = str(row[3]).strip() if row[3] else ''
        if not link_url.startswith(('http://', 'https://')):
            return jsonify({'success': False, 'error': f'链接地址格式错误（必须以http://或https://开头）：{link_url}'}), 400
        # 链接地址含汉字/全角标点：链接本身有误，收集所有问题行后统一报错（不自动改）
        illegal_chars = find_illegal_url_chars(link_url)
        if illegal_chars:
            illegal_url_rows.append({'row': row_idx, 'chars': illegal_chars})
            continue
        if not work_name:
            return jsonify({'success': False, 'error': f'侵权链接"{link_url}"缺少作品名称'}), 400
        # 清洗链接名称中的 emoji：百度 /upload 含 emoji 会返回 code=500，必须先去掉
        cleaned_name, changed = strip_emoji(link_name)
        if changed:
            emoji_cleaned.append({'row': row_idx, 'original': link_name, 'cleaned': cleaned_name})
            link_name = cleaned_name
        all_links.append({
            'link_name': link_name,
            'link_url': link_url,
            'work_name': work_name,
        })

    # 链接地址含汉字/全角标点：禁止投诉，列出所有问题行让用户删除后重传
    if illegal_url_rows:
        detail = '；'.join(f"第{r['row']}行（含「{r['chars']}」）" for r in illegal_url_rows[:10])
        more = f"，等共{len(illegal_url_rows)}行" if len(illegal_url_rows) > 10 else ''
        return jsonify({'success': False, 'error':
            f'以下行的链接地址含汉字或全角标点，链接无法投诉，请删除这些字符后重新上传：\n{detail}{more}'}), 400

    if not all_links:
        return jsonify({'success': False, 'error': '"侵权链接"中没有有效数据'}), 400

    # 校验链接地址是否重复（全局判重，只要链接地址相同即视为重复，与作品名称无关）
    link_positions = {}  # {url: [行号列表]}
    for idx, link in enumerate(all_links):
        key = link['link_url']
        if key not in link_positions:
            link_positions[key] = []
        link_positions[key].append(idx + 2)  # +2 因为第1行是标题，idx从0开始
    duplicate_errors = []
    for url, rows in link_positions.items():
        if len(rows) > 1:
            duplicate_errors.append(f'第{rows[0]}行与第{rows[1]}行链接地址重复')
    if duplicate_errors:
        return jsonify({'success': False, 'error': '侵权链接中存在重复，请删除后重新上传：\n' + '\n'.join(duplicate_errors[:5])}), 400

    # 按作品名称分组链接
    links_by_work = {}
    for link in all_links:
        wn = link['work_name']
        if wn not in links_by_work:
            links_by_work[wn] = []
        links_by_work[wn].append({'link_name': link['link_name'], 'url_address': link['link_url']})

    # 校验：以 Sheet3 侵权链接中的作品为准，Sheet2 多出来的忽略
    work_names_in_list = {w['work_name'] for w in works_list}
    work_names_in_links = set(links_by_work.keys())

    # Sheet3 中有链接但 Sheet2 中没有对应作品的，记录警告但不阻断
    extra_links_warnings = []
    extra_link_works = work_names_in_links - work_names_in_list
    if extra_link_works:
        for wn in extra_link_works:
            extra_links_warnings.append(f'{wn}（上传的文件作品列表中未找到该作品配置）')
            del links_by_work[wn]

    # 构建结果（只处理 Sheet3 中有链接的作品，按 Sheet3 中出现的顺序）
    # 先收集 Sheet3 中链接的作品出现顺序
    link_work_order = []
    for link in all_links:
        wn = link['work_name']
        if wn not in link_work_order and wn in links_by_work:
            link_work_order.append(wn)

    works_config = []
    total_links = 0
    total_batches = 0
    for wn in link_work_order:
        # 从 Sheet2 中找到对应的作品配置
        work_info = next((w for w in works_list if w['work_name'] == wn), None)
        if not work_info:
            continue
        work_links = links_by_work[wn]
        link_count = len(work_links)
        batch_count = math.ceil(link_count / 200)
        total_links += link_count
        total_batches += batch_count
        works_config.append({
            **work_info,
            'links': work_links,
            'link_count': link_count,
            'batch_count': batch_count,
        })

    if not works_config:
        return jsonify({'success': False, 'error': '没有可投诉的作品（侵权链接中的作品在作品列表中均未找到配置）'}), 400

    return jsonify({
        'success': True,
        'complaint_product': complaint_product,
        'complaint_type_code': BAIDU_COMPLAINT_TYPE_MAP[complaint_product],
        'infringe_type': infringe_type,
        'upload_filename': Path(file.filename).name,
        'works': works_config,
        'total_works': len(works_config),
        'total_links': total_links,
        'total_batches': total_batches,
        'skipped_works': extra_links_warnings,
        'emoji_cleaned': emoji_cleaned,
    })


@app.route('/api/baidu/submit', methods=['POST'])
@login_required
def baidu_submit():
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': '请求数据为空'}), 400

    cookie = data.get('cookie', '').strip()
    collect_account = data.get('collect_account', '').strip()
    complaint_product = data.get('complaint_product', '').strip()
    complaint_type_code = data.get('complaint_type_code')
    infringe_type = data.get('infringe_type')
    works_config = data.get('works', [])
    skipped_works = data.get('skipped_works', [])
    upload_filename = data.get('upload_filename', '').strip()

    if not cookie:
        return jsonify({'success': False, 'error': 'Cookie不能为空'}), 400
    if not collect_account:
        return jsonify({'success': False, 'error': '请选择投诉账号'}), 400
    if not complaint_product:
        return jsonify({'success': False, 'error': '投诉产品不能为空'}), 400

    # 防重复：同账号+同文件名已有非失败记录则拒绝
    if upload_filename:
        _sess = get_db_session()
        try:
            dup = _sess.execute(text("""
                SELECT task_id FROM complaints
                WHERE collect_account=:acc AND upload_filename=:fn
                  AND platform_code='baidu' AND status NOT IN ('failed')
                LIMIT 1
            """), {'acc': collect_account, 'fn': upload_filename}).fetchone()
        finally:
            _sess.close()
        if dup:
            return jsonify({'success': False, 'error': f'文件「{upload_filename}」已投诉过（任务 {dup[0]}），请勿重复提交'}), 400
    if not works_config and not skipped_works:
        return jsonify({'success': False, 'error': '作品列表不能为空'}), 400

    # 验证 Cookie
    try:
        resp = requests.get(
            f'{BAIDU_API_BASE}/login/check',
            headers={'Cookie': cookie, 'User-Agent': 'Mozilla/5.0'},
            timeout=10
        )
        result = resp.json()
        if result.get('code') != 200 or not result.get('data', {}).get('uid'):
            return jsonify({'success': False, 'error': 'Cookie已失效，请更新后重试'}), 401
    except Exception as e:
        return jsonify({'success': False, 'error': f'Cookie验证失败：{str(e)}'}), 500

    # 计算批次（只算可投诉的作品）
    total_links = 0
    total_batches = 0
    for work in works_config:
        link_count = len(work.get('links', []))
        total_links += link_count
        total_batches += math.ceil(link_count / 200)

    # 所有作品名称（包含跳过的）
    all_work_names = [w['work_name'] for w in works_config] + [w['work_name'] for w in skipped_works]

    # 创建提交目录
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    submission_id = f"{timestamp}_{uuid4().hex[:8]}"
    submission_dir = os.path.join(app.config['BAIDU_SUBMISSION_FOLDER'], submission_id)
    os.makedirs(submission_dir, exist_ok=True)

    # 保存 submission.json
    submission_data = {
        'submission_id': submission_id,
        'platform_code': 'baidu',
        'collect_account': collect_account,
        'cookie': cookie,
        'complaint_product': complaint_product,
        'complaint_type_code': complaint_type_code,
        'works_config': works_config,
        'skipped_works': skipped_works,
        'total_works': len(works_config),
        'total_links': total_links,
        'total_batches': total_batches,
        'submitted_at': datetime.now().isoformat(),
    }
    with open(os.path.join(submission_dir, 'submission.json'), 'w', encoding='utf-8') as f:
        json.dump(submission_data, f, ensure_ascii=False, indent=2)

    task_id = f'baidu_{submission_id}'

    # 写入数据库
    session = get_db_session()
    try:
        work_names_str = ', '.join(all_work_names)
        session.execute(text("""
            INSERT INTO complaints
            (complaint_id, task_id, platform_code, collect_account, cookie_snapshot,
             identity_type, agent_name, principal_name,
             complaint_category, complaint_type, module_name, content_type,
             description_text, work_name, total_links, batch_size, batch_count, status, submitted_at, operator, upload_filename)
            VALUES (:sid, :tid, 'baidu', :account, :cookie,
                    :identity_type, :agent_name, '',
                    :complaint_category, :complaint_type, :module_name, :content_type,
                    :desc, :work_name, :rows, 200, :batches, 'queued', NOW(), :operator, :upload_filename)
        """), {
            'sid': submission_id,
            'tid': task_id,
            'account': collect_account,
            'cookie': cookie[:100] + '...',
            'identity_type': '代理人',
            'agent_name': collect_account,
            'complaint_category': '知识产权',
            'complaint_type': complaint_product,
            'module_name': complaint_product,
            'content_type': '版权',
            'desc': complaint_product,
            'work_name': work_names_str[:5000],
            'rows': total_links,
            'batches': total_batches,
            'operator': get_current_user(),
            'upload_filename': upload_filename,
        })

        batch_no = 0
        for work in works_config:
            links = work.get('links', [])
            for chunk_start in range(0, len(links), 200):
                batch_no += 1
                chunk_end = min(chunk_start + 200, len(links))
                session.execute(text("""
                    INSERT INTO complaint_batches
                    (batch_id, complaint_id, batch_no, work_name, batch_filename, start_row, end_row, row_count, status)
                    VALUES (:bid, :sid, :bno, :wname, :fname, :sr, :er, :rc, 'pending')
                """), {
                    'bid': uuid4().hex[:12],
                    'sid': submission_id,
                    'bno': batch_no,
                    'wname': work['work_name'],
                    'fname': f"{work['work_name']}_part{batch_no}",
                    'sr': chunk_start + 1,
                    'er': chunk_end,
                    'rc': chunk_end - chunk_start,
                })

        # 写入百度作品子表（可投诉的）
        work_idx = 0
        for work in works_config:
            work_links = work.get('links', [])
            session.execute(text("""
                INSERT INTO submission_works
                (complaint_id, work_index, work_name, platform_code, description,
                 actual_name, actual_url, link_count, batch_count, status)
                VALUES (:sid, :widx, :wname, 'baidu', :desc,
                        :aname, :aurl, :lcount, :bcount, 'pending')
            """), {
                'sid': submission_id,
                'widx': work_idx,
                'wname': work['work_name'],
                'desc': work.get('description', ''),
                'aname': work.get('actual_name', ''),
                'aurl': work.get('actual_url', ''),
                'lcount': len(work_links),
                'bcount': math.ceil(len(work_links) / 200),
            })
            work_idx += 1

        # 写入百度作品子表（跳过的，直接标记 failed）
        for sw in skipped_works:
            session.execute(text("""
                INSERT INTO submission_works
                (complaint_id, work_index, work_name, platform_code, description,
                 actual_name, actual_url, link_count, batch_count, status, error_message)
                VALUES (:sid, :widx, :wname, 'baidu', '',
                        '', '', 0, 0, 'skipped', :err)
            """), {
                'sid': submission_id,
                'widx': work_idx,
                'wname': sw['work_name'],
                'err': sw.get('reason', '未找到已通过审核的权属记录，请在百度投诉原平台进行投诉'),
            })
            work_idx += 1

        session.commit()
    except Exception as e:
        session.rollback()
        return jsonify({'success': False, 'error': f'数据库写入失败：{str(e)}'}), 500
    finally:
        session.close()

    # 入队（只包含可投诉的作品）
    if works_config:
        task_payload = {
            'task_id': task_id,
            'submission_id': submission_id,
            'cookie': cookie,
            'complaint_product': complaint_product,
            'complaint_type_code': complaint_type_code,
            'infringe_type': infringe_type,
            'works_config': works_config,
            'total_batches': total_batches,
        }
        enqueue_baidu_task(task_payload)

    tasks[task_id] = {
        'status': 'queued' if works_config else 'completed',
        'submitted_at': datetime.now().isoformat(),
        'total_batches': total_batches,
    }

    # 如果没有可投诉的作品，直接标记完成
    if not works_config:
        session = get_db_session()
        try:
            skipped_numbers = [f"未找到已通过审核的权属记录:{sw['work_name']}" for sw in skipped_works]
            session.execute(text("""
                UPDATE complaints SET status='completed', completed_at=NOW(),
                complaint_numbers_json=:nums WHERE task_id=:tid
            """), {'nums': json.dumps(skipped_numbers, ensure_ascii=False), 'tid': task_id})
            session.commit()
        except:
            session.rollback()
        finally:
            session.close()

    return jsonify({
        'success': True,
        'task_id': task_id,
        'submission_id': submission_id,
        'total_works': len(works_config),
        'skipped_works': len(skipped_works),
        'total_links': total_links,
        'total_batches': total_batches,
    })


@app.route('/api/baidu/task/<task_id>', methods=['GET'])
@login_required
def baidu_task_status(task_id):
    session = get_db_session()
    try:
        row = session.execute(text("""
            SELECT task_id, submission_id, status, current_batch, batch_count,
                   completed_batches, failed_batches, complaint_numbers_json,
                   error_message, submitted_at, started_at, completed_at
            FROM complaints WHERE task_id = :tid
        """), {'tid': task_id}).fetchone()
        if not row:
            mem = tasks.get(task_id)
            if mem:
                return jsonify({'success': True, 'task': mem})
            return jsonify({'success': False, 'error': '任务不存在'}), 404
        return jsonify({
            'success': True,
            'task': {
                'task_id': row.task_id,
                'submission_id': row.submission_id,
                'status': row.status,
                'current_batch': row.current_batch,
                'batch_count': row.batch_count,
                'completed_batches': row.completed_batches,
                'failed_batches': row.failed_batches,
                'complaint_numbers': json.loads(row.complaint_numbers_json) if row.complaint_numbers_json else [],
                'error_message': row.error_message,
                'submitted_at': normalize_datetime(row.submitted_at),
                'started_at': normalize_datetime(row.started_at),
                'completed_at': normalize_datetime(row.completed_at),
            }
        })
    finally:
        session.close()


@app.route('/api/baidu/status_list', methods=['GET'])
@login_required
def baidu_status_list():
    session = get_db_session()
    try:
        rows = session.execute(text("""
            SELECT complaint_id AS submission_id, collect_account, work_name, total_links AS excel_rows,
                   batch_count, submitted_at,
                   task_id, status, complaint_numbers_json, error_message,
                   completed_at, operator
            FROM complaints
            WHERE platform_code = 'baidu'
            ORDER BY submitted_at DESC
            LIMIT 50
        """)).fetchall()

        status_map = {
            'queued': '等待中',
            'running': '执行中',
            'completed': '已完成',
            'failed': '失败',
            'partial_failed': '部分失败',
        }

        result = []
        for row in rows:
            complaint_numbers = []
            if row.complaint_numbers_json:
                try:
                    complaint_numbers = json.loads(row.complaint_numbers_json)
                except:
                    pass
            result.append({
                'submission_id': row.submission_id,
                'task_id': row.task_id,
                'collect_account': row.collect_account,
                'work_name': row.work_name,
                'total_links': row.excel_rows,
                'batch_count': row.batch_count,
                'status': status_map.get(row.status, row.status or '等待中'),
                'complaint_numbers': complaint_numbers,
                'error_message': row.error_message,
                'submitted_at': normalize_datetime(row.submitted_at),
                'completed_at': normalize_datetime(row.completed_at),
                'operator': row.operator or '',
            })

        return jsonify({'success': True, 'data': result})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        session.close()


@app.route('/api/baidu/export_excel/<submission_id>', methods=['GET'])
@login_required
def baidu_export_excel(submission_id):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    session = get_db_session()
    try:
        # 获取主表信息
        sub = session.execute(text("""
            SELECT complaint_id AS submission_id, collect_account, submitted_at,
                   complaint_numbers_json
            FROM complaints
            WHERE complaint_id = :sid AND platform_code = 'baidu'
        """), {'sid': submission_id}).fetchone()
        if not sub:
            return jsonify({'success': False, 'error': '记录不存在'}), 404

        # 获取作品列表（包含跳过的）。feedback_numbers 为每作品自己的单号列表，
        # 一部作品多批次时会有多个单号——这是正确配对的权威来源。
        works = session.execute(text("""
            SELECT work_name, status, error_message, feedback_numbers FROM submission_works
            WHERE complaint_id = :sid ORDER BY work_index
        """), {'sid': submission_id}).fetchall()

        complaint_numbers = []
        if sub.complaint_numbers_json:
            try:
                complaint_numbers = json.loads(sub.complaint_numbers_json)
            except:
                pass

        # 提交时间格式化
        submitted_at = ''
        if sub.submitted_at:
            submitted_at = sub.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if hasattr(sub.submitted_at, 'strftime') else str(sub.submitted_at)

        # 生成 Excel
        wb = Workbook()
        ws = wb.active
        ws.title = '投诉结果'
        ws.append(['采集时间', '采集账号', '作品名称', '反馈单号'])
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # 优先用每作品自己的 feedback_numbers（一作品多批次 → 多行）。
        # 老记录该列为空，退回旧的「逐作品从扁平 complaint_numbers 取一个」兜底。
        has_per_work = any(getattr(w, 'feedback_numbers', None) for w in works)
        number_idx = 0
        for work in works:
            if work.status == 'skipped':
                ws.append([submitted_at, sub.collect_account, work.work_name,
                           work.error_message or '未找到已通过审核的权属记录，请在百度投诉原平台进行投诉'])
                continue

            if has_per_work:
                # 用本作品自己的单号列表，逐个单号一行
                nums = []
                raw = getattr(work, 'feedback_numbers', None)
                if raw:
                    try:
                        nums = json.loads(raw) if isinstance(raw, str) else list(raw)
                    except (TypeError, json.JSONDecodeError):
                        nums = []
                if not nums:
                    nums = ['']
                for fn in nums:
                    ws.append([submitted_at, sub.collect_account, work.work_name, str(fn)])
            else:
                # 兜底（老记录）：扁平列表按顺序取一个
                fn = complaint_numbers[number_idx] if number_idx < len(complaint_numbers) else ''
                ws.append([submitted_at, sub.collect_account, work.work_name, str(fn)])
                number_idx += 1

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 55

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f'baidu_result_{submission_id}.xlsx'
        return send_file(buf, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    finally:
        session.close()


def _recover_baidu_partial(task_id, submission_id, reason):
    """百度脚本被超时/异常杀掉时，从增量文件 task_results/<task_id>.json 回收
    已成功作品的反馈单号，写回 complaints + submission_works，状态记 partial_failed。
    返回回收到的单号个数（0 表示没有可回收的）。"""
    recovered = load_task_result(f'baidu_{submission_id}') or load_task_result(task_id)
    if not recovered:
        return 0
    numbers = recovered.get('feedback_numbers') or []
    by_work = recovered.get('feedback_numbers_by_work') or []
    real_numbers = [n for n in numbers if not (isinstance(n, str) and (n.startswith('未获取到单号:') or n.startswith('投诉失败:')))]
    if not real_numbers and not by_work:
        return 0
    session = get_db_session()
    try:
        session.execute(text("""
            UPDATE complaints SET status='partial_failed', completed_at=NOW(),
                complaint_numbers_json=:nums, error_message=:err
            WHERE task_id=:tid
        """), {
            'nums': json.dumps(numbers, ensure_ascii=False) if numbers else None,
            'err': reason,
            'tid': task_id,
        })
        for grp in by_work:
            session.execute(text("""
                UPDATE submission_works SET feedback_numbers=:nums
                WHERE complaint_id=:sid AND work_name=:wname
            """), {
                'nums': json.dumps(grp.get('numbers', []), ensure_ascii=False),
                'sid': submission_id,
                'wname': grp.get('work_name', ''),
            })
        session.commit()
    except Exception:
        session.rollback()
    finally:
        session.close()
    return len(real_numbers)


def run_baidu_complaint_script(task_id, cookie, complaint_product, complaint_type_code, works_config, total_batches, infringe_type=None):
    import sys

    script_path = os.path.join(os.path.dirname(__file__), 'baidu_complaint_backend.py')
    submission_id = task_id[6:] if task_id.startswith('baidu_') else task_id

    session = get_db_session()
    try:
        session.execute(text("""
            UPDATE complaints SET status='running', started_at=NOW() WHERE task_id=:tid
        """), {'tid': task_id})
        session.commit()
    except:
        session.rollback()
    finally:
        session.close()

    tasks[task_id] = tasks.get(task_id, {})
    tasks[task_id]['status'] = 'running'
    tasks[task_id]['started_at'] = datetime.now().isoformat()

    import tempfile
    works_config_file = tempfile.NamedTemporaryFile(
        mode='w', suffix='.json', delete=False, encoding='utf-8'
    )
    works_config_file.write(json.dumps(works_config, ensure_ascii=False))
    works_config_file.close()

    cmd = [
        sys.executable, script_path,
        '--task-id', task_id,
        '--cookie', cookie,
        '--complaint-type-code', str(complaint_type_code),
        '--works-config-file', works_config_file.name,
    ]
    if infringe_type is not None:
        cmd += ['--infringe-type', str(infringe_type)]

    timeout_seconds = max(120, total_batches * 30)

    try:
        proc = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=timeout_seconds,
            cwd=os.path.dirname(__file__),
        )
        try:
            os.unlink(works_config_file.name)
        except Exception:
            pass

        stdout = proc.stdout or ''
        stderr = proc.stderr or ''

        # 保存日志
        log_dir = app.config['TASK_RESULT_FOLDER']
        os.makedirs(log_dir, exist_ok=True)
        log_path = os.path.join(log_dir, f'{task_id}.log')
        with open(log_path, 'w', encoding='utf-8') as f:
            f.write(f"=== STDOUT ===\n{stdout}\n\n=== STDERR ===\n{stderr}\n")

        # 解析 JSON 结果
        result_data = None
        if 'JSON_RESULT_START' in stdout and 'JSON_RESULT_END' in stdout:
            json_str = stdout.split('JSON_RESULT_START')[1].split('JSON_RESULT_END')[0].strip()
            try:
                result_data = json.loads(json_str)
            except:
                pass

        session = get_db_session()
        try:
            if result_data:
                final_status = result_data.get('status', 'completed')
                complaint_numbers = result_data.get('feedback_numbers', [])
                error_msg = result_data.get('error_message', '')
                completed_batches = result_data.get('completed_batches', 0)
                failed_batches = result_data.get('failed_batches', 0)

                session.execute(text("""
                    UPDATE complaints
                    SET status=:st, completed_at=NOW(),
                        complaint_numbers_json=:nums,
                        completed_batches=:cb, failed_batches=:fb,
                        error_message=:err
                    WHERE task_id=:tid
                """), {
                    'st': final_status,
                    'nums': json.dumps(complaint_numbers, ensure_ascii=False) if complaint_numbers else None,
                    'cb': completed_batches,
                    'fb': failed_batches,
                    'err': error_msg or None,
                    'tid': task_id,
                })

                # 更新各批次状态
                batch_results = result_data.get('batch_results', [])
                for br in batch_results:
                    session.execute(text("""
                        UPDATE complaint_batches
                        SET status=:st, complaint_number=:cn, error_message=:err
                        WHERE complaint_id=:sid AND batch_no=:bno
                    """), {
                        'st': br.get('status', 'completed'),
                        'cn': br.get('feedback_number'),
                        'err': br.get('error'),
                        'sid': submission_id,
                        'bno': br.get('batch_no'),
                    })

                # 更新百度作品子表（权属详情 + 状态）
                works_detail = result_data.get('works_detail', [])
                for wd in works_detail:
                    wd_status = wd.get('status', 'completed')
                    works_category_name = BAIDU_WORKS_CATEGORY_MAP.get(wd.get('works_category'), '')
                    session.execute(text("""
                        UPDATE submission_works
                        SET cp_id=:cpid, owner_type=:ot, works_category=:wc,
                            works_category_name=:wcn, contact_name=:cn,
                            status=:st, error_message=:err
                        WHERE complaint_id=:sid AND work_index=:widx
                    """), {
                        'cpid': wd.get('cp_id', ''),
                        'ot': wd.get('owner_type') or 0,
                        'wc': wd.get('works_category') or 0,
                        'wcn': works_category_name,
                        'cn': wd.get('contact_name', ''),
                        'st': wd_status,
                        'err': wd.get('error'),
                        'sid': submission_id,
                        'widx': wd.get('work_index', 0),
                    })

                # 写入每作品的反馈单号分组（供导出时「一作品多批次=多单号」正确配对）。
                # 后端按 work_name 分组，这里按 work_name 匹配回各 submission_works 行。
                for grp in result_data.get('feedback_numbers_by_work', []):
                    session.execute(text("""
                        UPDATE submission_works
                        SET feedback_numbers=:nums
                        WHERE complaint_id=:sid AND work_name=:wname
                    """), {
                        'nums': json.dumps(grp.get('numbers', []), ensure_ascii=False),
                        'sid': submission_id,
                        'wname': grp.get('work_name', ''),
                    })
            else:
                session.execute(text("""
                    UPDATE complaints
                    SET status='failed', completed_at=NOW(),
                        error_message=:err
                    WHERE task_id=:tid
                """), {
                    'err': f'脚本执行异常，退出码：{proc.returncode}',
                    'tid': task_id,
                })

            session.commit()
        except:
            session.rollback()
        finally:
            session.close()

        if result_data:
            tasks[task_id]['status'] = result_data.get('status', 'completed')
            tasks[task_id]['feedback_numbers'] = result_data.get('feedback_numbers', [])
        else:
            tasks[task_id]['status'] = 'failed'
            tasks[task_id]['error'] = f'退出码：{proc.returncode}'

    except subprocess.TimeoutExpired:
        # 超时被杀：先尝试从增量文件回收已成功作品的单号
        recovered_n = _recover_baidu_partial(task_id, submission_id, '脚本执行超时')
        if recovered_n > 0:
            tasks[task_id]['status'] = 'partial_failed'
            tasks[task_id]['error'] = f'执行超时（已回收{recovered_n}个单号）'
        else:
            session = get_db_session()
            try:
                session.execute(text("""
                    UPDATE complaints SET status='failed', completed_at=NOW(),
                    error_message='脚本执行超时' WHERE task_id=:tid
                """), {'tid': task_id})
                session.commit()
            except:
                session.rollback()
            finally:
                session.close()
            tasks[task_id]['status'] = 'failed'
            tasks[task_id]['error'] = '执行超时'

    except Exception as e:
        session = get_db_session()
        try:
            session.execute(text("""
                UPDATE complaints SET status='failed', completed_at=NOW(),
                error_message=:err WHERE task_id=:tid
            """), {'err': str(e), 'tid': task_id})
            session.commit()
        except:
            session.rollback()
        finally:
            session.close()
        tasks[task_id]['status'] = 'failed'
        tasks[task_id]['error'] = str(e)


def run_quark_complaint_script(task_id, cookie, module, content_type, works_config, total_batches):
    import sys, tempfile
    script_path = os.path.join(os.path.dirname(__file__), 'quark_complaint_backend.py')
    submission_id = task_id[len('quark_'):] if task_id.startswith('quark_') else task_id

    db = get_db_session()
    try:
        db.execute(text("UPDATE complaints SET status='running', started_at=NOW() WHERE task_id=:tid"), {'tid': task_id})
        db.commit()
    except:
        db.rollback()
    finally:
        db.close()

    tasks[task_id] = tasks.get(task_id, {})
    tasks[task_id]['status'] = 'running'
    tasks[task_id]['started_at'] = datetime.now().isoformat()

    wc_file = tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False, encoding='utf-8')
    wc_file.write(json.dumps(works_config, ensure_ascii=False))
    wc_file.close()

    cmd = [
        sys.executable, script_path,
        '--task-id', task_id,
        '--cookie', cookie,
        '--works-config-file', wc_file.name,
        '--module', str(module),
        '--content-type', str(content_type),
    ]

    try:
        proc = subprocess.run(cmd, capture_output=True, text=True,
                              timeout=max(120, total_batches * 30),
                              cwd=os.path.dirname(__file__))
        try:
            os.unlink(wc_file.name)
        except Exception:
            pass

        stdout = proc.stdout or ''
        stderr = proc.stderr or ''

        log_dir = app.config['TASK_RESULT_FOLDER']
        os.makedirs(log_dir, exist_ok=True)
        with open(os.path.join(log_dir, f'{task_id}.log'), 'w', encoding='utf-8') as f:
            f.write(f"=== STDOUT ===\n{stdout}\n\n=== STDERR ===\n{stderr}\n")

        result_data = None
        if 'JSON_RESULT_START' in stdout and 'JSON_RESULT_END' in stdout:
            try:
                result_data = json.loads(stdout.split('JSON_RESULT_START')[1].split('JSON_RESULT_END')[0].strip())
            except Exception:
                pass

        db = get_db_session()
        try:
            if result_data:
                db.execute(text("""
                    UPDATE complaints
                    SET status=:st, completed_at=NOW(),
                        complaint_numbers_json=:nums,
                        completed_batches=:cb, failed_batches=:fb,
                        error_message=:err
                    WHERE task_id=:tid
                """), {
                    'st': result_data.get('status', 'completed'),
                    'nums': json.dumps(result_data.get('feedback_numbers', []), ensure_ascii=False),
                    'cb': result_data.get('completed_batches', 0),
                    'fb': result_data.get('failed_batches', 0),
                    'err': result_data.get('error_message') or None,
                    'tid': task_id,
                })
                for br in result_data.get('batch_results', []):
                    db.execute(text("""
                        UPDATE complaint_batches
                        SET status=:st, complaint_number=:cn, error_message=:err
                        WHERE complaint_id=:sid AND batch_no=:bno
                    """), {
                        'st': br.get('status', 'completed'),
                        'cn': br.get('feedback_number'),
                        'err': br.get('error'),
                        'sid': submission_id,
                        'bno': br.get('batch_no'),
                    })
                # 按作品写回单号分组（成功=单号，失败=「投诉失败: 原因」），
                # 供导出时一作品多批次正确配对、失败原因直接显示在单号列。
                for grp in result_data.get('feedback_numbers_by_work', []):
                    db.execute(text("""
                        UPDATE submission_works
                        SET feedback_numbers=:nums, status=:st
                        WHERE complaint_id=:sid AND work_name=:wname
                    """), {
                        'nums': json.dumps(grp.get('numbers', []), ensure_ascii=False),
                        'st': grp.get('status', 'completed'),
                        'sid': submission_id,
                        'wname': grp.get('work_name', ''),
                    })
                tasks[task_id]['status'] = result_data.get('status', 'completed')
            else:
                db.execute(text("UPDATE complaints SET status='failed', completed_at=NOW(), error_message=:err WHERE task_id=:tid"),
                           {'err': stderr[:500] or '脚本未返回结果', 'tid': task_id})
                tasks[task_id]['status'] = 'failed'
            db.commit()
        except Exception:
            db.rollback()
        finally:
            db.close()

    except subprocess.TimeoutExpired:
        try:
            os.unlink(wc_file.name)
        except Exception:
            pass
        db = get_db_session()
        try:
            db.execute(text("UPDATE complaints SET status='failed', completed_at=NOW(), error_message='脚本执行超时' WHERE task_id=:tid"), {'tid': task_id})
            db.commit()
        except:
            db.rollback()
        finally:
            db.close()
        tasks[task_id]['status'] = 'failed'
    except Exception as e:
        db = get_db_session()
        try:
            db.execute(text("UPDATE complaints SET status='failed', completed_at=NOW(), error_message=:err WHERE task_id=:tid"), {'err': str(e), 'tid': task_id})
            db.commit()
        except:
            db.rollback()
        finally:
            db.close()
        tasks[task_id]['status'] = 'failed'


# Blueprint 注册（必须在所有函数定义之后，避免循环引用）
from quark_routes import quark_bp
app.register_blueprint(quark_bp)


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5001, debug=False)
