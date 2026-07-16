# -*- coding: utf-8 -*-
"""腾讯文犀(ri.qq.com)投诉 Blueprint

与其他平台最大差异：**无 cookie，用 token 鉴权**。账号表 cookie_text 复用为 JSON blob：
  {"token":..., "sessionId":..., "uid":..., "subjectGroup":{账号固定主体}}

模板填中文名，upload_template 时用账号 token 实时拉 委托方下拉/产品列表，把中文名
转成 delegateCode / appId+appKey；权属证明沿用 static/imgs/剧名 目录匹配。
一作品一单，多链接合并（后端按 20 条/单自动拆多单）。
"""

import io
import json
import math
import os
import time as _time
from datetime import datetime
from functools import wraps
from uuid import uuid4

import requests
from flask import Blueprint, request, jsonify, send_file, session as flask_session, current_app
from openpyxl import load_workbook
from sqlalchemy import text

wenxi_bp = Blueprint('wenxi', __name__, url_prefix='/api/wenxi')

WENXI_API_BASE = 'https://ri.qq.com/api/v1'
LOGIN_EXPIRE_SECONDS = 43200
_UA = ('Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 '
       '(KHTML, like Gecko) Chrome/149.0.0.0 Safari/537.36')

# 内容类型（product_content_type）中文→code
WENXI_CONTENT_TYPE_MAP = {
    '视频': 1000, '音频': 1001, '图文': 1002, '其他': 1004,
}
# 权利类型（complaint_request_right_type）中文→code
WENXI_RIGHT_TYPE_MAP = {
    '著作权': 400, '名誉权': 401, '肖像权': 402, '隐私权': 403,
    '商誉权': 404, '商标权': 405, '其他': 406,
}
# 作品类型（right_work_type）中文→code；可留空
WENXI_WORK_TYPE_MAP = {
    '电影': 2300, '电视剧': 2301, '微短剧': 2310, '综艺': 2302, '动漫': 2303,
    '纪录片': 2304, '个创类短视频': 2308, '体育': 2305, '新闻': 2307,
    '漫剧': 2309, '其他': 2306,
}
# 代理机构全称 → 简称（与 app.py / weibo_routes 一致，授权委托书文件名用简称）
WENXI_AGENT_ORG_SHORT = {
    '北京和晞科技有限公司': '和晞科技',
    '北京柏蒙文化传媒有限公司': '柏蒙文化',
    '北京中惠信科科技有限公司': '中惠信科',
}


# ── 懒加载 app 模块符号（避免循环引用）────────────────────────────────────────

def _app():
    import app as _m
    return _m


def get_db_session():
    return _app().get_db_session()


def get_redis_client():
    return _app().get_redis_client()


def get_current_user():
    return flask_session.get('username', '')


def _tasks():
    return _app().tasks


def normalize_work_path_part(v):
    return (v or '').strip().replace('/', '_').replace('\\', '_')


def normalize_datetime(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value)


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = flask_session.get('token')
        if not token:
            return jsonify({'success': False, 'error': '未登录', 'login_required': True}), 401
        login_time = flask_session.get('login_time', 0)
        if _time.time() - login_time > LOGIN_EXPIRE_SECONDS:
            flask_session.clear()
            return jsonify({'success': False, 'error': '登录已过期，请重新登录', 'login_required': True}), 401
        return f(*args, **kwargs)
    return decorated


def enqueue_wenxi_task(payload: dict):
    payload['platform'] = 'wenxi'
    m = _app()
    get_redis_client().lpush(m.UNIFIED_QUEUE_NAME, json.dumps(payload, ensure_ascii=False))


# ── 文犀 token 鉴权工具 ──────────────────────────────────────────────────────

def _parse_auth_blob(cookie_text: str) -> dict:
    """账号 cookie_text 存的是 JSON blob，解析出 token/sessionId/uid/subjectGroup。

    兼容运营直接粘贴控制台输出的多种格式：
    1. 纯 JSON 文本：{"token":"...","uid":"..."}
    2. 控制台字面量（外层单引号 + \\n 转义）：'{\n  "token": "...",\n  ...}'
    3. 控制台字符串（外层单引号 + 真实换行）：'\n{...\n}'\n'
    """
    if not cookie_text:
        return {}
    text = cookie_text.strip()

    # 尝试1：直接 JSON 解析（用户粘贴纯 JSON，最理想情况）
    try:
        blob = json.loads(text)
        if isinstance(blob, dict):
            return blob
    except Exception:
        pass

    # 尝试2：ast.literal_eval 处理控制台字符串字面量
    # 识别外层单/双引号并处理 \n \t 等转义序列
    try:
        import ast
        inner = ast.literal_eval(text)
        if isinstance(inner, str):
            blob = json.loads(inner)
            if isinstance(blob, dict):
                return blob
    except Exception:
        pass

    # 尝试3：直接剥掉外层单/双引号后 JSON 解析
    # 处理：控制台输出带真实换行的单引号包裹字符串
    if len(text) >= 2 and text[0] in ('"', "'") and text[-1] == text[0]:
        try:
            blob = json.loads(text[1:-1])
            if isinstance(blob, dict):
                return blob
        except Exception:
            pass

    return {}


def _fetch_subject_group(auth: dict) -> dict:
    """从 /account/subject/auth/0/details 自动拉取账号自身主体（subjectGroup）。

    路径参数对服务端无意义，传 0 即可拿到当前登录账号的主体认证数据。
    返回组装好的 subjectGroup dict，失败返回空 dict。
    """
    try:
        resp = requests.get(f'{WENXI_API_BASE}/account/subject/auth/0/details',
                            headers=_wenxi_headers(auth), timeout=15)
        data = resp.json()
        if data.get('code') != 0 or not (data.get('data') or {}).get('group'):
            return {}
        g = data['data']['group']
        return {
            'isAgency': 1,
            'isTemp': 0,
            'agencyType': 801,
            'name': g.get('name', ''),
            'identityType': None,
            'identityNum': g.get('orgCode', ''),
            'contactNo': g.get('contactNumber', ''),
            'email': g.get('email', ''),
            'contactName': g.get('contactName') or g.get('name', ''),
            'subjectType': g.get('subjectType'),
            'id': g.get('id'),
            'orgType': g.get('orgType', 6600),
        }
    except Exception:
        return {}


def _wenxi_headers(auth: dict) -> dict:
    return {
        'Authorization': auth.get('token', ''),
        'sessionId': auth.get('sessionId', ''),
        'uid': auth.get('uid', ''),
        'User-Agent': _UA,
        'Origin': 'https://ri.qq.com',
        'Referer': 'https://ri.qq.com/initiate-complaint',
        'Accept': 'application/json, text/plain, */*',
    }


def _check_wenxi_login(auth: dict) -> bool:
    """/message/unread-total 返回 code:0 即有效。"""
    try:
        resp = requests.get(f'{WENXI_API_BASE}/message/unread-total',
                            headers=_wenxi_headers(auth), timeout=15)
        data = resp.json()
    except Exception:
        return False
    return data.get('code') == 0


def _fetch_delegate_droplist(auth: dict, acc_type: int = 801) -> dict:
    """委托方下拉：返回 {name: code}。"""
    resp = requests.get(f'{WENXI_API_BASE}/delegate/{acc_type}/drop-list',
                        headers=_wenxi_headers(auth), timeout=20)
    data = resp.json()
    if data.get('code') != 0:
        raise RuntimeError(f"获取委托方列表失败: {data.get('message', data)}")
    return {item.get('name', ''): item.get('code', '') for item in (data.get('data') or [])}


def _fetch_products(auth: dict) -> dict:
    """产品列表：返回 {name: {appId, appKey, contentType[]}}。"""
    resp = requests.get(f'{WENXI_API_BASE}/products/name',
                        headers=_wenxi_headers(auth), timeout=20)
    data = resp.json()
    if data.get('code') != 0:
        raise RuntimeError(f"获取产品列表失败: {data.get('message', data)}")
    out = {}
    for item in (data.get('data') or []):
        if item.get('isEnabled') != 1:
            continue
        out[item.get('name', '')] = {
            'appId': item.get('appId', ''),
            'appKey': item.get('appKey', ''),
            'contentType': item.get('contentType', []),
        }
    return out


# ── download_template ──────────────────────────────────────────────────────────

WENXI_SHEET1_FIELDS = [
    ('委托方组织/机构', '', '必填，须与文犀「委托管理」里的委托方全称一致（如 上海XX文化传播有限公司）'),
    ('代理机构名称', '', '必填，用于匹配授权委托书/营业执照目录（如 北京和晞科技有限公司）'),
    ('投诉产品', '', '必填，须与文犀投诉产品名一致，如：搜狗搜索/腾讯视频/腾讯新闻/企鹅号/微视/qq浏览器/腾讯体育/应用宝'),
    ('内容类型', '', '必填。可选：视频/音频/图文/其他（须为该产品支持的类型）'),
    ('权利类型', '', '必填。可选：著作权/名誉权/肖像权/隐私权/商誉权/商标权/其他'),
    ('作品类型', '', '可选。可选：电影/电视剧/微短剧/综艺/动漫/纪录片/个创类短视频/体育/新闻/漫剧/其他'),
    ('投诉描述', '', '必填，投诉理由描述'),
]


@wenxi_bp.route('/download_template', methods=['GET'])
@login_required
def wenxi_download_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()

    # Sheet1 表单内容
    ws1 = wb.active
    ws1.title = '表单内容'
    ws1.append(['字段', '值', '说明'])
    for cell in ws1[1]:
        cell.font = Font(bold=True)
    for name, default, remark in WENXI_SHEET1_FIELDS:
        ws1.append([name, default, remark])
    ws1.column_dimensions['A'].width = 24
    ws1.column_dimensions['B'].width = 40
    ws1.column_dimensions['C'].width = 70

    # Sheet2 批量导入链接（侵权链接 | 作品名称 | 首发地址）
    ws2 = wb.create_sheet('批量导入Excel')
    ws2.append(['侵权链接', '作品名称', '首发地址'])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for _ in range(3):
        ws2.append(['', '', ''])
    ws2.column_dimensions['A'].width = 60
    ws2.column_dimensions['B'].width = 30
    ws2.column_dimensions['C'].width = 60

    # Sheet3 填写说明
    ws3 = wb.create_sheet('填写说明')
    for line in [
        ['腾讯文犀版权投诉模版（仅支持机构代理场景）'],
        [''],
        ['Sheet1 表单内容（填中文名称，系统自动转编码）'],
        ['  委托方组织/机构：须与文犀「委托管理」里已通过的委托方全称完全一致'],
        ['  投诉产品：须与文犀投诉产品名一致（搜狗搜索/腾讯视频/腾讯新闻/企鹅号/微视/qq浏览器/腾讯体育/应用宝等）'],
        ['  内容类型：视频/音频/图文/其他（须为该产品支持的类型）'],
        ['  权利类型：著作权/名誉权/肖像权/隐私权/商誉权/商标权/其他'],
        ['  作品类型：可选，电影/电视剧/微短剧/综艺/动漫/纪录片/个创类短视频/体育/新闻/漫剧/其他'],
        [''],
        ['Sheet2 批量导入Excel'],
        ['  侵权链接：必填。同一作品链接超过 20 条时自动拆成多单提交'],
        ['  作品名称：必填，支持多部作品混合，系统按作品名分组'],
        ['  首发地址：必填，作品的原始/首发链接（每行填，同作品可相同）'],
        [''],
        ['证明文件说明（沿用其他平台约定，放 static/imgs/ 下）'],
        ['  权属证明：static/imgs/剧名/<作品目录>/ 下「证明文件_*」'],
        ['  授权委托书由文犀「委托管理」维护，本系统提交时自动带上，无需本地准备'],
    ]:
        ws3.append(line)
    ws3.column_dimensions['A'].width = 90

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='wenxi_template.xlsx',
    )


# ── 文件匹配工具（沿用其他平台）─────────────────────────────────────────────

def _paren(s):
    return (s or '').replace('（', '(').replace('）', ')')


def _match_file(dir_path, prefix, *must_contain):
    """在目录里找 prefix 开头、且包含所有 must_contain 片段的文件，返回绝对路径或空。"""
    if not os.path.isdir(dir_path):
        return ''
    for f in sorted(os.listdir(dir_path)):
        if f.startswith('._') or not f.startswith(prefix):
            continue
        if all(_paren(seg) in _paren(f) for seg in must_contain if seg):
            return os.path.join(dir_path, f)
    return ''


def _get_account_auth(collect_account):
    """按采集账号取 cookie_text 并解析出 auth blob。返回 (auth_dict, error_str)。

    若 subjectGroup 缺失，自动调 /account/subject/auth/0/details 补全，
    无需运营手动配置。
    """
    _s = get_db_session()
    try:
        acc_row = _s.execute(text("""
            SELECT cookie_text FROM accounts
            WHERE platform_code='wenxi' AND account_user=:acc LIMIT 1
        """), {'acc': collect_account}).fetchone()
    finally:
        _s.close()
    if not acc_row or not acc_row.cookie_text:
        return {}, '未找到该采集账号或账号未配置 token'
    auth = _parse_auth_blob(acc_row.cookie_text)
    if not auth.get('token'):
        return {}, 'token 解析失败，请在账号管理中重新粘贴文犀登录信息（JSON）'
    # subjectGroup 缺失时自动拉取（账号固定主体，拉一次即可）
    if not auth.get('subjectGroup'):
        sg = _fetch_subject_group(auth)
        if sg:
            auth['subjectGroup'] = sg
    return auth, ''


@wenxi_bp.route('/upload_template', methods=['POST'])
@login_required
def wenxi_upload_template():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '未上传文件'}), 400
    file = request.files['file']
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ('.xlsx', '.xls'):
        return jsonify({'success': False, 'error': '仅支持 .xlsx / .xls 格式'}), 400

    collect_account = request.form.get('collect_account', '').strip()
    if not collect_account:
        return jsonify({'success': False, 'error': '请先选择采集账号，再上传模版'}), 400

    # 取账号 token（用于实时拉委托方/产品列表并校验登录态）
    auth, auth_err = _get_account_auth(collect_account)
    if auth_err:
        return jsonify({'success': False, 'error': auth_err}), 400
    if not _check_wenxi_login(auth):
        return jsonify({'success': False, 'error': 'token 已失效，请在账号管理中更新文犀登录信息'}), 401

    try:
        wb = load_workbook(file, data_only=True)
    except Exception as e:
        return jsonify({'success': False, 'error': f'文件解析失败：{e}'}), 400
    if '表单内容' not in wb.sheetnames or '批量导入Excel' not in wb.sheetnames:
        return jsonify({'success': False, 'error': '模版缺少"表单内容"或"批量导入Excel"工作表'}), 400

    # 解析 Sheet1
    cfg = {}
    for row in wb['表单内容'].iter_rows(min_row=2, max_col=2, values_only=True):
        if row[0] and row[1] is not None and str(row[1]).strip() != '':
            cfg[str(row[0]).strip()] = str(row[1]).strip()

    def _g(key):
        return cfg.get(key, '').strip()

    delegate_name = _g('委托方组织/机构')
    agent_org = _g('代理机构名称')
    product_name = _g('投诉产品')
    content_type_name = _g('内容类型')
    right_type_name = _g('权利类型')
    work_type_name = _g('作品类型')       # 可选
    description = _g('投诉描述')

    required = {
        '委托方组织/机构': delegate_name, '代理机构名称': agent_org,
        '投诉产品': product_name, '内容类型': content_type_name,
        '权利类型': right_type_name, '投诉描述': description,
    }
    missing = [k for k, v in required.items() if not v]
    if missing:
        return jsonify({'success': False, 'error': 'Sheet1 缺少必填项：' + '、'.join(missing)}), 400

    # 中文名 → code（内容/权利/作品类型）
    content_type = WENXI_CONTENT_TYPE_MAP.get(content_type_name)
    if content_type is None:
        return jsonify({'success': False, 'error': f'内容类型「{content_type_name}」无效，可选：'
                        + '、'.join(WENXI_CONTENT_TYPE_MAP.keys())}), 400
    right_type = WENXI_RIGHT_TYPE_MAP.get(right_type_name)
    if right_type is None:
        return jsonify({'success': False, 'error': f'权利类型「{right_type_name}」无效，可选：'
                        + '、'.join(WENXI_RIGHT_TYPE_MAP.keys())}), 400
    work_type = None
    if work_type_name:
        work_type = WENXI_WORK_TYPE_MAP.get(work_type_name)
        if work_type is None:
            return jsonify({'success': False, 'error': f'作品类型「{work_type_name}」无效，可选：'
                            + '、'.join(WENXI_WORK_TYPE_MAP.keys())}), 400

    # 实时拉委托方下拉 + 产品列表，解析名称
    try:
        delegate_map = _fetch_delegate_droplist(auth, 801)
        product_map = _fetch_products(auth)
    except Exception as e:
        return jsonify({'success': False, 'error': f'拉取文犀基础数据失败：{e}'}), 502

    delegate_code = delegate_map.get(delegate_name)
    if not delegate_code:
        return jsonify({'success': False, 'error': f'委托方「{delegate_name}」不在该账号的委托列表中，'
                        f'可选：{("、".join(k for k in delegate_map if k)) or "（空）"}'}), 400

    product = product_map.get(product_name)
    if not product:
        return jsonify({'success': False, 'error': f'投诉产品「{product_name}」无效或未启用，'
                        f'可选：{("、".join(k for k in product_map if k)) or "（空）"}'}), 400
    # 注意：products/name 返回的 contentType 字段是产品自身的分类代码（如 [0]），
    # 不是投诉内容类型的约束范围，不做校验。所有产品均支持全部投诉内容类型。

    meta = {
        'appId': product['appId'],
        'appName': product_name,
        'appKey': product['appKey'],
        'contentType': content_type,
        'rightType': right_type,
        'workType': work_type,
        'description': description,
        'agencyType': 801,
    }

    # 解析 Sheet2（侵权链接 | 作品名称 | 首发地址）
    works_map, work_order, empty_rows = {}, [], 0
    for row in wb['批量导入Excel'].iter_rows(min_row=2, max_col=3, values_only=True):
        link = str(row[0]).strip() if row[0] else ''
        wn = str(row[1]).strip() if row[1] else ''
        origin = str(row[2]).strip() if len(row) > 2 and row[2] else ''
        if not link and not wn:
            empty_rows += 1
            if empty_rows >= 5:
                break
            continue
        empty_rows = 0
        if not link:
            return jsonify({'success': False, 'error': f'存在作品名但侵权链接为空（作品：{wn}）'}), 400
        if not wn:
            return jsonify({'success': False, 'error': f'存在链接但作品名为空（链接：{link[:60]}）'}), 400
        if wn not in works_map:
            works_map[wn] = {'links': [], 'origin_url': origin}
            work_order.append(wn)
        works_map[wn]['links'].append(link)
        if origin and not works_map[wn]['origin_url']:
            works_map[wn]['origin_url'] = origin

    if not works_map:
        return jsonify({'success': False, 'error': '"批量导入Excel"中没有有效数据'}), 400
    missing_origin = [wn for wn in work_order if not works_map[wn]['origin_url']]
    if missing_origin:
        return jsonify({'success': False, 'error': '以下作品缺少首发地址：\n'
                        + '\n'.join(f'  · {w}' for w in missing_origin)}), 400

    # 逐作品匹配权属证明（沿用其他平台：works 表拿 used_company/content_type/complaint_type 拼目录）
    static_imgs_dir = os.path.join(current_app.root_path, 'static', 'imgs')
    works_base_dir = os.path.join(static_imgs_dir, '剧名')

    works_config, match_errors = [], []
    for wn in work_order:
        db = get_db_session()
        try:
            rows = db.execute(text("""
                SELECT w.used_company,
                       ct.dict_name AS content_type, cpt.dict_name AS complaint_type
                FROM works w
                JOIN dictionaries ct  ON ct.dict_type='content_type'   AND ct.dict_code=CAST(w.content_type_id  AS CHAR)
                JOIN dictionaries cpt ON cpt.dict_type='complaint_type' AND cpt.dict_code=CAST(w.complaint_type_id AS CHAR)
                WHERE w.work_name=:wn
            """), {'wn': wn}).mappings().all()
        finally:
            db.close()

        if not rows:
            match_errors.append(f'「{wn}」在作品覆盖列表中不存在')
            continue
        r0 = rows[0]
        uc, ct, cpt = r0.get('used_company') or '', r0.get('content_type', ''), r0.get('complaint_type', '')
        dir_name = f"{normalize_work_path_part(wn)}_{normalize_work_path_part(uc)}_{normalize_work_path_part(ct)}_{normalize_work_path_part(cpt)}"
        drama_dir = os.path.join(works_base_dir, dir_name)
        proof_path = _match_file(drama_dir, '证明文件_')
        if not proof_path:
            match_errors.append(f'「{wn}」缺少权属证明（目录 {dir_name}/证明文件_*）')
            continue

        works_config.append({
            'work_name': wn,
            'links': works_map[wn]['links'],
            'origin_url': works_map[wn]['origin_url'],
            'proof_path': proof_path,
        })

    if not works_config:
        return jsonify({'success': False, 'error': '所有作品匹配失败：\n' + '\n'.join(match_errors)}), 400

    total_links = sum(len(w['links']) for w in works_config)
    total_batches = sum(math.ceil(len(w['links']) / 20) for w in works_config)

    resp_data = {
        'success': True,
        'filename': file.filename,
        'meta': meta,
        'delegate_code': delegate_code,
        'works': works_config,
        'total_links': total_links,
        'total_batches': total_batches,
        'delegate_name': delegate_name,
        'agent_org': agent_org,
        'product_name': product_name,
        'content_type_name': content_type_name,
        'right_type_name': right_type_name,
        'work_type_name': work_type_name,
        'description': description,
    }
    if match_errors:
        resp_data['warnings'] = match_errors
    return jsonify(resp_data)


# ── verify_cookie（校验账号 token）───────────────────────────────────────────

@wenxi_bp.route('/verify_cookie', methods=['POST'])
@login_required
def wenxi_verify_cookie():
    data = request.get_json() or {}
    collect_account = data.get('collect_account', '').strip()
    if not collect_account:
        return jsonify({'success': False, 'error': '请选择采集账号'}), 400
    auth, auth_err = _get_account_auth(collect_account)
    if auth_err:
        return jsonify({'success': False, 'error': auth_err}), 400
    try:
        if _check_wenxi_login(auth):
            return jsonify({'success': True})
        return jsonify({'success': False, 'error': 'token 已失效，请在账号管理中更新文犀登录信息'}), 401
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ── submit ─────────────────────────────────────────────────────────────────────

@wenxi_bp.route('/submit', methods=['POST'])
@login_required
def wenxi_submit():
    data = request.get_json() or {}
    collect_account = data.get('collect_account', '').strip()
    meta = data.get('meta', {})
    delegate_code = data.get('delegate_code', '').strip()
    works_config = data.get('works', [])
    upload_filename = data.get('upload_filename', '').strip()

    if not collect_account:
        return jsonify({'success': False, 'error': '请选择投诉账号'}), 400
    if not works_config:
        return jsonify({'success': False, 'error': '作品列表不能为空'}), 400
    if not delegate_code:
        return jsonify({'success': False, 'error': '缺少委托方，请重新上传模版'}), 400

    # 取账号 auth + 固化的 subjectGroup
    auth, auth_err = _get_account_auth(collect_account)
    if auth_err:
        return jsonify({'success': False, 'error': auth_err}), 400
    subject_group = auth.get('subjectGroup')
    if not subject_group:
        return jsonify({'success': False, 'error': '该账号未固化 subjectGroup（账号自身主体），'
                        '请在账号管理中补充一次真实提交的 subjectGroup'}), 400
    if not _check_wenxi_login(auth):
        return jsonify({'success': False, 'error': 'token 已失效，请在账号管理中更新文犀登录信息'}), 401

    # 防重复：同账号同文件未失败的记录已存在则拒绝
    if upload_filename:
        _s = get_db_session()
        try:
            dup = _s.execute(text("""
                SELECT task_id FROM complaints
                WHERE collect_account=:acc AND upload_filename=:fn
                  AND platform_code='wenxi' AND status NOT IN ('failed')
                LIMIT 1
            """), {'acc': collect_account, 'fn': upload_filename}).fetchone()
        finally:
            _s.close()
        if dup:
            return jsonify({'success': False, 'error': f'文件「{upload_filename}」已投诉过（任务 {dup[0]}），请勿重复提交'}), 400

    total_links = sum(len(w.get('links', [])) for w in works_config)
    total_batches = sum(math.ceil(len(w.get('links', [])) / 20) for w in works_config)
    all_work_names = [w['work_name'] for w in works_config]

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    submission_id = f"{timestamp}_{uuid4().hex[:8]}"
    task_id = f'wenxi_{submission_id}'

    delegate_name = data.get('delegate_name', '')
    agent_org = data.get('agent_org', '')

    db = get_db_session()
    try:
        submitted_at = datetime.now()
        estimated_finish_at = _app().compute_estimated_finish(db, total_batches, 'wenxi', submitted_at)
        db.execute(text("""
            INSERT INTO complaints
            (complaint_id, task_id, platform_code, collect_account, cookie_snapshot,
             identity_type, agent_name, principal_name,
             complaint_category, complaint_type, module_name, content_type,
             description_text, work_name, total_links, batch_size, batch_count,
             status, submitted_at, estimated_finish_at, operator, upload_filename)
            VALUES (:sid, :tid, 'wenxi', :account, :cookie,
                    '机构代理', :agent, :principal,
                    :rtype, :rtype, :product, :ctype,
                    :desc, :work_name, :rows, 20, :batches,
                    'queued', :submitted_at, :estimated_finish_at, :operator, :upload_filename)
        """), {
            'sid': submission_id,
            'tid': task_id,
            'account': collect_account,
            'cookie': (auth.get('token', '')[:60] + '...'),
            'agent': agent_org,
            'principal': delegate_name,
            'rtype': data.get('right_type_name', ''),
            'product': data.get('product_name', ''),
            'ctype': data.get('content_type_name', ''),
            'desc': meta.get('description', ''),
            'work_name': ', '.join(all_work_names)[:5000],
            'rows': total_links,
            'batches': total_batches,
            'submitted_at': submitted_at,
            'estimated_finish_at': estimated_finish_at,
            'operator': get_current_user(),
            'upload_filename': upload_filename,
        })

        batch_no = 0
        for work in works_config:
            links = work.get('links', [])
            for chunk_start in range(0, len(links), 20):
                batch_no += 1
                chunk_end = min(chunk_start + 20, len(links))
                db.execute(text("""
                    INSERT INTO complaint_batches
                    (batch_id, complaint_id, batch_no, work_name, batch_filename,
                     start_row, end_row, row_count, status)
                    VALUES (:bid, :sid, :bno, :wname, :fname, :sr, :er, :rc, 'pending')
                """), {
                    'bid': uuid4().hex[:12],
                    'sid': submission_id,
                    'bno': batch_no,
                    'wname': work['work_name'],
                    'fname': f"{work['work_name']}_part{batch_no}",
                    'sr': chunk_start + 1,
                    'er': chunk_end,
                    'rc': chunk_end - chunk_start,
                })

        for idx, work in enumerate(works_config):
            db.execute(text("""
                INSERT INTO submission_works
                (complaint_id, work_index, work_name, platform_code,
                 link_count, batch_count, status)
                VALUES (:sid, :widx, :wname, 'wenxi', :lcount, :bcount, 'pending')
            """), {
                'sid': submission_id,
                'widx': idx,
                'wname': work['work_name'],
                'lcount': len(work.get('links', [])),
                'bcount': math.ceil(len(work.get('links', [])) / 20),
            })

        db.commit()
    except Exception as e:
        db.rollback()
        return jsonify({'success': False, 'error': f'数据库写入失败：{e}'}), 500
    finally:
        db.close()

    enqueue_wenxi_task({
        'task_id': task_id,
        'submission_id': submission_id,
        'auth': auth,
        'meta': meta,
        'subject_group': subject_group,
        'delegate_code': delegate_code,
        'works_config': works_config,
        'total_batches': total_batches,
    })

    _tasks()[task_id] = {'status': 'queued', 'submitted_at': datetime.now().isoformat()}
    return jsonify({'success': True, 'task_id': task_id, 'submission_id': submission_id})


# ── status list ───────────────────────────────────────────────────────────────

@wenxi_bp.route('/status_list', methods=['GET'])
@login_required
def wenxi_status_list():
    db = get_db_session()
    try:
        rows = db.execute(text("""
            SELECT complaint_id AS submission_id, task_id, collect_account, work_name,
                   total_links, batch_count, submitted_at, estimated_finish_at, status,
                   complaint_numbers_json, error_message, operator
            FROM complaints
            WHERE platform_code = 'wenxi'
            ORDER BY submitted_at DESC
            LIMIT 50
        """)).fetchall()
        status_map = {
            'queued': '等待中', 'running': '执行中', 'completed': '已完成',
            'failed': '失败', 'partial_failed': '部分失败',
        }
        result = []
        for row in rows:
            complaint_numbers = []
            if row.complaint_numbers_json:
                try:
                    complaint_numbers = json.loads(row.complaint_numbers_json)
                except Exception:
                    pass
            result.append({
                'submission_id': row.submission_id,
                'task_id': row.task_id,
                'collect_account': row.collect_account,
                'work_name': row.work_name,
                'total_links': row.total_links,
                'batch_count': row.batch_count,
                'submitted_at': normalize_datetime(row.submitted_at),
                'estimated_finish_at': normalize_datetime(row.estimated_finish_at),
                'status': status_map.get(row.status, row.status or '等待中'),
                'complaint_numbers': complaint_numbers,
                'operator': row.operator or '',
            })
        return jsonify({'success': True, 'data': result})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        db.close()


# ── export excel ──────────────────────────────────────────────────────────────

@wenxi_bp.route('/export_excel/<submission_id>', methods=['GET'])
@login_required
def wenxi_export_excel(submission_id):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    db = get_db_session()
    try:
        sub = db.execute(text("""
            SELECT complaint_id, collect_account, submitted_at, complaint_numbers_json
            FROM complaints WHERE complaint_id = :sid AND platform_code = 'wenxi'
        """), {'sid': submission_id}).fetchone()
        if not sub:
            return jsonify({'success': False, 'error': '记录不存在'}), 404

        works = db.execute(text("""
            SELECT work_name, status, feedback_numbers FROM submission_works
            WHERE complaint_id = :sid ORDER BY work_index
        """), {'sid': submission_id}).fetchall()

        complaint_numbers = []
        if sub.complaint_numbers_json:
            try:
                complaint_numbers = json.loads(sub.complaint_numbers_json)
            except Exception:
                pass

        submitted_at = ''
        if sub.submitted_at:
            submitted_at = sub.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if hasattr(sub.submitted_at, 'strftime') else str(sub.submitted_at)

        wb = Workbook()
        ws = wb.active
        ws.title = '投诉结果'
        ws.append(['采集时间', '采集账号', '作品名称', '投诉单号'])
        for cell in ws[1]:
            cell.font = Font(bold=True)

        has_per_work = any(getattr(w, 'feedback_numbers', None) for w in works)
        number_idx = 0
        for work in works:
            if has_per_work:
                nums = []
                raw = getattr(work, 'feedback_numbers', None)
                if raw:
                    try:
                        nums = json.loads(raw) if isinstance(raw, str) else list(raw)
                    except (TypeError, json.JSONDecodeError):
                        nums = []
                if not nums:
                    nums = ['']
                for num in nums:
                    ws.append([submitted_at, sub.collect_account, work.work_name, str(num)])
            else:
                num = complaint_numbers[number_idx] if number_idx < len(complaint_numbers) else ''
                number_idx += 1
                ws.append([submitted_at, sub.collect_account, work.work_name, str(num)])

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 30

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True,
                         download_name=f'wenxi_{submission_id}.xlsx')
    finally:
        db.close()


# ── task status ────────────────────────────────────────────────────────────────

@wenxi_bp.route('/task/<task_id>', methods=['GET'])
@login_required
def wenxi_task_status(task_id):
    db = get_db_session()
    try:
        row = db.execute(text("""
            SELECT task_id, status, batch_count, completed_batches, failed_batches,
                   complaint_numbers_json, error_message,
                   submitted_at, started_at, completed_at
            FROM complaints WHERE task_id = :tid
        """), {'tid': task_id}).fetchone()
        if not row:
            mem = _tasks().get(task_id)
            if mem:
                return jsonify({'success': True, 'task': mem})
            return jsonify({'success': False, 'error': '任务不存在'}), 404
        return jsonify({
            'success': True,
            'task': {
                'task_id': row.task_id,
                'status': row.status,
                'batch_count': row.batch_count,
                'completed_batches': row.completed_batches,
                'failed_batches': row.failed_batches,
                'complaint_numbers': json.loads(row.complaint_numbers_json) if row.complaint_numbers_json else [],
                'error_message': row.error_message,
                'submitted_at': normalize_datetime(row.submitted_at),
                'started_at': normalize_datetime(row.started_at),
                'completed_at': normalize_datetime(row.completed_at),
            }
        })
    finally:
        db.close()
