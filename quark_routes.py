# -*- coding: utf-8 -*-
"""夸克投诉 Blueprint"""

import json
import math
import os
import io
import re
import time as _time
from datetime import datetime
from functools import wraps
from uuid import uuid4

import requests
from flask import Blueprint, request, jsonify, send_file, session as flask_session, current_app
from openpyxl import Workbook, load_workbook
from sqlalchemy import text

quark_bp = Blueprint('quark', __name__, url_prefix='/api/quark')


# ── 懒加载 app 模块符号（避免循环引用）────────────────────────────────────────

def _app():
    import app as _m
    return _m


def get_db_session():
    return _app().get_db_session()


def get_redis_client():
    return _app().get_redis_client()


def get_current_user():
    return flask_session.get('username', '')


def normalize_work_path_part(v):
    return (v or '').strip().replace('/', '_').replace('\\', '_')


def normalize_datetime(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value)


LOGIN_EXPIRE_SECONDS = 43200


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = flask_session.get('token')
        if not token:
            return jsonify({'success': False, 'error': '未登录', 'login_required': True}), 401
        login_time = flask_session.get('login_time', 0)
        if _time.time() - login_time > LOGIN_EXPIRE_SECONDS:
            flask_session.clear()
            return jsonify({'success': False, 'error': '登录已过期，请重新登录', 'login_required': True}), 401
        return f(*args, **kwargs)
    return decorated


def enqueue_quark_task(payload: dict):
    payload['platform'] = 'quark'
    m = _app()
    get_redis_client().lpush(m.UNIFIED_QUEUE_NAME, json.dumps(payload, ensure_ascii=False))


def _tasks():
    return _app().tasks

QUARK_API_BASE = 'https://ipp.quark.cn'
QUARK_MODULE_MAP = {3: '夸克网盘', 8: '夸克搜索', 1: '夸克图片', 4: '夸克日报'}
QUARK_CONTENT_TYPE_MAP = {1: '影视剧集', 2: '其他视频', 3: '小说', 4: '漫画', 5: '图片', 7: '文章', 8: '软件/游戏', 6: '其他'}


def _quark_headers(cookie: str) -> dict:
    xtstk = ''
    for part in cookie.split(';'):
        p = part.strip()
        if p.startswith('cmptstk='):
            xtstk = p[len('cmptstk='):]
            break
    return {
        'Cookie': cookie,
        'x-requested-with': 'XMLHttpRequest',
        'xtstk': xtstk,
        'User-Agent': 'Mozilla/5.0',
        'Referer': 'https://ipp.quark.cn/',
    }




# ── verify_cookie ──────────────────────────────────────────────────────────────

@quark_bp.route('/verify_cookie', methods=['POST'])
@login_required
def quark_verify_cookie():
    data = request.get_json() or {}
    cookie = data.get('cookie', '').strip()
    if not cookie:
        return jsonify({'success': False, 'error': 'Cookie不能为空'}), 400
    try:
        resp = requests.get(
            f'{QUARK_API_BASE}/api/complain/accuse',
            headers=_quark_headers(cookie),
            timeout=10,
        )
        result = resp.json()
        if result.get('code') == 200:
            return jsonify({'success': True})
        return jsonify({'success': False, 'error': f'Cookie无效，code={result.get("code")}'}), 401
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ── download_template ──────────────────────────────────────────────────────────

@quark_bp.route('/download_template', methods=['GET'])
@login_required
def quark_download_template():
    import pandas as pd

    sheet1_data = {
        '字段': ['您的身份', '代理人/权利人', '被代理人（权利人）信息', '投诉大类', '投诉类型', '功能模块', '内容类型', '投诉内容描述'],
        '值':   ['代理人', '北京和晞科技有限公司', '', '', '', '', '', ''],
        '可选值(备注)': [
            '权利人、代理人',
            '北京和晞科技有限公司',
            '填写被代理人名称，须与夸克账号已登记的被代理人名称完全一致',
            '知识产权、人身权',
            '知识产权→著作权(含视频、图文、图集等)/商标/专利/其他知识产权；人身权→名誉/商誉权/姓名/名称权/肖像权/隐私权/其他人身权益',
            '夸克网盘、夸克图片、夸克日报、夸克搜索',
            '影视剧集、其他视频、小说、漫画、图片、文章、软件/游戏、其他',
            '可在描述中写 ${work_title}，投诉每部作品时会自动替换成该作品名',
        ],
    }
    df1 = pd.DataFrame(sheet1_data)

    sheet2_headers = ['侵权链接', '对应原创链接/对应访问码', '作品名称']
    df2 = pd.DataFrame([['', '', ''], ['', '', ''], ['', '', '']], columns=sheet2_headers)

    sheet3_lines = [
        ['Sheet1 表单填写说明'], [''],
        ['字段名', '填写说明'],
        ['您的身份', '必填，权利人 或 代理人，默认代理人'],
        ['代理人/权利人', '必填，选择代理人名称'],
        ['被代理人（权利人）信息', '代理人身份时必填，与夸克账号已登记的被代理人名称完全一致'],
        ['功能模块', '必填，夸克网盘/夸克图片/夸克日报/夸克搜索'],
        ['内容类型', '必填，影视剧集/其他视频/小说/漫画/图片/文章/软件/游戏/其他'],
        ['投诉大类', '必填，知识产权 或 人身权'],
        ['投诉类型', '必填，知识产权→著作权(含视频、图文、图集等)/商标/专利/其他知识产权；人身权→名誉/商誉权/姓名/名称权/肖像权/隐私权/其他人身权益'],
        ['投诉内容描述', '必填，可用 ${work_title} 作为作品名占位符'],
        [''],
        ['Sheet2 批量导入Excel填写说明'], [''],
        ['字段名', '填写说明'],
        ['侵权链接', '必填'],
        ['对应原创链接/对应访问码', '选填'],
        ['作品名称', '必填，支持多部作品混合填写，系统按作品名自动分组'],
        [''],
        ['证明文件说明'], [''],
        ['证明文件', '系统根据作品名在 static/imgs/剧名/ 下匹配「证明文件_*」'],
        ['其他证明', '系统自动匹配授权委托书、营业执照（被代理人）、营业执照（代理人）'],
    ]
    df3 = pd.DataFrame(sheet3_lines)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df1.to_excel(writer, sheet_name='表单内容', index=False)
        df2.to_excel(writer, sheet_name='批量导入Excel', index=False)
        df3.to_excel(writer, sheet_name='填写说明', index=False, header=False)
    buf.seek(0)
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='quark_template.xlsx',
    )


# ── upload_template ────────────────────────────────────────────────────────────

QUARK_MODULE_NAME_MAP = {'夸克网盘': 3, '夸克搜索': 8, '夸克图片': 1, '夸克日报': 4}
QUARK_CONTENT_TYPE_NAME_MAP = {'影视剧集': 1, '其他视频': 2, '小说': 3, '漫画': 4, '图片': 5, '文章': 7, '软件/游戏': 8, '其他': 6}
QUARK_TYPE_MAP = {'知识产权': 9, '人身权': 10}
QUARK_SUB_TYPE_MAP = {
    '著作权(含视频、图文、图集等)': 11, '商标': 12, '专利': 13, '其他知识产权': 14,
    '名誉/商誉权': 15, '姓名/名称权': 16, '肖像权': 17, '隐私权': 18, '其他人身权益': 19,
}


@quark_bp.route('/upload_template', methods=['POST'])
@login_required
def quark_upload_template():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '未上传文件'}), 400
    file = request.files['file']
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ('.xlsx', '.xls'):
        return jsonify({'success': False, 'error': '仅支持 .xlsx / .xls 格式'}), 400

    collect_account = request.form.get('collect_account', '').strip()

    try:
        wb = load_workbook(file, data_only=True)
    except Exception as e:
        return jsonify({'success': False, 'error': f'文件解析失败：{e}'}), 400

    if '表单内容' not in wb.sheetnames:
        return jsonify({'success': False, 'error': '缺少"表单内容"工作表，请使用正确模版'}), 400
    if '批量导入Excel' not in wb.sheetnames:
        return jsonify({'success': False, 'error': '缺少"批量导入Excel"工作表，请使用正确模版'}), 400

    # 解析 Sheet1
    config = {}
    for row in wb['表单内容'].iter_rows(min_row=2, max_col=2, values_only=True):
        if row[0] and row[1] is not None:
            config[str(row[0]).strip()] = str(row[1]).strip()

    module_name = config.get('功能模块', '').strip()
    content_type_name = config.get('内容类型', '').strip()
    type_name = config.get('投诉大类', '').strip()
    sub_type_name = config.get('投诉类型', '').strip()
    description_tpl = config.get('投诉内容描述', '').strip()
    identity = config.get('您的身份', '代理人').strip()
    agent_name = config.get('代理人/权利人', '').strip()
    proxy_name = config.get('被代理人（权利人）信息', '').strip()

    if not agent_name:
        return jsonify({'success': False, 'error': '"代理人/权利人"不能为空'}), 400
    if identity == '代理人' and not proxy_name:
        return jsonify({'success': False, 'error': '代理人身份时"被代理人（权利人）信息"不能为空'}), 400
    if not module_name:
        return jsonify({'success': False, 'error': '"功能模块"不能为空'}), 400
    if not content_type_name:
        return jsonify({'success': False, 'error': '"内容类型"不能为空'}), 400
    if not type_name:
        return jsonify({'success': False, 'error': '"投诉大类"不能为空'}), 400
    if not sub_type_name:
        return jsonify({'success': False, 'error': '"投诉类型"不能为空'}), 400
    if not description_tpl:
        return jsonify({'success': False, 'error': '"投诉内容描述"不能为空'}), 400

    module = QUARK_MODULE_NAME_MAP.get(module_name)
    if module is None:
        return jsonify({'success': False, 'error': f'功能模块「{module_name}」不支持，可选：夸克网盘、夸克图片、夸克日报、夸克搜索'}), 400
    content_type = QUARK_CONTENT_TYPE_NAME_MAP.get(content_type_name)
    if content_type is None:
        return jsonify({'success': False, 'error': f'内容类型「{content_type_name}」不支持，可选：' + '、'.join(QUARK_CONTENT_TYPE_NAME_MAP.keys())}), 400
    complaint_type = QUARK_TYPE_MAP.get(type_name)
    if complaint_type is None:
        return jsonify({'success': False, 'error': f'投诉大类「{type_name}」不支持，可选：知识产权、人身权'}), 400
    complaint_sub_type = QUARK_SUB_TYPE_MAP.get(sub_type_name)
    if complaint_sub_type is None:
        return jsonify({'success': False, 'error': f'投诉类型「{sub_type_name}」不支持，请查看填写说明'}), 400
    if not description_tpl:
        return jsonify({'success': False, 'error': '投诉内容描述不能为空'}), 400

    # 解析 Sheet2（3列：侵权链接 | 对应原创链接/对应访问码 | 作品名称）
    works_map = {}   # work_name -> {'links': [], 'originals': []}
    work_order = []
    empty_rows = 0
    for row in wb['批量导入Excel'].iter_rows(min_row=2, max_col=3, values_only=True):
        link      = str(row[0]).strip() if row[0] else ''
        original  = str(row[1]).strip() if row[1] else ''
        work_name = str(row[2]).strip() if row[2] else ''
        if not link and not work_name:
            empty_rows += 1
            if empty_rows >= 5:
                break
            continue
        empty_rows = 0
        if not link:
            return jsonify({'success': False, 'error': f'存在作品名但侵权链接为空（作品：{work_name}）'}), 400
        if not work_name:
            return jsonify({'success': False, 'error': f'存在链接但作品名为空（链接：{link[:60]}）'}), 400
        if work_name not in works_map:
            works_map[work_name] = {'links': [], 'originals': []}
            work_order.append(work_name)
        works_map[work_name]['links'].append(link)
        works_map[work_name]['originals'].append(original)

    if not works_map:
        return jsonify({'success': False, 'error': '"批量导入Excel"中没有有效数据'}), 400

    # 获取 used_company（代理人所属公司）
    used_company = ''
    if collect_account:
        db = get_db_session()
        try:
            row = db.execute(text("""
                SELECT used_company FROM accounts
                WHERE platform_code='quark' AND account_user=:acc LIMIT 1
            """), {'acc': collect_account}).fetchone()
            if row:
                used_company = (row.used_company or '').strip()
        finally:
            db.close()

    static_imgs_dir = os.path.join(current_app.root_path, 'static', 'imgs')
    works_base_dir  = os.path.join(static_imgs_dir, '剧名')

    def _paren(s):
        return (s or '').replace('（', '(').replace('）', ')')

    def _company_match(name, filename):
        return _paren(name) in _paren(filename)

    works_config = []
    match_errors = []

    for wn in work_order:
        links = works_map[wn]['links']
        originals = works_map[wn]['originals']
        description = description_tpl.replace('${work_title}', wn)
        pn = proxy_name  # 全局被代理人，来自 Sheet1

        # 查 works 表，拿 content_type + complaint_type + used_company
        db = get_db_session()
        try:
            rows = db.execute(text("""
                SELECT w.work_name, w.used_company, w.principal_name,
                       ct.dict_name AS content_type, cpt.dict_name AS complaint_type
                FROM works w
                JOIN dictionaries ct  ON ct.dict_type='content_type'  AND ct.dict_code=CAST(w.content_type_id  AS CHAR)
                JOIN dictionaries cpt ON cpt.dict_type='complaint_type' AND cpt.dict_code=CAST(w.complaint_type_id AS CHAR)
                WHERE w.work_name=:wn
            """), {'wn': wn}).mappings().all()
        finally:
            db.close()

        if not rows:
            match_errors.append(f'「{wn}」在作品覆盖列表中不存在')
            continue

        # 按 principal_name 匹配被代理人
        matched = [r for r in rows if _paren(r.get('principal_name') or '') == _paren(pn)]
        if not matched:
            match_errors.append(f'「{wn}」的被代理人「{pn}」不匹配')
            continue

        # 按 used_company 缩小
        if used_company:
            narrowed = [r for r in matched if (r.get('used_company') or '') == used_company]
            if narrowed:
                matched = narrowed

        row0 = matched[0]
        uc   = row0.get('used_company') or used_company
        ct   = row0.get('content_type', '')
        cpt  = row0.get('complaint_type', '')
        dir_name = f"{normalize_work_path_part(wn)}_{normalize_work_path_part(uc)}_{normalize_work_path_part(ct)}_{normalize_work_path_part(cpt)}"
        drama_dir = os.path.join(works_base_dir, dir_name)

        if not os.path.isdir(drama_dir):
            match_errors.append(f'「{wn}」作品目录不存在：{dir_name}')
            continue

        # 证明文件（必须）
        proof_path = ''
        other_paths = []
        for f in os.listdir(drama_dir):
            if f.startswith('证明文件_') and not f.startswith('._'):
                proof_path = os.path.join(drama_dir, f)
                break
        for f in sorted(os.listdir(drama_dir)):
            if f.startswith('其他证明_') and not f.startswith('._'):
                other_paths.append(os.path.join(drama_dir, f))

        # 授权委托书
        auth_dir = os.path.join(static_imgs_dir, '授权委托书')
        if os.path.isdir(auth_dir):
            for f in os.listdir(auth_dir):
                if f.startswith('授权委托书_') and not f.startswith('._') and _company_match(pn, f):
                    other_paths.append(os.path.join(auth_dir, f))
                    break

        # 营业执照(被代理人)
        biz_dir = os.path.join(static_imgs_dir, '营业执照')
        if os.path.isdir(biz_dir):
            for f in os.listdir(biz_dir):
                if f.startswith('营业执照_') and not f.startswith('._') and _company_match(pn, f):
                    other_paths.append(os.path.join(biz_dir, f))
                    break

        # 营业执照(代理人)
        if uc and os.path.isdir(biz_dir):
            for f in os.listdir(biz_dir):
                if f.startswith('营业执照_') and not f.startswith('._') and _company_match(uc, f):
                    other_paths.append(os.path.join(biz_dir, f))
                    break

        if not proof_path:
            match_errors.append(f'「{wn}」缺少证明文件')
            continue

        works_config.append({
            'work_name':          wn,
            'proxy_name':         pn,
            'description':        description,
            'links':              links,
            'originals':          originals,
            'proof_path':         proof_path,
            'other_paths':        other_paths,
            'complaint_type':     complaint_type,
            'complaint_sub_type': complaint_sub_type,
        })

    if match_errors and not works_config:
        return jsonify({'success': False, 'error': '所有作品匹配失败：\n' + '\n'.join(match_errors)}), 400

    total_links   = sum(len(w['links']) for w in works_config)
    total_batches = sum(math.ceil(len(w['links']) / 200) for w in works_config)

    resp_data = {
        'success':              True,
        'filename':             file.filename,
        'works':                works_config,
        'total_links':          total_links,
        'total_batches':        total_batches,
        'module':               module,
        'content_type':         content_type,
        'module_name':          module_name,
        'content_type_name':    content_type_name,
        'type_name':            type_name,
        'sub_type_name':        sub_type_name,
        'identity':             identity,
        'agent_name':           agent_name,
        'proxy_name':           proxy_name,
    }
    if match_errors:
        resp_data['warnings'] = match_errors
    return jsonify(resp_data)


# ── submit ─────────────────────────────────────────────────────────────────────

@quark_bp.route('/submit', methods=['POST'])
@login_required
def quark_submit():
    data = request.get_json() or {}
    cookie = data.get('cookie', '').strip()
    collect_account = data.get('collect_account', '').strip()
    module = data.get('module', 3)
    content_type = data.get('content_type', 6)
    works_config = data.get('works', [])
    upload_filename = data.get('upload_filename', '').strip()

    if not cookie:
        return jsonify({'success': False, 'error': 'Cookie不能为空'}), 400
    if not collect_account:
        return jsonify({'success': False, 'error': '请选择投诉账号'}), 400
    if not works_config:
        return jsonify({'success': False, 'error': '作品列表不能为空'}), 400

    # 防重复
    if upload_filename:
        _s = get_db_session()
        try:
            dup = _s.execute(text("""
                SELECT task_id FROM complaints
                WHERE collect_account=:acc AND upload_filename=:fn
                  AND platform_code='quark' AND status NOT IN ('failed')
                LIMIT 1
            """), {'acc': collect_account, 'fn': upload_filename}).fetchone()
        finally:
            _s.close()
        if dup:
            return jsonify({'success': False, 'error': f'文件「{upload_filename}」已投诉过（任务 {dup[0]}），请勿重复提交'}), 400

    # 验证 Cookie
    try:
        resp = requests.get(
            f'{QUARK_API_BASE}/api/complain/accuse',
            headers=_quark_headers(cookie),
            timeout=10,
        )
        if resp.json().get('code') != 200:
            return jsonify({'success': False, 'error': 'Cookie已失效，请更新后重试'}), 401
    except Exception as e:
        return jsonify({'success': False, 'error': f'Cookie验证失败：{e}'}), 500

    total_links = sum(len(w.get('links', [])) for w in works_config)
    total_batches = sum(math.ceil(len(w.get('links', [])) / 200) for w in works_config)
    all_work_names = [w['work_name'] for w in works_config]

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    submission_id = f"{timestamp}_{uuid4().hex[:8]}"
    task_id = f'quark_{submission_id}'

    # 写入数据库
    db = get_db_session()
    try:
        db.execute(text("""
            INSERT INTO complaints
            (complaint_id, task_id, platform_code, collect_account, cookie_snapshot,
             identity_type, agent_name, principal_name,
             complaint_category, complaint_type, module_name, content_type,
             description_text, work_name, total_links, batch_size, batch_count,
             status, submitted_at, operator, upload_filename)
            VALUES (:sid, :tid, 'quark', :account, :cookie,
                    '代理人', :account, '',
                    '知识产权', :module_name, :module_name, :content_type_name,
                    :module_name, :work_name, :rows, 200, :batches,
                    'queued', NOW(), :operator, :upload_filename)
        """), {
            'sid': submission_id,
            'tid': task_id,
            'account': collect_account,
            'cookie': cookie[:100] + '...',
            'module_name': QUARK_MODULE_MAP.get(module, f'module={module}'),
            'content_type_name': QUARK_CONTENT_TYPE_MAP.get(content_type, f'type={content_type}'),
            'work_name': ', '.join(all_work_names)[:5000],
            'rows': total_links,
            'batches': total_batches,
            'operator': get_current_user(),
            'upload_filename': upload_filename,
        })

        batch_no = 0
        for work in works_config:
            links = work.get('links', [])
            for chunk_start in range(0, len(links), 200):
                batch_no += 1
                chunk_end = min(chunk_start + 200, len(links))
                db.execute(text("""
                    INSERT INTO complaint_batches
                    (batch_id, complaint_id, batch_no, work_name, batch_filename,
                     start_row, end_row, row_count, status)
                    VALUES (:bid, :sid, :bno, :wname, :fname, :sr, :er, :rc, 'pending')
                """), {
                    'bid': uuid4().hex[:12],
                    'sid': submission_id,
                    'bno': batch_no,
                    'wname': work['work_name'],
                    'fname': f"{work['work_name']}_part{batch_no}",
                    'sr': chunk_start + 1,
                    'er': chunk_end,
                    'rc': chunk_end - chunk_start,
                })

        for idx, work in enumerate(works_config):
            db.execute(text("""
                INSERT INTO submission_works
                (complaint_id, work_index, work_name, platform_code,
                 link_count, batch_count, status)
                VALUES (:sid, :widx, :wname, 'quark', :lcount, :bcount, 'pending')
            """), {
                'sid': submission_id,
                'widx': idx,
                'wname': work['work_name'],
                'lcount': len(work.get('links', [])),
                'bcount': math.ceil(len(work.get('links', [])) / 200),
            })

        db.commit()
    except Exception as e:
        db.rollback()
        return jsonify({'success': False, 'error': f'数据库写入失败：{e}'}), 500
    finally:
        db.close()

    enqueue_quark_task({
        'task_id': task_id,
        'submission_id': submission_id,
        'cookie': cookie,
        'module': module,
        'content_type': content_type,
        'works_config': works_config,
        'total_batches': total_batches,
    })

    _tasks()[task_id] = {'status': 'queued', 'submitted_at': datetime.now().isoformat()}

    return jsonify({'success': True, 'task_id': task_id, 'submission_id': submission_id})


# ── task status ────────────────────────────────────────────────────────────────

@quark_bp.route('/task/<task_id>', methods=['GET'])
@login_required
def quark_task_status(task_id):
    db = get_db_session()
    try:
        row = db.execute(text("""
            SELECT task_id, status, batch_count, completed_batches, failed_batches,
                   complaint_numbers_json, error_message,
                   submitted_at, started_at, completed_at
            FROM complaints WHERE task_id = :tid
        """), {'tid': task_id}).fetchone()
        if not row:
            mem = _tasks().get(task_id)
            if mem:
                return jsonify({'success': True, 'task': mem})
            return jsonify({'success': False, 'error': '任务不存在'}), 404
        return jsonify({
            'success': True,
            'task': {
                'task_id': row.task_id,
                'status': row.status,
                'batch_count': row.batch_count,
                'completed_batches': row.completed_batches,
                'failed_batches': row.failed_batches,
                'complaint_numbers': json.loads(row.complaint_numbers_json) if row.complaint_numbers_json else [],
                'error_message': row.error_message,
                'submitted_at': normalize_datetime(row.submitted_at),
                'started_at': normalize_datetime(row.started_at),
                'completed_at': normalize_datetime(row.completed_at),
            }
        })
    finally:
        db.close()
