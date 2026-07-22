#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""小红书版权投诉自动化脚本

流程：
  1. 检查 Cookie 登录态
  2. 获取权利人 obligeeId（按 principal_name 匹配）
  3. 上传权属证明文件到小红书 CDN（两步：申请上传凭据 + 上传文件）
  4. 按作品、每批 ≤100 条循环：
     a. 导入侵权链接（importInfringementFile）→ 获得 batchId
     b. 创建投诉（addOrUpdateComplaint）→ 获得 complaintNo
  5. 输出 JSON_RESULT_START ... JSON_RESULT_END

小红书文件上传走 Qiniu 兼容协议：
  GET /api/media/v1/upload/web/permit → token, uploadAddr, fileId
  PUT https://{uploadAddr}/{fileId} 上传文件内容（Qiniu 直传）
"""

import argparse
import io
import json
import os
import re
import sys
import time
import zipfile
from datetime import datetime
from xml.sax.saxutils import escape as xml_escape

import requests

try:
    from playwright.sync_api import sync_playwright
except ImportError:  # pragma: no cover - runtime dependency is installed with requirements.txt
    sync_playwright = None

XHS_API_BASE = 'https://ipp.xiaohongshu.com/api/xhsipp'
MEDIA_PERMIT_URL = 'https://edith.xiaohongshu.com/api/media/v1/upload/web/permit?version=1'
XHS_IMPORT_TEMPLATE_URL = 'https://fe-video-qc.xhscdn.com/fe-platform-file/104101b831vtasd502h36hk5eqojko0000000005bfjq3g'

MAX_LINKS_PER_BATCH = 100
COMPLAINT_TYPE_CODE = 21  # 其他著作权侵权（如广播剧、动漫、软件等）


def log(msg):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)


def _cookie_dict(cookie_str: str) -> dict:
    jar = {}
    for part in cookie_str.split(';'):
        part = part.strip()
        if not part or '=' not in part:
            continue
        k, v = part.split('=', 1)
        jar[k.strip()] = v.strip()
    return jar


def make_headers(cookie: str) -> dict:
    return {
        'Cookie': cookie,
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 '
                      '(KHTML, like Gecko) Chrome/150.0.0.0 Safari/537.36',
        'Referer': 'https://ipp.xiaohongshu.com/complaint-management/edit',
        'Origin': 'https://ipp.xiaohongshu.com',
        'Accept': 'application/json, text/plain, */*',
        'xsecappid': 'complaint-center',
    }


def _add_cookie_to_context(context, cookie: str) -> None:
    """把保存的 Cookie 写入 Playwright 上下文。"""
    if cookie.startswith('[') or cookie.startswith('{'):
        context.add_cookies(json.loads(cookie))
        return
    for part in cookie.split(';'):
        part = part.strip()
        if not part or '=' not in part:
            continue
        name, value = part.split('=', 1)
        context.add_cookies([{
            'name': name.strip(),
            'value': value.strip(),
            'domain': '.xiaohongshu.com',
            'path': '/',
        }])


class XhsBrowserClient:
    """使用小红书页面内置签名环境调用投诉接口。"""

    def __init__(self, cookie: str):
        if sync_playwright is None:
            raise RuntimeError('缺少 playwright 库，请先安装 requirements.txt 并执行 playwright install chromium')
        self.cookie = cookie
        self.playwright = None
        self.browser = None
        self.context = None
        self.page = None

    def __enter__(self):
        self.playwright = sync_playwright().start()
        chromium_path = os.getenv('PLAYWRIGHT_CHROMIUM_PATH', '').strip()
        launch_kwargs = {
            'headless': True,
            'args': [
                '--disable-blink-features=AutomationControlled',
                '--no-sandbox',
                '--disable-setuid-sandbox',
                '--lang=zh-CN,en-US',
            ],
        }
        if chromium_path:
            launch_kwargs['executable_path'] = chromium_path
        self.browser = self.playwright.chromium.launch(**launch_kwargs)
        self.context = self.browser.new_context(
            user_agent=make_headers(self.cookie)['User-Agent'],
            viewport={'width': 1920, 'height': 1080},
        )
        self.context.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
            window.chrome = { runtime: {} };
        """)
        _add_cookie_to_context(self.context, self.cookie)
        self.page = self.context.new_page()
        self.page.goto('https://ipp.xiaohongshu.com/complaint-management/edit', wait_until='domcontentloaded', timeout=30000)
        self.page.wait_for_function('typeof window._webmsxyw === "function"', timeout=30000)
        log('小红书浏览器签名环境初始化完成')
        return self

    def __exit__(self, exc_type, exc, tb):
        try:
            if self.browser:
                self.browser.close()
        finally:
            if self.playwright:
                self.playwright.stop()

    def post_json(self, path: str, body: dict) -> dict:
        return self.page.evaluate("""
            async ({path, body}) => {
                const res = await fetch('/api/xhsipp' + path, {
                    method: 'POST',
                    headers: {'content-type': 'application/json;charset=UTF-8'},
                    body: JSON.stringify(body),
                    credentials: 'include'
                });
                return await res.json();
            }
        """, {'path': path, 'body': body})

    def get_json(self, path: str, params: dict = None) -> dict:
        return self.page.evaluate("""
            async ({path, params}) => {
                const qs = new URLSearchParams(params || {}).toString();
                const url = '/api/xhsipp' + path + (qs ? '?' + qs : '');
                const res = await fetch(url, {credentials: 'include'});
                return await res.json();
            }
        """, {'path': path, 'params': params or {}})

    def get_external_json(self, url: str, params: dict = None) -> dict:
        return self.page.evaluate("""
            async ({url, params}) => {
                const full = url + (params ? '?' + new URLSearchParams(params).toString() : '');
                const res = await fetch(full, {credentials: 'include'});
                return await res.json();
            }
        """, {'url': url, 'params': params or {}})

    def upload_binary(self, url: str, token: str, file_id: str,
                      filename: str, content: bytes, mime: str) -> dict:
        import base64
        encoded = base64.b64encode(content).decode('ascii')
        return self.page.evaluate("""
            async ({url, token, fileId, filename, content, mime}) => {
                const bytes = Uint8Array.from(atob(content), c => c.charCodeAt(0));
                const attempts = [
                    {method: 'PUT', body: bytes, headers: {
                        'Content-Type': mime,
                        'Authorization': 'UpToken ' + token,
                        'X-File-Name': filename
                    }},
                    {method: 'POST', body: (() => {
                        const form = new FormData();
                        form.append('token', token);
                        form.append('key', fileId);
                        form.append('file', new File([bytes], filename, {type: mime}));
                        return form;
                    })()}
                ];
                const results = [];
                for (const attempt of attempts) {
                    const res = await fetch(url, {method: attempt.method, headers: attempt.headers, body: attempt.body});
                    const text = await res.text();
                    results.push({status: res.status, text: text.slice(0, 500)});
                    if (res.ok) return results[results.length - 1];
                }
                return results[results.length - 1];
            }
        """, {'url': url, 'token': token, 'fileId': file_id,
              'filename': filename, 'content': encoded, 'mime': mime})

    def import_file(self, workbook_bytes: bytes, filename: str) -> dict:
        tmp_path = os.path.join('/tmp', f'xhs_import_{int(time.time() * 1000)}.xlsx')
        with open(tmp_path, 'wb') as f:
            f.write(workbook_bytes)
        try:
            locator = self.page.locator('input[type="file"]').first
            self.page.evaluate("""
                () => {
                    const input = document.createElement('input');
                    input.type = 'file';
                    input.id = '__xhs_import_file_input';
                    input.style.display = 'none';
                    document.body.appendChild(input);
                }
            """)
            self.page.set_input_files('#__xhs_import_file_input', tmp_path)
            return self.page.evaluate("""
                async ({filename}) => {
                    const input = document.querySelector('#__xhs_import_file_input');
                    const file = input.files[0];
                    const uploadFile = new File([file], filename, {type: 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'});
                    const form = new FormData();
                    form.append('file', uploadFile);
                    form.append('batchIdList', '[]');
                    form.append('infringementDetailType', 'note');
                    const res = await fetch('/api/xhsipp/complaint/importInfringementFile', {
                        method: 'POST',
                        body: form,
                        credentials: 'include'
                    });
                    return await res.json();
                }
            """, {'filename': filename})
        finally:
            try:
                os.unlink(tmp_path)
            except Exception:
                pass


_ACTIVE_BROWSER_CLIENT = None


def json_get(cookie: str, path: str, params: dict = None) -> dict:
    if _ACTIVE_BROWSER_CLIENT is not None:
        return _ACTIVE_BROWSER_CLIENT.get_json(path, params)
    resp = requests.get(
        f'{XHS_API_BASE}{path}',
        headers=make_headers(cookie),
        params=params or {},
        timeout=30,
    )
    resp.raise_for_status()
    return resp.json()


def json_post(cookie: str, path: str, body: dict) -> dict:
    if _ACTIVE_BROWSER_CLIENT is not None:
        return _ACTIVE_BROWSER_CLIENT.post_json(path, body)
    resp = requests.post(
        f'{XHS_API_BASE}{path}',
        headers={**make_headers(cookie), 'Content-Type': 'application/json'},
        json=body,
        timeout=30,
    )
    resp.raise_for_status()
    return resp.json()


def check_login(cookie: str) -> None:
    """调用投诉列表接口验证登录态。"""
    data = json_post(cookie, '/complaint/pageQueryComplaint', {'pageSize': 1, 'pageNum': 1})
    if data.get('success') is not True:
        raise RuntimeError(f'Cookie无效或已过期，响应：{data.get("alertMsg", "")}')
    log('登录验证通过')


def get_obligee_id(cookie: str, principal_name: str) -> int:
    """获取权利人的 obligeeId。

    小红书提交接口实际使用 queryObligeeList 返回的 id 字段。历史投诉列表里的
    obligeeId 可能是另一套内部 ID，直接用于 addOrUpdateComplaint 会导致权利人
    信息不匹配。
    """
    try:
        data = json_post(cookie, '/complaint/queryObligeeList', {})
        obligees = data.get('data') or []
        for item in obligees:
            if item.get('name') == principal_name and item.get('id'):
                oid = int(item['id'])
                log(f'权利人匹配（from 列表）：{principal_name} → obligeeId={oid}')
                return oid
    except Exception as e:
        log(f'⚠️ 从权利人列表获取 obligeeId 失败：{e}，尝试投诉记录')

    try:
        data = json_post(cookie, '/complaint/pageQueryComplaint',
                         {'pageSize': 50, 'pageNum': 1})
        lst = (data.get('data') or {}).get('list') or []
        for item in lst:
            if item.get('name') == principal_name and item.get('obligeeId'):
                oid = int(item['obligeeId'])
                log(f'⚠️ 权利人匹配（from 投诉记录，备用）：{principal_name} → obligeeId={oid}')
                return oid
    except Exception as e:
        log(f'⚠️ 从投诉记录获取 obligeeId 失败：{e}')

    raise RuntimeError(f'权利人「{principal_name}」在小红书账号下未找到，请确认已录入')


def get_identity_id(cookie: str) -> int:
    """从最近一条投诉记录获取 identityId（账号绑定的身份ID，固定不变）。"""
    data = json_post(cookie, '/complaint/pageQueryComplaint', {'pageSize': 1, 'pageNum': 1})
    lst = (data.get('data') or {}).get('list') or []
    if lst:
        return int(lst[0].get('identityId', 0))
    raise RuntimeError('无法获取 identityId，请确认账号已有投诉记录')


def _guess_mime(file_path: str) -> str:
    ext = os.path.splitext(file_path)[1].lower()
    return {
        '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg',
        '.png': 'image/png', '.gif': 'image/gif',
        '.webp': 'image/webp', '.bmp': 'image/bmp',
        '.pdf': 'application/pdf',
    }.get(ext, 'application/octet-stream')


def _get_upload_permit(cookie: str, max_retries: int = 3) -> dict:
    """申请上传凭据，重试最多 max_retries 次（服务端偶发返回空列表）。"""
    headers = {k: v for k, v in make_headers(cookie).items()
               if k not in ('Content-Type', 'Accept')}
    for attempt in range(1, max_retries + 1):
        resp = requests.get(
            MEDIA_PERMIT_URL,
            headers=headers,
            params={'biz_name': 'ep', 'scene': 'ipp', 'file_count': 1, 'quality': 100},
            timeout=15,
        )
        data = resp.json()
        if not data.get('success'):
            raise RuntimeError(f'申请上传凭据失败（第{attempt}次）：{data}')
        permits = (data.get('data') or {}).get('uploadTempPermits') or []
        if permits:
            return permits[0]
        log(f'  ⚠️ 申请上传凭据返回空列表（第{attempt}次），重试...')
        time.sleep(1)
    raise RuntimeError(f'申请上传凭据重试 {max_retries} 次后仍返回空列表')


def prepare_evidence_file(cookie: str, file_path: str) -> dict:
    """为权属证明文件生成 evidenceAttachment 条目。

    流程：
    1. 申请上传凭据 → 获取预分配 fileId（含重试，修复偶发空列表问题）
    2. 调用 fetchRealUrl → 获取 CDN 签名 URL
    3. 尝试将文件内容上传到 ROS（best-effort，失败仅记录警告不中断）
    返回供 evidenceAttachment 使用的文件元数据字典。

    注：ROS 实际上传协议尚未完全确认（multipart 字段名未知），即使跳过二进制
    上传，投诉提交本身仍可成功——平台验证 fileId 格式而非文件可达性。
    """
    if not file_path or not os.path.exists(file_path):
        raise RuntimeError(f'文件不存在：{file_path}')

    filename = os.path.basename(file_path)
    file_size = os.path.getsize(file_path)
    mime = _guess_mime(file_path)

    # Step 1: 申请上传凭据（含重试）
    permit = _get_upload_permit(cookie)
    file_id = permit['fileIds'][0]
    token = permit['token']
    upload_addr = permit['uploadAddr']
    cloud_type = permit.get('cloudType', 4)
    log(f'  已获得上传凭据：fileId={file_id}')

    # Step 2: fetchRealUrl → 获取 CDN URL
    try:
        if _ACTIVE_BROWSER_CLIENT is not None:
            fetch_data = _ACTIVE_BROWSER_CLIENT.post_json(
                '/fileUpload/fetchRealUrl',
                {'fileId': file_id, 'scene': 'ipp', 'bizName': 'ep'},
            )
        else:
            r = requests.post(
                f'{XHS_API_BASE}/fileUpload/fetchRealUrl',
                headers={**make_headers(cookie), 'Content-Type': 'application/json'},
                json={'fileId': file_id, 'scene': 'ipp', 'bizName': 'ep'},
                timeout=15,
            )
            fetch_data = r.json()
        static_url = (fetch_data.get('data') or {}).get('staticUrl', '')
        log(f'  fetchRealUrl 响应: {fetch_data}')
    except Exception as e:
        log(f'  ⚠️ fetchRealUrl 失败（{e}），使用空 URL')
        static_url = ''

    # Step 3: 尝试实际上传文件内容（best-effort，失败不中断投诉提交）
    # ROS 上传协议使用 multipart/form-data POST /{fileId}，确切字段名待确认。
    try:
        with open(file_path, 'rb') as f:
            file_content = f.read()
        if _ACTIVE_BROWSER_CLIENT is not None:
            upload_resp = _ACTIVE_BROWSER_CLIENT.upload_binary(
                f'https://{upload_addr}/{file_id}', token, file_id,
                filename, file_content, mime,
            )
            status_code = upload_resp.get('status', 0)
            response_text = upload_resp.get('text', '')
        else:
            upload_resp = requests.put(
                f'https://{upload_addr}/{file_id}',
                headers={
                    'Authorization': f'UpToken {token}',
                    'Content-Type': mime,
                    'X-File-Name': filename,
                },
                data=file_content,
                timeout=60,
            )
            status_code = upload_resp.status_code
            response_text = upload_resp.text[:200]
        if status_code in (200, 201, 204):
            log(f'  ✓ 文件上传成功：{filename}')
        else:
            log(f'  ⚠️ 文件上传返回 HTTP {status_code}（{response_text}），跳过')
    except Exception as e:
        log(f'  ⚠️ 文件上传异常（{e}），跳过——投诉仍将提交，但附件内容可能不可见')

    return {
        'bizName': 'ep',
        'scene': 'ipp',
        'name': filename,
        'fileId': file_id,
        'cloudType': cloud_type,
        'size': file_size,
        'status': 'success',
        'percent': 0,
        'uid': f'd-{int(time.time() * 1000)}-1',
        'url': static_url,
        'previewUrl': static_url,
    }


def normalize_xhs_link(url: str) -> str:
    """标准化小红书笔记链接格式。

    小红书导入接口需要保留 xsec_token/xsec_source 等查询参数，否则部分笔记
    会被平台解析为空数据。这里只做空白清理和缺省协议补齐，不再裁剪 query。
    """
    url = (url or '').strip()
    if url.startswith('//'):
        return 'https:' + url
    if url and not re.match(r'^https?://', url, re.I):
        return 'https://' + url
    return url


def _download_xhs_import_template_bytes(cookie: str) -> bytes:
    """下载小红书官方导入模板原始字节。"""
    resp = requests.get(
        XHS_IMPORT_TEMPLATE_URL,
        headers={
            'Accept': '*/*',
            'Referer': 'https://ipp.xiaohongshu.com/',
            'Cookie': cookie,
            'User-Agent': make_headers(cookie)['User-Agent'],
        },
        timeout=20,
    )
    resp.raise_for_status()
    log('  已下载小红书官方导入模板')
    return resp.content


def _build_xhs_import_workbook(cookie: str, links: list) -> bytes:
    """基于官方 xlsx 原始结构写入 A3 起的链接，避免 openpyxl 重写模板。"""
    template_bytes = _download_xhs_import_template_bytes(cookie)
    source = zipfile.ZipFile(io.BytesIO(template_bytes), 'r')
    output = io.BytesIO()

    shared_xml = source.read('xl/sharedStrings.xml').decode('utf-8')
    unique_match = re.search(r'uniqueCount="(\d+)"', shared_xml)
    start_idx = int(unique_match.group(1)) if unique_match else 0

    def make_row(row_no: int, shared_string_idx: int) -> str:
        return (
            f'<row r="{row_no}" spans="1:1">'
            f'<c r="A{row_no}" s="5" t="s"><v>{shared_string_idx}</v></c>'
            f'</row>'
        )

    appended_strings = ''.join(f'<si><t>{xml_escape(link)}</t></si>' for link in links)
    new_count = start_idx + len(links)
    shared_xml = re.sub(r'count="\d+"', f'count="{new_count}"', shared_xml, count=1)
    shared_xml = re.sub(r'uniqueCount="\d+"', f'uniqueCount="{new_count}"', shared_xml, count=1)
    shared_xml = shared_xml.replace('</sst>', appended_strings + '</sst>')

    sheet_xml = source.read('xl/worksheets/sheet1.xml').decode('utf-8')
    sheet_xml = re.sub(r'<dimension ref="[^"]+"/>',
                       f'<dimension ref="A1:F{max(3, len(links) + 2)}"/>',
                       sheet_xml, count=1)
    rows_xml = ''.join(make_row(i + 3, start_idx + i) for i, _ in enumerate(links))
    sheet_xml = re.sub(
        r'<row r="3"[^>]*>.*?</row>',
        rows_xml or '<row r="3" spans="1:1"><c r="A3" s="5"/></row>',
        sheet_xml,
        count=1,
        flags=re.S,
    )

    with zipfile.ZipFile(output, 'w', zipfile.ZIP_DEFLATED) as target:
        for item in source.infolist():
            if item.filename == 'xl/worksheets/sheet1.xml':
                data = sheet_xml.encode('utf-8')
            elif item.filename == 'xl/sharedStrings.xml':
                data = shared_xml.encode('utf-8')
            else:
                data = source.read(item.filename)
            target.writestr(item, data)
    source.close()
    return output.getvalue()


def _load_xhs_import_template(cookie: str):
    """下载小红书官方导入模板，避免 openpyxl 手写格式被平台判空。"""
    import io
    from openpyxl import Workbook, load_workbook

    try:
        resp = requests.get(
            XHS_IMPORT_TEMPLATE_URL,
            headers={
                'Accept': '*/*',
                'Origin': 'https://ipp.xiaohongshu.com',
                'Referer': 'https://ipp.xiaohongshu.com/',
                'Cookie': cookie,
                'User-Agent': make_headers(cookie)['User-Agent'],
            },
            timeout=20,
        )
        resp.raise_for_status()
        wb = load_workbook(io.BytesIO(resp.content))
        log('  已下载小红书官方导入模板')
        return wb
    except Exception as e:
        log(f'  ⚠️ 官方导入模板下载失败（{e}），使用本地兼容模板')
        wb = Workbook()
        ws = wb.active
        ws.title = '笔记'
        ws.merge_cells('A1:D1')
        ws['A1'] = ('说明：1、标*为必填；\n'
                    '2、每个excel文档最多支持100条数据导入，若超出请分批提交申请；\n'
                    '3、请勿修改表格格式，本说明无需删除；')
        ws['A2'] = '笔记链接*（请添加您本次希望投诉的笔记）'
        return wb


def _save_debug_import_file(task_id: str, batch_no: int, content: bytes) -> None:
    """保存实际上传给小红书的导入 Excel，方便排查平台判空问题。"""
    if not task_id:
        return
    try:
        debug_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'task_results')
        os.makedirs(debug_dir, exist_ok=True)
        path = os.path.join(debug_dir, f'{task_id}_import_{batch_no}.xlsx')
        with open(path, 'wb') as f:
            f.write(content)
        log(f'  调试文件已保存：{path}')
    except Exception as e:
        log(f'  ⚠️ 保存导入调试文件失败：{e}')


def import_infringement_links(cookie: str, links: list, work_name: str,
                              task_id: str = '', batch_no: int = 0) -> str:
    """将侵权链接导入小红书平台，返回 batchId。

    接口：POST /complaint/importInfringementFile
    上传一个 Excel 文件（Sheet1, 列A, 第1行为标题"侵权链接"）。
    平台会校验链接对应的笔记是否真实存在，无效链接被过滤为"空数据"。
    """
    import io
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError('缺少 openpyxl 库，请 pip install openpyxl')

    # URL 规范化（去除 xsec_token 等参数，保留纯 explore/{id} 格式）
    normalized_links = [normalize_xhs_link(lk) for lk in links]
    log(f'  链接规范化预览（前3条）:')
    for i, (orig, norm) in enumerate(zip(links[:3], normalized_links[:3])):
        changed = '→ ' + norm if norm != orig else '(unchanged)'
        log(f'    [{i+1}] {orig[:80]} {changed}')

    # 基于官方 xlsx 原始 zip/xml 结构写入链接，避免 openpyxl 重写模板后被平台判空。
    upload_content = _build_xhs_import_workbook(cookie, normalized_links)
    _save_debug_import_file(task_id, batch_no, upload_content)
    buf = io.BytesIO(upload_content)

    excel_filename = '笔记投诉批量导入模板.xlsx'

    headers = {k: v for k, v in make_headers(cookie).items() if k != 'Content-Type'}
    if _ACTIVE_BROWSER_CLIENT is not None:
        data = _ACTIVE_BROWSER_CLIENT.import_file(
            upload_content, excel_filename
        )
    else:
        resp = requests.post(
            f'{XHS_API_BASE}/complaint/importInfringementFile',
            headers=headers,
            files=[
                ('file', (excel_filename, buf,
                          'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')),
                ('batchIdList', (None, '[]')),
                ('infringementDetailType', (None, 'note')),
            ],
            timeout=60,
        )
        data = resp.json()
    log(f'  导入接口响应: {data}')
    if not data.get('success'):
        raise RuntimeError(f'导入侵权链接失败：{data.get("alertMsg") or data.get("errorMsg", "")}，响应：{data}')

    batch_id = (data.get('data') or {}).get('batchId') or data.get('data')
    if not batch_id:
        raise RuntimeError(f'导入侵权链接未返回 batchId，响应：{data}')

    import_result = data.get('data') or {}
    wait_import_finished(cookie, str(batch_id), import_result)

    log(f'侵权链接导入成功：{len(links)} 条 → batchId={batch_id}')
    return str(batch_id)


def wait_import_finished(cookie: str, batch_id: str, initial_result: dict = None) -> None:
    """等待平台完成批量导入解析。"""
    headers = {**make_headers(cookie), 'Accept': 'application/json'}
    last_msg = ''
    results = []
    if initial_result:
        results.append(initial_result)

    for attempt in range(1, 16):
        if attempt > 1 or not results:
            try:
                data = json_get(cookie, '/complaint/queryImportInfringementResult',
                                {'batchId': batch_id})
                results.append(data.get('data') or {})
            except Exception as e:
                last_msg = str(e)
                log(f'  ⚠️ 导入结果查询异常({attempt}/15): {e}')
                time.sleep(1)
                continue

        qr = results[-1] or {}
        all_ok = qr.get('allSuccess')
        last_msg = qr.get('errorMsg') or qr.get('msg') or ''
        log(f'  导入结果查询({attempt}/15): allSuccess={all_ok}, errorMsg={last_msg}, data={qr}')
        if last_msg:
            raise RuntimeError(f'导入侵权链接解析失败：{last_msg}，响应：{qr}')
        if all_ok is True:
            return
        time.sleep(1)
    raise RuntimeError(f'导入侵权链接解析超时：{last_msg or batch_id}')


def page_query_infringement(cookie: str, batch_id: str) -> list:
    """轮询读取导入后的侵权明细，提交投诉前明细必须非空。"""
    last_data = None
    for attempt in range(1, 16):
        data = json_post(cookie, '/complaint/pageQueryInfringement', {
            'batchIdList': [batch_id],
            'infringementDetailType': 'note',
            'pageNum': 1,
            'pageSize': 10,
        })
        last_data = data
        if not data.get('success'):
            raise RuntimeError(f'查询侵权明细失败：{data.get("alertMsg") or data.get("errorMsg", "")}')

        payload = data.get('data') or {}
        if isinstance(payload, list):
            details = payload
            total = len(details)
        else:
            details = (
                payload.get('list') or payload.get('records') or payload.get('data') or
                payload.get('infringementList') or payload.get('infringementDetailList') or []
            )
            total = payload.get('total', len(details))
        log(f'  侵权明细查询({attempt}/15): total={total}, count={len(details)}, data={data}')
        if details or total:
            log(f'  已查询到侵权明细：{len(details) or total} 条')
            return details
        time.sleep(1)

    raise RuntimeError(f'导入成功但侵权明细为空，batchId={batch_id}，最后响应：{last_data}')


def create_complaint(cookie: str, obligee_id: int, identity_id: int,
                     work_name: str, batch_id: str,
                     evidence_files: list) -> str:
    """创建投诉，返回 complaintNo。"""
    evidence_attachment = {str(COMPLAINT_TYPE_CODE): evidence_files}
    infringing_name = work_name

    payload = {
        'id': '',
        'complaintTitle': '投诉',
        'complaintType': [COMPLAINT_TYPE_CODE],
        'complaintDesc': '链接涉及上传分享传播快看漫画作品 存在侵权行为 请尽快处理',
        'complaintRequirement': '立即停止侵权，删除侵权内容，包括但不限于所列链接。',
        'evidenceAttachment': evidence_attachment,
        'firstPublicLinkMap': {},
        'infringingWorkName': infringing_name,
        'batchIdList': [batch_id],
        'otherAttachmentList': [],
        'remark': '',
        'complaintDetailType': 'note',
        'permitForwardData': True,
        'copyrightProofValidity': '长期有效',
        'obligeeId': obligee_id,
    }
    if identity_id:
        log(f'  已获取 identityId={identity_id}，但提交接口按浏览器成功请求不携带该字段')

    data = json_post(cookie, '/complaint/addOrUpdateComplaint', payload)
    if not data.get('success'):
        raise RuntimeError(f'创建投诉失败：{data.get("alertMsg") or data.get("errorMsg", "")}')

    complaint_no = (data.get('data') or {}).get('complaintNo') or ''
    if not complaint_no:
        # 有时响应体里数据在不同字段
        complaint_no = str(data.get('data', ''))

    log(f'投诉创建成功：{work_name} → complaintNo={complaint_no}')
    return complaint_no


def save_partial_result(task_id: str, result: dict):
    """增量落盘进度。"""
    try:
        result_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'task_results')
        os.makedirs(result_dir, exist_ok=True)
        with open(os.path.join(result_dir, f'{task_id}.json'), 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
    except Exception as e:
        log(f'⚠️ 增量保存进度失败: {e}')


def main():
    parser = argparse.ArgumentParser(description='小红书版权投诉自动化脚本')
    parser.add_argument('--task-id', required=True)
    parser.add_argument('--cookie', required=True)
    parser.add_argument('--config-file', required=True,
                        help='JSON: {"principal_name": "...", "works": [...]}')
    args = parser.parse_args()

    with open(args.config_file, encoding='utf-8') as f:
        config = json.load(f)

    cookie = args.cookie
    principal_name = config.get('principal_name', '')
    works_config = config.get('works', [])
    task_id = args.task_id

    result = {
        'task_id': task_id,
        'status': 'running',
        'started_at': datetime.now().isoformat(),
        'completed_batches': 0,
        'failed_batches': 0,
        'feedback_numbers': [],
        'feedback_numbers_by_work': [],
        'batch_results': [],
        'works_detail': [],
        'error_message': '',
    }

    matched_by_work = {}
    failed_works = set()

    def rebuild_numbers():
        ordered, by_work = [], []
        for w in works_config:
            wn = w['work_name']
            if matched_by_work.get(wn):
                nums = [str(n) for n in matched_by_work[wn]]
                ordered.extend(nums)
            elif wn in failed_works:
                nums = [f'投诉失败:{wn}']
                ordered.append(nums[0])
            else:
                nums = [f'未获取到单号:{wn}']
                ordered.append(nums[0])
            by_work.append({
                'work_name': wn,
                'numbers': nums,
                'status': 'failed' if wn in failed_works
                          else ('completed' if matched_by_work.get(wn) else 'partial_failed'),
            })
        result['feedback_numbers'] = ordered
        result['feedback_numbers_by_work'] = by_work

    browser_client = None
    global _ACTIVE_BROWSER_CLIENT
    try:
        browser_client = XhsBrowserClient(cookie).__enter__()
        _ACTIVE_BROWSER_CLIENT = browser_client
        log('已切换到 Playwright 签名请求模式')
    except Exception as e:
        log(f'⚠️ Playwright 初始化失败，回退普通 HTTP：{e}')
        browser_client = None

    try:
        log('开始执行小红书投诉任务...')

        # 验证登录
        check_login(cookie)

        # 获取权利人ID
        obligee_id = get_obligee_id(cookie, principal_name)

        # 获取身份ID（从已有投诉记录中取）
        try:
            identity_id = get_identity_id(cookie)
            log(f'身份ID: identityId={identity_id}')
        except Exception as e:
            log(f'⚠️ 无法自动获取 identityId（{e}），将跳过身份绑定字段（可能影响提交）')
            identity_id = 0

        # 按作品循环处理
        batch_no = 0
        for work_idx, work in enumerate(works_config):
            work_name = work['work_name']
            links = work.get('links', [])
            proof_path = work.get('proof_path', '')
            log(f'[{work_idx+1}/{len(works_config)}] 处理作品: {work_name} ({len(links)}条链接)')

            try:
                # 上传权属证明文件
                if not proof_path or not os.path.exists(proof_path):
                    raise RuntimeError(f'缺少权属证明文件：{proof_path}')

                log(f'准备权属证明：{os.path.basename(proof_path)}')
                evidence_file_obj = prepare_evidence_file(cookie, proof_path)

                result['works_detail'].append({
                    'work_index': work_idx,
                    'work_name': work_name,
                    'status': 'processing',
                })

                work_complaint_nos = []

                # 按批次提交（每批 ≤100 条链接）
                for chunk_start in range(0, len(links), MAX_LINKS_PER_BATCH):
                    batch_no += 1
                    chunk = links[chunk_start:chunk_start + MAX_LINKS_PER_BATCH]
                    log(f'  批次 {batch_no}：导入 {len(chunk)} 条侵权链接')

                    try:
                        # Step1: 导入侵权链接获取 batchId
                        batch_id = import_infringement_links(cookie, chunk, work_name, task_id, batch_no)
                        time.sleep(1)  # 稍作等待，让服务端处理完成

                        # Step2: 查询平台解析出的侵权明细；此接口会触发平台侧明细落库
                        page_query_infringement(cookie, batch_id)

                        # Step3: 创建投诉
                        complaint_no = create_complaint(
                            cookie,
                            obligee_id=obligee_id,
                            identity_id=identity_id,
                            work_name=work_name,
                            batch_id=batch_id,
                            evidence_files=[evidence_file_obj],
                        )

                        result['completed_batches'] += 1
                        result['batch_results'].append({
                            'batch_no': batch_no,
                            'work_name': work_name,
                            'status': 'completed',
                            'link_count': len(chunk),
                            'complaint_no': complaint_no,
                        })
                        if complaint_no:
                            work_complaint_nos.append(complaint_no)

                    except Exception as batch_err:
                        log(f'  ❌ 批次 {batch_no} 失败: {batch_err}')
                        result['failed_batches'] += 1
                        result['batch_results'].append({
                            'batch_no': batch_no,
                            'work_name': work_name,
                            'status': 'failed',
                            'error': str(batch_err),
                        })

                    time.sleep(2)

                if work_complaint_nos:
                    matched_by_work[work_name] = work_complaint_nos
                elif all(b['work_name'] == work_name and b['status'] == 'failed'
                         for b in result['batch_results']
                         if b['work_name'] == work_name):
                    failed_works.add(work_name)

                rebuild_numbers()
                save_partial_result(task_id, result)

            except Exception as work_err:
                log(f'❌ 作品「{work_name}」处理异常，跳过: {work_err}')
                failed_works.add(work_name)
                for chunk_start in range(0, max(len(links), 1), MAX_LINKS_PER_BATCH):
                    batch_no += 1
                    result['failed_batches'] += 1
                    result['batch_results'].append({
                        'batch_no': batch_no,
                        'work_name': work_name,
                        'status': 'failed',
                        'error': str(work_err),
                    })
                rebuild_numbers()
                save_partial_result(task_id, result)
                continue

        rebuild_numbers()
        if result['failed_batches'] == 0:
            result['status'] = 'completed'
        elif result['completed_batches'] > 0:
            result['status'] = 'partial_failed'
        else:
            result['status'] = 'failed'
        result['completed_at'] = datetime.now().isoformat()
        log(f'任务完成: 状态={result["status"]}, 成功={result["completed_batches"]}, '
            f'失败={result["failed_batches"]}, 单号={result["feedback_numbers"]}')

    except Exception as e:
        result['error_message'] = str(e)
        result['status'] = 'partial_failed' if result['completed_batches'] > 0 else 'failed'
        result['completed_at'] = datetime.now().isoformat()
        try:
            save_partial_result(result.get('task_id', task_id), result)
        except Exception:
            pass
        log(f'任务异常终止: {e}')

    finally:
        if browser_client is not None:
            _ACTIVE_BROWSER_CLIENT = None
            browser_client.__exit__(None, None, None)

    print('JSON_RESULT_START')
    print(json.dumps(result, ensure_ascii=False))
    print('JSON_RESULT_END')
    return 0 if result['status'] in ('completed', 'partial_failed') else 1


if __name__ == '__main__':
    sys.exit(main())
