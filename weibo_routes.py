# -*- coding: utf-8 -*-
"""微博投诉 Blueprint

沿用夸克模式：Sheet1 表单内容 + Sheet2 批量导入链接；证件材料按 static/imgs 目录
匹配后由后端脚本每次上传拿 picid。与夸克的不同：
  - 微博表单字段多、身份文本字段全部来自 Sheet1（DB 未存这些）；
  - 提交前需图形验证码（后端脚本内部 ddddocr+重试处理，路由层无感）；
  - 一单=1部作品，侵权链接>100 自动拆多单（后端处理）。
v1 仅支持机构代理场景。
"""

import io
import json
import math
import os
import time as _time
from datetime import datetime
from functools import wraps
from uuid import uuid4

import requests
from flask import Blueprint, request, jsonify, send_file, session as flask_session, current_app
from openpyxl import load_workbook
from sqlalchemy import text

weibo_bp = Blueprint('weibo', __name__, url_prefix='/api/weibo')

WEIBO_API_BASE = 'https://service.account.weibo.com'
LOGIN_EXPIRE_SECONDS = 43200

# 微博投诉表单 4 组单选项的「名称 → 编码」映射（2026-07-01 从投诉页面抓取）。
# 让用户在模板里填中文名，系统转成提交用的编码。
# 【重要】权利作品类型/投诉内容 的选项随「权利类型分组」联动变化，当前 3 张映射
# 对应的是「著作权」组的下级选项。故 v1 只放开著作权组（各项加「著作权-」前缀，
# 与商标权/专利权分离）；如需商标权/专利权，须另抓其下级选项映射再放开。
WEIBO_RIGHTS_TYPE_MAP = {
    '著作权-资源买卖分享下载': '4',
    '著作权-未经授权播放或发布': '5',
    '著作权-抄袭或未经授权搬运': '6',
    '著作权-盗版软件': '7',
    '著作权-其他': '1',
    # 以下商标权/专利权组暂不开放（下级选项不同，未抓取映射）：
    # '商标权-假货推广': '8', '商标权-盗用或仿冒注册商标': '9', '商标权-其他': '2', '专利权': '3',
}
# 权利作品类型（提交字段 class_id）
WEIBO_CLASS_ID_MAP = {
    '电影': '0', '电视剧': '1', '综艺': '2', '音乐': '3', '动漫': '4',
    '体育节目': '5', '短视频': '7', '游戏': '8', '图片': '9', '文章': '10', '其他': '6',
}
# 权利来源（提交字段 empower_type）
WEIBO_EMPOWER_TYPE_MAP = {
    '原始权利': '3', '独家授权': '1', '非独家授权': '2',
}
# 投诉内容（提交字段 c_content）
WEIBO_C_CONTENT_MAP = {
    '视频': '1', '音乐': '2', '图片': '3', '文章': '4', '微博内容': '6', '其他': '5',
}
# rights_type 中「需要填原作品链接」的编码：著作权类抄袭/未经授权搬运
WEIBO_RIGHTS_NEED_ORIGINAL = {'6'}

# 代理机构全称 → 简称（与 app.py map_agent_to_used_company 一致）。
# 授权委托书文件名用简称（如「授权委托书_<被代理人>_中惠信科_*」），
# 而 Sheet1 填的是全称，故匹配前先转简称。营业执照文件名用全称，不经此转换。
WEIBO_AGENT_ORG_SHORT = {
    '北京和晞科技有限公司': '和晞科技',
    '北京柏蒙文化传媒有限公司': '柏蒙文化',
    '北京中惠信科科技有限公司': '中惠信科',
}


# ── 懒加载 app 模块符号（避免循环引用）────────────────────────────────────────

def _app():
    import app as _m
    return _m


def get_db_session():
    return _app().get_db_session()


def get_redis_client():
    return _app().get_redis_client()


def get_current_user():
    return flask_session.get('username', '')


def _tasks():
    return _app().tasks


def normalize_work_path_part(v):
    return (v or '').strip().replace('/', '_').replace('\\', '_')


def normalize_datetime(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value)


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = flask_session.get('token')
        if not token:
            return jsonify({'success': False, 'error': '未登录', 'login_required': True}), 401
        login_time = flask_session.get('login_time', 0)
        if _time.time() - login_time > LOGIN_EXPIRE_SECONDS:
            flask_session.clear()
            return jsonify({'success': False, 'error': '登录已过期，请重新登录', 'login_required': True}), 401
        return f(*args, **kwargs)
    return decorated


def enqueue_weibo_task(payload: dict):
    payload['platform'] = 'weibo'
    m = _app()
    get_redis_client().lpush(m.UNIFIED_QUEUE_NAME, json.dumps(payload, ensure_ascii=False))


def _weibo_headers(cookie: str) -> dict:
    return {
        'Cookie': cookie,
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 '
                      '(KHTML, like Gecko) Chrome/149.0.0.0 Safari/537.36',
        'Referer': f'{WEIBO_API_BASE}/rights/movie',
    }


import re as _re


def _check_weibo_login(cookie: str) -> bool:
    resp = requests.get(f'{WEIBO_API_BASE}/rights/movie', headers=_weibo_headers(cookie), timeout=15)
    m = _re.search(r"\$CONFIG\['islogin'\]\s*=\s*(\d)", resp.text)
    return bool(m and m.group(1) == '1')


def _fetch_link_limit(cookie: str):
    """抓 /rights/movie，判定该账号每批侵权链接上限。

    直接镜像微博前端 movie.js 的判断（唯一权威依据）：
        if ($('#whitelist').val() == 1) { 上限1条 } else { 上限100条 }
    jQuery `$('#whitelist')` 取页面里【第一个】id=whitelist 元素的值（id 重复时只取第一个）。
    故这里取页面第一个 whitelist 的 value：== '1' → 受限每批1条；否则/缺失 → 每批100条。
    （实测：受限账号页面第一个 whitelist=1 → 限1条，与页面报错"投诉链接不能大于1个"一致。）

    返回 (per_batch_limit, is_limited, ok)：
      - per_batch_limit: 100 或 1
      - is_limited:      True 表示受限(1条/批)
      - ok:              是否成功读到登录态页面（读失败/未登录时降级默认1，ok=False）

    降级策略：探测失败时保守取每批1条——顶多慢，不会因超限被拒还白耗验证码；
    绝不乐观取100（受限账号会触发"投诉链接不能大于1个"失败）。
    """
    try:
        resp = requests.get(f'{WEIBO_API_BASE}/rights/movie', headers=_weibo_headers(cookie), timeout=15)
        html = resp.text
    except Exception:
        return 1, False, False
    # 未登录/异常页面：无法判定，保守降级默认1
    if "$CONFIG['islogin'] = 1" not in html and "islogin'] = 1" not in html:
        return 1, False, False
    # 取页面第一个 whitelist（镜像 jQuery $('#whitelist')）
    m = _re.search(r'id="whitelist"\s+value="([^"]*)"', html)
    val = (m.group(1).strip() if m else '')
    is_limited = (val == '1')
    return (1 if is_limited else 100), is_limited, True


# ── verify_cookie ──────────────────────────────────────────────────────────────

@weibo_bp.route('/verify_cookie', methods=['POST'])
@login_required
def weibo_verify_cookie():
    data = request.get_json() or {}
    cookie = data.get('cookie', '').strip()
    if not cookie:
        return jsonify({'success': False, 'error': 'Cookie不能为空'}), 400
    try:
        if _check_weibo_login(cookie):
            return jsonify({'success': True})
        return jsonify({'success': False, 'error': 'Cookie无效或已过期（未登录）'}), 401
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ── download_template ──────────────────────────────────────────────────────────

# Sheet1 字段定义：(字段名, 默认值, 说明)。投诉场景编码目前仅实测过"著作权-搬运"这套，
# 故以数字编码+默认值+说明呈现，不做易错的中文映射。
# Sheet1 字段定义：(字段名, 默认值, 说明)。投诉场景 4 项填「中文名称」，系统转编码。
WEIBO_SHEET1_FIELDS = [
    ('被代理人名称', '', '必填，须与被代理人营业执照/证明文件目录一致'),
    ('被代理人法人', '', '必填，被代理人营业执照上的法定代表人'),
    ('被代理人统一社会信用代码', '', '必填'),
    ('代理机构名称', '北京和晞科技有限公司', '必填，默认和晞科技'),
    ('代理机构法人', '', '必填，代理机构营业执照上的法定代表人'),
    ('代理机构统一社会信用代码', '', '必填'),
    ('机构联系人姓名', '', '必填'),
    ('机构联系人电话', '', '必填'),
    ('机构联系人身份证号', '', '必填'),
    ('权利类型', '著作权-抄袭或未经授权搬运',
     '必填。当前仅支持著作权类，可选：著作权-资源买卖分享下载/著作权-未经授权播放或发布/'
     '著作权-抄袭或未经授权搬运/著作权-盗版软件/著作权-其他。'
     '（选「著作权-抄袭或未经授权搬运」时需填原作品链接）'),
    ('权利作品类型', '综艺',
     '必填。可选：电影/电视剧/综艺/音乐/动漫/体育节目/短视频/游戏/图片/文章/其他'),
    ('权利来源', '独家授权', '必填。可选：原始权利/独家授权/非独家授权'),
    ('投诉内容', '微博内容', '必填。可选：视频/音乐/图片/文章/微博内容/其他'),
    ('投诉理由', '链接涉及上传分享传播独播作品存在侵权行为 请尽快处理', '必填'),
    ('处理要求', '删除', '必填'),
]


@weibo_bp.route('/download_template', methods=['GET'])
@login_required
def weibo_download_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()

    # Sheet1 表单内容
    ws1 = wb.active
    ws1.title = '表单内容'
    ws1.append(['字段', '值', '说明'])
    for cell in ws1[1]:
        cell.font = Font(bold=True)
    for name, default, remark in WEIBO_SHEET1_FIELDS:
        ws1.append([name, default, remark])
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 40
    ws1.column_dimensions['C'].width = 60

    # Sheet2 批量导入链接
    ws2 = wb.create_sheet('批量导入Excel')
    ws2.append(['侵权链接', '原作品链接', '作品名称'])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for _ in range(3):
        ws2.append(['', '', ''])
    ws2.column_dimensions['A'].width = 55
    ws2.column_dimensions['B'].width = 55
    ws2.column_dimensions['C'].width = 30

    # Sheet3 填写说明
    ws3 = wb.create_sheet('填写说明')
    for line in [
        ['微博版权投诉模版（v1 仅支持机构代理场景）'],
        [''],
        ['Sheet1 表单内容'],
        ['身份文本字段全部在 Sheet1 填写（数据库不存这些）'],
        ['投诉场景 4 项填「中文名称」，系统自动转编码，无需填数字：'],
        ['  权利类型（当前仅支持著作权类）：著作权-资源买卖分享下载/著作权-未经授权播放或发布/'],
        ['           著作权-抄袭或未经授权搬运/著作权-盗版软件/著作权-其他'],
        ['  权利作品类型：电影/电视剧/综艺/音乐/动漫/体育节目/短视频/游戏/图片/文章/其他'],
        ['  权利来源：原始权利/独家授权/非独家授权'],
        ['  投诉内容：视频/音乐/图片/文章/微博内容/其他'],
        [''],
        ['Sheet2 批量导入Excel'],
        ['侵权链接：必填，微博链接。一部作品超过每批上限时会自动拆成多单提交；'],
        ['  每批上限由账号权限决定（上传模板时自动检测）：普通账号每批100条，受限账号每批1条'],
        ['原作品链接：权利类型为「抄袭或未经授权搬运」时必填'],
        ['作品名称：必填，支持多部作品混合，系统按作品名分组；提交时自动用《》包裹'],
        [''],
        ['证明文件说明（沿用夸克约定，放 static/imgs/ 下）'],
        ['权属证明：static/imgs/剧名/<作品目录>/ 下「证明文件_*」'],
        ['被代理人营业执照：static/imgs/营业执照/「营业执照_<被代理人名称>*」'],
        ['代理机构营业执照：static/imgs/营业执照/「营业执照_<代理机构名称>*」'],
        ['授权委托书：static/imgs/授权委托书/「授权委托书_<被代理人>_<代理机构>*」'],
        ['机构联系人身份证正/反面：无需单独准备，系统复用代理机构营业执照图'],
    ]:
        ws3.append(line)
    ws3.column_dimensions['A'].width = 80

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='weibo_template.xlsx',
    )


# ── upload_template ──────────────────────────────────────────────────────────

def _paren(s):
    return (s or '').replace('（', '(').replace('）', ')')


def _match_file(dir_path, prefix, *must_contain):
    """在目录里找 prefix 开头、且包含所有 must_contain 片段的文件，返回绝对路径或空。"""
    if not os.path.isdir(dir_path):
        return ''
    for f in sorted(os.listdir(dir_path)):
        if f.startswith('._') or not f.startswith(prefix):
            continue
        if all(_paren(seg) in _paren(f) for seg in must_contain if seg):
            return os.path.join(dir_path, f)
    return ''


@weibo_bp.route('/upload_template', methods=['POST'])
@login_required
def weibo_upload_template():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '未上传文件'}), 400
    file = request.files['file']
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ('.xlsx', '.xls'):
        return jsonify({'success': False, 'error': '仅支持 .xlsx / .xls 格式'}), 400

    collect_account = request.form.get('collect_account', '').strip()

    try:
        wb = load_workbook(file, data_only=True)
    except Exception as e:
        return jsonify({'success': False, 'error': f'文件解析失败：{e}'}), 400
    if '表单内容' not in wb.sheetnames or '批量导入Excel' not in wb.sheetnames:
        return jsonify({'success': False, 'error': '模版缺少"表单内容"或"批量导入Excel"工作表'}), 400

    # 解析 Sheet1
    cfg = {}
    for row in wb['表单内容'].iter_rows(min_row=2, max_col=2, values_only=True):
        if row[0] and row[1] is not None and str(row[1]).strip() != '':
            cfg[str(row[0]).strip()] = str(row[1]).strip()

    def _g(key):
        return cfg.get(key, '').strip()

    principal_name = _g('被代理人名称')
    agent_org = _g('代理机构名称')
    org_agt_name = _g('机构联系人姓名')

    rights_type_name = _g('权利类型')
    class_name = _g('权利作品类型')
    empower_name = _g('权利来源')
    c_content_name = _g('投诉内容')

    required = {
        '被代理人名称': principal_name, '被代理人法人': _g('被代理人法人'),
        '被代理人统一社会信用代码': _g('被代理人统一社会信用代码'),
        '代理机构名称': agent_org, '代理机构法人': _g('代理机构法人'),
        '代理机构统一社会信用代码': _g('代理机构统一社会信用代码'),
        '机构联系人姓名': org_agt_name, '机构联系人电话': _g('机构联系人电话'),
        '机构联系人身份证号': _g('机构联系人身份证号'),
        '权利类型': rights_type_name, '权利作品类型': class_name,
        '权利来源': empower_name, '投诉内容': c_content_name,
    }
    missing = [k for k, v in required.items() if not v]
    if missing:
        return jsonify({'success': False, 'error': 'Sheet1 缺少必填项：' + '、'.join(missing)}), 400

    # 中文名称 → 编码。名称无效则报错，附可选项，避免提交时才被微博拒。
    def _to_code(name, mapping, field_cn):
        code = mapping.get(name)
        if code is None:
            return None, f'{field_cn}「{name}」无效，可选：' + '、'.join(mapping.keys())
        return code, ''

    code_errors = []
    rights_type, e = _to_code(rights_type_name, WEIBO_RIGHTS_TYPE_MAP, '权利类型')
    if e: code_errors.append(e)
    class_id, e = _to_code(class_name, WEIBO_CLASS_ID_MAP, '权利作品类型')
    if e: code_errors.append(e)
    empower_type, e = _to_code(empower_name, WEIBO_EMPOWER_TYPE_MAP, '权利来源')
    if e: code_errors.append(e)
    c_content, e = _to_code(c_content_name, WEIBO_C_CONTENT_MAP, '投诉内容')
    if e: code_errors.append(e)
    if code_errors:
        return jsonify({'success': False, 'error': '\n'.join(code_errors)}), 400

    # 组装后端 form（文本字段 + 场景编码），picid 由后端上传获得
    form = {
        'med_name': principal_name,
        'med_legname': _g('被代理人法人'),
        'med_idnum': _g('被代理人统一社会信用代码'),
        'org_name': agent_org,
        'org_legname': _g('代理机构法人'),
        'org_idnum': _g('代理机构统一社会信用代码'),
        'org_agt_name': org_agt_name,
        'org_agt_tel': _g('机构联系人电话'),
        'org_agt_idnum': _g('机构联系人身份证号'),
        'rights_type': rights_type,
        'class_id': class_id,
        'c_content': c_content,
        'empower_type': empower_type,
        'dpt_reason': _g('投诉理由'),
        'deal_req': _g('处理要求') or '删除',
        # 回显用（前端展示名称，不参与提交）
        '_rights_type_name': rights_type_name,
        '_class_name': class_name,
        '_empower_name': empower_name,
        '_c_content_name': c_content_name,
    }
    need_original = (rights_type in WEIBO_RIGHTS_NEED_ORIGINAL)

    # 解析 Sheet2（侵权链接 | 原作品链接 | 作品名称）
    works_map, work_order, empty_rows = {}, [], 0
    for row in wb['批量导入Excel'].iter_rows(min_row=2, max_col=3, values_only=True):
        link = str(row[0]).strip() if row[0] else ''
        original = str(row[1]).strip() if row[1] else ''
        wn = str(row[2]).strip() if row[2] else ''
        if not link and not wn:
            empty_rows += 1
            if empty_rows >= 5:
                break
            continue
        empty_rows = 0
        if not link:
            return jsonify({'success': False, 'error': f'存在作品名但侵权链接为空（作品：{wn}）'}), 400
        if not wn:
            return jsonify({'success': False, 'error': f'存在链接但作品名为空（链接：{link[:60]}）'}), 400
        if need_original and not original:
            return jsonify({'success': False, 'error': f'权利类型「{rights_type_name}」需填原作品链接（作品：{wn}）'}), 400
        if wn not in works_map:
            works_map[wn] = {'links': [], 'originals': []}
            work_order.append(wn)
        works_map[wn]['links'].append(link)
        works_map[wn]['originals'].append(original)

    if not works_map:
        return jsonify({'success': False, 'error': '"批量导入Excel"中没有有效数据'}), 400

    # PLACEHOLDER_MATCH_FILES
    static_imgs_dir = os.path.join(current_app.root_path, 'static', 'imgs')
    works_base_dir = os.path.join(static_imgs_dir, '剧名')
    biz_dir = os.path.join(static_imgs_dir, '营业执照')
    auth_dir = os.path.join(static_imgs_dir, '授权委托书')
    # 身份证目录暂不用：机构联系人身份证正反面复用代理机构营业执照图

    # 共享证件路径（同一被代理人+代理机构固定，后端上传一次复用）
    # 授权委托书文件名用代理机构【简称】（如「..._中惠信科_*」），Sheet1 填的是全称，
    # 故匹配前把全称转简称；未在映射表内则退回原值。营业执照文件名用全称，不转换。
    agent_org_short = WEIBO_AGENT_ORG_SHORT.get(agent_org, agent_org)
    _org_pic = _match_file(biz_dir, '营业执照_', agent_org)
    shared_paths = {
        'obusiness_path': _match_file(biz_dir, '营业执照_', principal_name),
        'org_pic_path': _org_pic,
        'org_empower_path': _match_file(auth_dir, '授权委托书_', principal_name, agent_org_short),
        # 机构联系人身份证正/反面：暂复用代理机构营业执照图（同一张用2次），
        # 不单独准备身份证文件。后端会分别上传各得 picid 填入正反面两个字段。
        'org_agt_pic1_path': _org_pic,
        'org_agt_pic2_path': _org_pic,
    }
    form.update(shared_paths)

    shared_missing = []
    if not shared_paths['obusiness_path']:
        shared_missing.append(f'被代理人营业执照（营业执照_{principal_name}*）')
    if not shared_paths['org_pic_path']:
        shared_missing.append(f'代理机构营业执照（营业执照_{agent_org}*）')
    if not shared_paths['org_empower_path']:
        shared_missing.append(f'授权委托书（授权委托书_{principal_name}_{agent_org}*）')
    # TODO(身份证后期改)：暂不校验联系人身份证正反面，避免阻塞其余流程测试。
    #   org_agt_pic1_path / org_agt_pic2_path 仍会匹配，文件存在则照常带上。
    if shared_missing:
        return jsonify({'success': False, 'error': '缺少共享证件材料：\n' + '\n'.join(shared_missing)}), 400

    # 逐作品匹配权属证明（沿用夸克：works 表拿 used_company/content_type/complaint_type 拼目录）
    works_config, match_errors = [], []
    for wn in work_order:
        db = get_db_session()
        try:
            rows = db.execute(text("""
                SELECT w.used_company,
                       ct.dict_name AS content_type, cpt.dict_name AS complaint_type
                FROM works w
                JOIN dictionaries ct  ON ct.dict_type='content_type'   AND ct.dict_code=CAST(w.content_type_id  AS CHAR)
                JOIN dictionaries cpt ON cpt.dict_type='complaint_type' AND cpt.dict_code=CAST(w.complaint_type_id AS CHAR)
                WHERE w.work_name=:wn AND w.principal_name=:pn
            """), {'wn': wn, 'pn': principal_name}).mappings().all()
        finally:
            db.close()

        if not rows:
            match_errors.append(f'「{wn}」在作品覆盖列表中不存在或被代理人不匹配（{principal_name}）')
            continue
        r0 = rows[0]
        uc, ct, cpt = r0.get('used_company') or '', r0.get('content_type', ''), r0.get('complaint_type', '')
        dir_name = f"{normalize_work_path_part(wn)}_{normalize_work_path_part(uc)}_{normalize_work_path_part(ct)}_{normalize_work_path_part(cpt)}"
        drama_dir = os.path.join(works_base_dir, dir_name)
        proof_path = _match_file(drama_dir, '证明文件_')
        if not proof_path:
            match_errors.append(f'「{wn}」缺少权属证明（目录 {dir_name}/证明文件_*）')
            continue

        works_config.append({
            'work_name': wn,
            'links': works_map[wn]['links'],
            'original_urls': works_map[wn]['originals'],
            'proof_path': proof_path,
        })

    if match_errors and not works_config:
        return jsonify({'success': False, 'error': '所有作品匹配失败：\n' + '\n'.join(match_errors)}), 400

    total_links = sum(len(w['links']) for w in works_config)

    # 抓页面判定每批上限（读账号 cookie 抓一次页面，只读不提交）。
    # 读失败或未选账号时【保守降级为 1 条/批】（whitelist_checked=False，前端提示"未能确认"）：
    # 宁可慢也不因超限被拒——若乐观按100，受限账号会报"投诉链接不能大于1个"且白耗验证码。
    per_batch_limit = 1
    is_whitelisted = False
    whitelist_checked = False
    if collect_account:
        _s = get_db_session()
        try:
            acc_row = _s.execute(text("""
                SELECT cookie_text FROM accounts
                WHERE platform_code='weibo' AND account_user=:acc LIMIT 1
            """), {'acc': collect_account}).fetchone()
        finally:
            _s.close()
        if acc_row and acc_row.cookie_text:
            per_batch_limit, is_whitelisted, whitelist_checked = _fetch_link_limit(acc_row.cookie_text)

    total_batches = sum(math.ceil(len(w['links']) / per_batch_limit) for w in works_config)

    resp_data = {
        'success': True,
        'filename': file.filename,
        'form': form,
        'works': works_config,
        'total_links': total_links,
        'total_batches': total_batches,
        'per_batch_limit': per_batch_limit,
        'is_whitelisted': is_whitelisted,
        'whitelist_checked': whitelist_checked,
        'principal_name': principal_name,
        'agent_org': agent_org,
        # 前端展示用名称（非编码）
        'rights_type_name': rights_type_name,
        'class_name': class_name,
        'empower_name': empower_name,
        'c_content_name': c_content_name,
    }
    if match_errors:
        resp_data['warnings'] = match_errors
    return jsonify(resp_data)


# ── submit ─────────────────────────────────────────────────────────────────────

@weibo_bp.route('/submit', methods=['POST'])
@login_required
def weibo_submit():
    data = request.get_json() or {}
    cookie = data.get('cookie', '').strip()
    collect_account = data.get('collect_account', '').strip()
    form = data.get('form', {})
    works_config = data.get('works', [])
    upload_filename = data.get('upload_filename', '').strip()
    # 每批链接上限（受限账号=1，普通=100），来自 upload_template 抓 whitelist 的结果。
    # 缺失/非法时【保守降级为1】——宁可慢也不因超限被拒；绝不乐观取100。
    try:
        per_batch_limit = int(data.get('per_batch_limit') or 1)
    except (TypeError, ValueError):
        per_batch_limit = 1
    if per_batch_limit <= 0:
        per_batch_limit = 1

    if not cookie:
        return jsonify({'success': False, 'error': 'Cookie不能为空'}), 400
    if not collect_account:
        return jsonify({'success': False, 'error': '请选择投诉账号'}), 400
    if not works_config:
        return jsonify({'success': False, 'error': '作品列表不能为空'}), 400

    # 防重复：同账号同文件未失败的记录已存在则拒绝
    if upload_filename:
        _s = get_db_session()
        try:
            dup = _s.execute(text("""
                SELECT task_id FROM complaints
                WHERE collect_account=:acc AND upload_filename=:fn
                  AND platform_code='weibo' AND status NOT IN ('failed')
                LIMIT 1
            """), {'acc': collect_account, 'fn': upload_filename}).fetchone()
        finally:
            _s.close()
        if dup:
            return jsonify({'success': False, 'error': f'文件「{upload_filename}」已投诉过（任务 {dup[0]}），请勿重复提交'}), 400

    # 验证 Cookie
    try:
        if not _check_weibo_login(cookie):
            return jsonify({'success': False, 'error': 'Cookie已失效，请更新后重试'}), 401
    except Exception as e:
        return jsonify({'success': False, 'error': f'Cookie验证失败：{e}'}), 500

    total_links = sum(len(w.get('links', [])) for w in works_config)
    total_batches = sum(math.ceil(len(w.get('links', [])) / per_batch_limit) for w in works_config)
    all_work_names = [w['work_name'] for w in works_config]

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    submission_id = f"{timestamp}_{uuid4().hex[:8]}"
    task_id = f'weibo_{submission_id}'

    principal_name = data.get('principal_name', '') or form.get('med_name', '')
    agent_org = data.get('agent_org', '') or form.get('org_name', '')

    db = get_db_session()
    try:
        submitted_at = datetime.now()
        estimated_finish_at = _app().compute_estimated_finish(db, total_batches, 'weibo', submitted_at)
        db.execute(text("""
            INSERT INTO complaints
            (complaint_id, task_id, platform_code, collect_account, cookie_snapshot,
             identity_type, agent_name, principal_name,
             complaint_category, complaint_type, module_name, content_type,
             description_text, work_name, total_links, batch_size, batch_count,
             status, submitted_at, estimated_finish_at, operator, upload_filename)
            VALUES (:sid, :tid, 'weibo', :account, :cookie,
                    '机构代理', :agent, :principal,
                    '著作权', :rights, '微博', '',
                    :dpt, :work_name, :rows, :batch_size, :batches,
                    'queued', :submitted_at, :estimated_finish_at, :operator, :upload_filename)
        """), {
            'batch_size': per_batch_limit,
            'sid': submission_id,
            'tid': task_id,
            'account': collect_account,
            'cookie': cookie[:100] + '...',
            'agent': agent_org,
            'principal': principal_name,
            'rights': f"rights_type={form.get('rights_type', '')}",
            'dpt': form.get('dpt_reason', ''),
            'work_name': ', '.join(all_work_names)[:5000],
            'rows': total_links,
            'batches': total_batches,
            'submitted_at': submitted_at,
            'estimated_finish_at': estimated_finish_at,
            'operator': get_current_user(),
            'upload_filename': upload_filename,
        })

        batch_no = 0
        for work in works_config:
            links = work.get('links', [])
            for chunk_start in range(0, len(links), per_batch_limit):
                batch_no += 1
                chunk_end = min(chunk_start + per_batch_limit, len(links))
                db.execute(text("""
                    INSERT INTO complaint_batches
                    (batch_id, complaint_id, batch_no, work_name, batch_filename,
                     start_row, end_row, row_count, status)
                    VALUES (:bid, :sid, :bno, :wname, :fname, :sr, :er, :rc, 'pending')
                """), {
                    'bid': uuid4().hex[:12],
                    'sid': submission_id,
                    'bno': batch_no,
                    'wname': work['work_name'],
                    'fname': f"{work['work_name']}_part{batch_no}",
                    'sr': chunk_start + 1,
                    'er': chunk_end,
                    'rc': chunk_end - chunk_start,
                })

        for idx, work in enumerate(works_config):
            db.execute(text("""
                INSERT INTO submission_works
                (complaint_id, work_index, work_name, platform_code,
                 link_count, batch_count, status)
                VALUES (:sid, :widx, :wname, 'weibo', :lcount, :bcount, 'pending')
            """), {
                'sid': submission_id,
                'widx': idx,
                'wname': work['work_name'],
                'lcount': len(work.get('links', [])),
                'bcount': math.ceil(len(work.get('links', [])) / per_batch_limit),
            })

        db.commit()
    except Exception as e:
        db.rollback()
        return jsonify({'success': False, 'error': f'数据库写入失败：{e}'}), 500
    finally:
        db.close()

    enqueue_weibo_task({
        'task_id': task_id,
        'submission_id': submission_id,
        'cookie': cookie,
        'form': form,
        'works_config': works_config,
        'total_batches': total_batches,
        'per_batch_limit': per_batch_limit,
    })

    _tasks()[task_id] = {'status': 'queued', 'submitted_at': datetime.now().isoformat()}
    return jsonify({'success': True, 'task_id': task_id, 'submission_id': submission_id})


# ── status list ───────────────────────────────────────────────────────────────

@weibo_bp.route('/status_list', methods=['GET'])
@login_required
def weibo_status_list():
    db = get_db_session()
    try:
        rows = db.execute(text("""
            SELECT complaint_id AS submission_id, task_id, collect_account, work_name,
                   total_links, batch_count, submitted_at, estimated_finish_at, status,
                   complaint_numbers_json, error_message, operator
            FROM complaints
            WHERE platform_code = 'weibo'
            ORDER BY submitted_at DESC
            LIMIT 50
        """)).fetchall()
        status_map = {
            'queued': '等待中', 'running': '执行中', 'completed': '已完成',
            'failed': '失败', 'partial_failed': '部分失败',
        }
        result = []
        for row in rows:
            complaint_numbers = []
            if row.complaint_numbers_json:
                try:
                    complaint_numbers = json.loads(row.complaint_numbers_json)
                except Exception:
                    pass
            result.append({
                'submission_id': row.submission_id,
                'task_id': row.task_id,
                'collect_account': row.collect_account,
                'work_name': row.work_name,
                'total_links': row.total_links,
                'batch_count': row.batch_count,
                'submitted_at': normalize_datetime(row.submitted_at),
                'estimated_finish_at': normalize_datetime(row.estimated_finish_at),
                'status': status_map.get(row.status, row.status or '等待中'),
                'complaint_numbers': complaint_numbers,
                'operator': row.operator or '',
            })
        return jsonify({'success': True, 'data': result})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        db.close()


# ── export excel ──────────────────────────────────────────────────────────────

@weibo_bp.route('/export_excel/<submission_id>', methods=['GET'])
@login_required
def weibo_export_excel(submission_id):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    db = get_db_session()
    try:
        sub = db.execute(text("""
            SELECT complaint_id, collect_account, submitted_at, complaint_numbers_json
            FROM complaints WHERE complaint_id = :sid AND platform_code = 'weibo'
        """), {'sid': submission_id}).fetchone()
        if not sub:
            return jsonify({'success': False, 'error': '记录不存在'}), 404

        works = db.execute(text("""
            SELECT work_name, status, feedback_numbers FROM submission_works
            WHERE complaint_id = :sid ORDER BY work_index
        """), {'sid': submission_id}).fetchall()

        complaint_numbers = []
        if sub.complaint_numbers_json:
            try:
                complaint_numbers = json.loads(sub.complaint_numbers_json)
            except Exception:
                pass

        submitted_at = ''
        if sub.submitted_at:
            submitted_at = sub.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if hasattr(sub.submitted_at, 'strftime') else str(sub.submitted_at)

        wb = Workbook()
        ws = wb.active
        ws.title = '投诉结果'
        ws.append(['采集时间', '采集账号', '作品名称', '投诉单号(rdid)'])
        for cell in ws[1]:
            cell.font = Font(bold=True)

        has_per_work = any(getattr(w, 'feedback_numbers', None) for w in works)
        number_idx = 0
        for work in works:
            if has_per_work:
                nums = []
                raw = getattr(work, 'feedback_numbers', None)
                if raw:
                    try:
                        nums = json.loads(raw) if isinstance(raw, str) else list(raw)
                    except (TypeError, json.JSONDecodeError):
                        nums = []
                if not nums:
                    nums = ['']
                for num in nums:
                    ws.append([submitted_at, sub.collect_account, work.work_name, str(num)])
            else:
                num = complaint_numbers[number_idx] if number_idx < len(complaint_numbers) else ''
                number_idx += 1
                ws.append([submitted_at, sub.collect_account, work.work_name, str(num)])

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 30

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True,
                         download_name=f'weibo_{submission_id}.xlsx')
    finally:
        db.close()


# ── task status ────────────────────────────────────────────────────────────────

@weibo_bp.route('/task/<task_id>', methods=['GET'])
@login_required
def weibo_task_status(task_id):
    db = get_db_session()
    try:
        row = db.execute(text("""
            SELECT task_id, status, batch_count, completed_batches, failed_batches,
                   complaint_numbers_json, error_message,
                   submitted_at, started_at, completed_at
            FROM complaints WHERE task_id = :tid
        """), {'tid': task_id}).fetchone()
        if not row:
            mem = _tasks().get(task_id)
            if mem:
                return jsonify({'success': True, 'task': mem})
            return jsonify({'success': False, 'error': '任务不存在'}), 404
        return jsonify({
            'success': True,
            'task': {
                'task_id': row.task_id,
                'status': row.status,
                'batch_count': row.batch_count,
                'completed_batches': row.completed_batches,
                'failed_batches': row.failed_batches,
                'complaint_numbers': json.loads(row.complaint_numbers_json) if row.complaint_numbers_json else [],
                'error_message': row.error_message,
                'submitted_at': normalize_datetime(row.submitted_at),
                'started_at': normalize_datetime(row.started_at),
                'completed_at': normalize_datetime(row.completed_at),
            }
        })
    finally:
        db.close()
