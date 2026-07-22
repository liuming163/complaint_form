# -*- coding: utf-8 -*-
"""小红书投诉 Blueprint

投诉流程：
  1. 用户上传 Excel 模板（包含作品名称、侵权链接）
  2. 系统匹配权属证明文件
  3. 后端脚本上传证明文件到小红书 CDN
  4. 导入侵权链接获得 batchId
  5. 创建投诉记录（addOrUpdateComplaint）

v1 仅支持「知识产权/著作权/其他著作权侵权（如广播剧、动漫、软件等）」类型。
"""

import io
import json
import math
import os
import time as _time
from datetime import datetime
from functools import wraps
from uuid import uuid4

import requests
from flask import Blueprint, request, jsonify, send_file, session as flask_session, current_app
from openpyxl import load_workbook
from sqlalchemy import text

xiaohongshu_bp = Blueprint('xiaohongshu', __name__, url_prefix='/api/xiaohongshu')

XHS_API_BASE = 'https://ipp.xiaohongshu.com/api/xhsipp'
LOGIN_EXPIRE_SECONDS = 43200

# 投诉类型固定：complaintId=21 对应「其他著作权侵权（如广播剧、动漫、软件等）」
XHS_COMPLAINT_TYPE_CODE = 21
XHS_COMPLAINT_TYPE_PATH = '知识产权/著作权(包含抄袭、搬运等)/其他著作权侵权（如广播剧、动漫、软件等）'

# 投诉内容类型：笔记
XHS_COMPLAINT_DETAIL_TYPE = 'note'

# 每批最多链接数（小红书限制）
MAX_LINKS_PER_BATCH = 100


# ── 懒加载 app 模块符号（避免循环引用）────────────────────────────────────────

def _app():
    import app as _m
    return _m


def get_db_session():
    return _app().get_db_session()


def get_redis_client():
    return _app().get_redis_client()


def get_current_user():
    return flask_session.get('username', '')


def _tasks():
    return _app().tasks


def normalize_work_path_part(v):
    return (v or '').strip().replace('/', '_').replace('\\', '_')


def normalize_datetime(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value)


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = flask_session.get('token')
        if not token:
            return jsonify({'success': False, 'error': '未登录', 'login_required': True}), 401
        login_time = flask_session.get('login_time', 0)
        if _time.time() - login_time > LOGIN_EXPIRE_SECONDS:
            flask_session.clear()
            return jsonify({'success': False, 'error': '登录已过期，请重新登录', 'login_required': True}), 401
        return f(*args, **kwargs)
    return decorated


def enqueue_xiaohongshu_task(payload: dict):
    payload['platform'] = 'xiaohongshu'
    m = _app()
    get_redis_client().lpush(m.UNIFIED_QUEUE_NAME, json.dumps(payload, ensure_ascii=False))


def _xhs_headers(cookie: str) -> dict:
    return {
        'Cookie': cookie,
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 '
                      '(KHTML, like Gecko) Chrome/150.0.0.0 Safari/537.36',
        'Referer': 'https://ipp.xiaohongshu.com/complaint-management/edit',
        'Origin': 'https://ipp.xiaohongshu.com',
        'Accept': 'application/json, text/plain, */*',
        'xsecappid': 'complaint-center',
    }


def _check_xhs_login(cookie: str) -> bool:
    """检查小红书 Cookie 是否有效（查询投诉列表）"""
    try:
        resp = requests.post(
            f'{XHS_API_BASE}/complaint/pageQueryComplaint',
            headers={**_xhs_headers(cookie), 'Content-Type': 'application/json'},
            json={'pageSize': 1, 'pageNum': 1},
            timeout=15
        )
        data = resp.json()
        return data.get('success') is True
    except Exception:
        return False


# ── verify_cookie ──────────────────────────────────────────────────────────────

@xiaohongshu_bp.route('/verify_cookie', methods=['POST'])
@login_required
def xiaohongshu_verify_cookie():
    data = request.get_json() or {}
    cookie = data.get('cookie', '').strip()
    if not cookie:
        return jsonify({'success': False, 'error': 'Cookie不能为空'}), 400
    try:
        if _check_xhs_login(cookie):
            return jsonify({'success': True})
        return jsonify({'success': False, 'error': 'Cookie无效或已过期'}), 401
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ── download_template ──────────────────────────────────────────────────────────

@xiaohongshu_bp.route('/download_template', methods=['GET'])
@login_required
def xiaohongshu_download_template():
    from openpyxl import Workbook
    from openpyxl.styles import Alignment

    wb = Workbook()

    # Sheet「笔记」需保持小红书官方导入模板格式：第 3 行开始填链接。
    ws1 = wb.active
    ws1.title = '笔记'
    ws1.merge_cells('A1:D1')
    ws1['A1'] = ('说明：1、标*为必填；\n'
                 '2、每个excel文档最多支持100条数据导入，若超出请分批提交申请；\n'
                 '3、请勿修改表格格式，本说明无需删除；')
    ws1['A1'].alignment = Alignment(wrap_text=True, vertical='center')
    ws1['A2'] = '笔记链接*（请添加您本次希望投诉的笔记）'
    ws1.append(['', '', '', '', '', ''])
    ws1.column_dimensions['A'].width = 64.5
    ws1.column_dimensions['B'].width = 26.5
    ws1.column_dimensions['C'].width = 22.3
    ws1.column_dimensions['D'].width = 33.1
    ws1.row_dimensions[1].height = 79

    # B 列是本系统额外使用的作品名，不会上传给小红书平台。
    ws1['B2'] = '作品名称（系统使用，必填）'

    # Sheet2 填写说明（本系统额外使用，用户在这里填作品名映射说明）
    ws2 = wb.create_sheet('填写说明')
    for line in [
        ['小红书版权投诉模版'],
        [''],
        ['Sheet「笔记」'],
        ['A列笔记链接：必填，小红书笔记链接。每批最多100条，超过自动拆分'],
        ['B列作品名称：必填，支持多部作品混合，系统按作品名分组（不会上传给小红书）'],
        [''],
        ['证明文件说明'],
        ['权属证明：static/imgs/剧名/<作品目录>/ 下「证明文件_*」'],
        ['  （作品目录格式：<作品名>_<使用公司>_<内容类型>_<投诉类型>）'],
        [''],
        ['投诉类型：固定为「知识产权/著作权/其他著作权侵权（如广播剧、动漫、软件等）」'],
        ['投诉内容：笔记'],
        ['投诉标题：投诉'],
        ['侵权详情描述：链接涉及上传分享传播快看漫画作品 存在侵权行为 请尽快处理'],
        ['投诉请求：立即停止侵权，删除侵权内容，包括但不限于所列链接。'],
        ['允许转发权属证明材料：是'],
    ]:
        ws2.append(line)
    ws2.column_dimensions['A'].width = 80

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='xiaohongshu_template.xlsx',
    )


# ── upload_template ──────────────────────────────────────────────────────────

def _match_file(dir_path, prefix, *must_contain, exclude_exts=None):
    """在目录里找 prefix 开头、且包含所有 must_contain 片段的文件，返回绝对路径或空。"""
    if not os.path.isdir(dir_path):
        return ''
    excluded = {ext.lower() for ext in (exclude_exts or [])}
    for f in sorted(os.listdir(dir_path)):
        if f.startswith('._') or not f.startswith(prefix):
            continue
        if os.path.splitext(f)[1].lower() in excluded:
            continue
        if all(seg in f for seg in must_contain if seg):
            return os.path.join(dir_path, f)
    return ''


@xiaohongshu_bp.route('/upload_template', methods=['POST'])
@login_required
def xiaohongshu_upload_template():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '未上传文件'}), 400
    file = request.files['file']
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ('.xlsx', '.xls'):
        return jsonify({'success': False, 'error': '仅支持 .xlsx / .xls 格式'}), 400

    collect_account = request.form.get('collect_account', '').strip()
    principal_name = request.form.get('principal_name', '').strip()

    if not principal_name:
        return jsonify({'success': False, 'error': '请选择被代理人'}), 400

    try:
        wb = load_workbook(file, data_only=True)
    except Exception as e:
        return jsonify({'success': False, 'error': f'文件解析失败：{e}'}), 400
    if '笔记' not in wb.sheetnames:
        return jsonify({'success': False, 'error': '模版缺少"笔记"工作表'}), 400

    # 解析官方模板：A2 为表头，A3 起为链接；B列为本系统额外作品名。
    works_map, work_order, empty_rows = {}, [], 0
    for row in wb['笔记'].iter_rows(min_row=3, max_col=2, values_only=True):
        link = str(row[0]).strip() if row[0] else ''
        wn = str(row[1]).strip() if len(row) > 1 and row[1] else ''
        if not link and not wn:
            empty_rows += 1
            if empty_rows >= 5:
                break
            continue
        empty_rows = 0
        if not link:
            return jsonify({'success': False, 'error': f'存在作品名但笔记链接为空（作品：{wn}）'}), 400
        if not wn:
            return jsonify({'success': False, 'error': f'存在链接但作品名为空（链接：{link[:60]}）'}), 400
        if wn not in works_map:
            works_map[wn] = {'links': []}
            work_order.append(wn)
        works_map[wn]['links'].append(link)

    if not works_map:
        return jsonify({'success': False, 'error': '"笔记"工作表中没有有效数据'}), 400

    # 逐作品匹配权属证明（按用户选定的被代理人过滤）
    static_imgs_dir = os.path.join(current_app.root_path, 'static', 'imgs')
    works_base_dir = os.path.join(static_imgs_dir, '剧名')

    works_config, match_errors = [], []
    for wn in work_order:
        db = get_db_session()
        try:
            rows = db.execute(text("""
                SELECT w.used_company,
                       ct.dict_name AS content_type, cpt.dict_name AS complaint_type
                FROM works w
                JOIN dictionaries ct  ON ct.dict_type='content_type'   AND ct.dict_code=CAST(w.content_type_id  AS CHAR)
                JOIN dictionaries cpt ON cpt.dict_type='complaint_type' AND cpt.dict_code=CAST(w.complaint_type_id AS CHAR)
                WHERE w.work_name=:wn AND w.principal_name=:pn
            """), {'wn': wn, 'pn': principal_name}).mappings().all()
        finally:
            db.close()

        if not rows:
            match_errors.append(f'「{wn}」在作品覆盖列表中不存在或被代理人不匹配（{principal_name}）')
            continue
        r0 = rows[0]
        uc, ct, cpt = r0.get('used_company') or '', r0.get('content_type', ''), r0.get('complaint_type', '')
        dir_name = f"{normalize_work_path_part(wn)}_{normalize_work_path_part(uc)}_{normalize_work_path_part(ct)}_{normalize_work_path_part(cpt)}"
        drama_dir = os.path.join(works_base_dir, dir_name)
        proof_path = _match_file(drama_dir, '证明文件_')
        if not proof_path:
            match_errors.append(f'「{wn}」缺少权属证明（目录 {dir_name}/证明文件_*）')
            continue

        works_config.append({
            'work_name': wn,
            'links': works_map[wn]['links'],
            'proof_path': proof_path,
        })

    if match_errors and not works_config:
        return jsonify({'success': False, 'error': '所有作品匹配失败：\n' + '\n'.join(match_errors)}), 400

    total_links = sum(len(w['links']) for w in works_config)
    total_batches = sum(math.ceil(len(w['links']) / MAX_LINKS_PER_BATCH) for w in works_config)

    resp_data = {
        'success': True,
        'filename': file.filename,
        'works': works_config,
        'total_links': total_links,
        'total_batches': total_batches,
        'principal_name': principal_name,
        'complaint_type': XHS_COMPLAINT_TYPE_PATH,
    }
    if match_errors:
        resp_data['warnings'] = match_errors
    return jsonify(resp_data)


# ── submit ─────────────────────────────────────────────────────────────────────

@xiaohongshu_bp.route('/submit', methods=['POST'])
@login_required
def xiaohongshu_submit():
    data = request.get_json() or {}
    cookie = data.get('cookie', '').strip()
    collect_account = data.get('collect_account', '').strip()
    # principal_name 由前端从 upload_template 响应中传回（auto-detected）
    principal_name = data.get('principal_name', '').strip()
    works_config = data.get('works', [])
    upload_filename = data.get('upload_filename', '').strip()

    if not cookie:
        return jsonify({'success': False, 'error': 'Cookie不能为空'}), 400
    if not collect_account:
        return jsonify({'success': False, 'error': '请选择投诉账号'}), 400
    if not works_config:
        return jsonify({'success': False, 'error': '作品列表不能为空'}), 400

    # 防重复：同账号同文件未失败的记录已存在则拒绝
    if upload_filename:
        _s = get_db_session()
        try:
            dup = _s.execute(text("""
                SELECT task_id FROM complaints
                WHERE collect_account=:acc AND upload_filename=:fn
                  AND platform_code='xiaohongshu' AND status NOT IN ('failed')
                LIMIT 1
            """), {'acc': collect_account, 'fn': upload_filename}).fetchone()
        finally:
            _s.close()
        if dup:
            return jsonify({'success': False, 'error': f'文件「{upload_filename}」已投诉过（任务 {dup[0]}），请勿重复提交'}), 400

    # 验证 Cookie
    try:
        if not _check_xhs_login(cookie):
            return jsonify({'success': False, 'error': 'Cookie已失效，请更新后重试'}), 401
    except Exception as e:
        return jsonify({'success': False, 'error': f'Cookie验证失败：{e}'}), 500

    total_links = sum(len(w.get('links', [])) for w in works_config)
    total_batches = sum(math.ceil(len(w.get('links', [])) / MAX_LINKS_PER_BATCH) for w in works_config)
    all_work_names = [w['work_name'] for w in works_config]

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    submission_id = f"{timestamp}_{uuid4().hex[:8]}"
    task_id = f'xiaohongshu_{submission_id}'

    db = get_db_session()
    try:
        submitted_at = datetime.now()
        estimated_finish_at = _app().compute_estimated_finish(db, total_batches, 'xiaohongshu', submitted_at)
        db.execute(text("""
            INSERT INTO complaints
            (complaint_id, task_id, platform_code, collect_account, cookie_snapshot,
             identity_type, agent_name, principal_name,
             complaint_category, complaint_type, module_name, content_type,
             description_text, work_name, total_links, batch_size, batch_count,
             status, submitted_at, estimated_finish_at, operator, upload_filename)
            VALUES (:sid, :tid, 'xiaohongshu', :account, :cookie,
                    '', '', :principal,
                    '知识产权', '著作权', '笔记', '',
                    :desc_text, :work_name, :rows, :batch_size, :batches,
                    'queued', :submitted_at, :estimated_finish_at, :operator, :upload_filename)
        """), {
            'batch_size': MAX_LINKS_PER_BATCH,
            'sid': submission_id,
            'tid': task_id,
            'account': collect_account,
            'cookie': cookie[:100] + '...',
            'principal': principal_name,
            'desc_text': '链接涉及上传分享传播快看漫画作品 存在侵权行为 请尽快处理',
            'work_name': ', '.join(all_work_names)[:5000],
            'rows': total_links,
            'batches': total_batches,
            'submitted_at': submitted_at,
            'estimated_finish_at': estimated_finish_at,
            'operator': get_current_user(),
            'upload_filename': upload_filename,
        })

        batch_no = 0
        for work in works_config:
            links = work.get('links', [])
            for chunk_start in range(0, len(links), MAX_LINKS_PER_BATCH):
                batch_no += 1
                chunk_end = min(chunk_start + MAX_LINKS_PER_BATCH, len(links))
                db.execute(text("""
                    INSERT INTO complaint_batches
                    (batch_id, complaint_id, batch_no, work_name, batch_filename,
                     start_row, end_row, row_count, status)
                    VALUES (:bid, :sid, :bno, :wname, :fname, :sr, :er, :rc, 'pending')
                """), {
                    'bid': uuid4().hex[:12],
                    'sid': submission_id,
                    'bno': batch_no,
                    'wname': work['work_name'],
                    'fname': f"{work['work_name']}_part{batch_no}",
                    'sr': chunk_start + 1,
                    'er': chunk_end,
                    'rc': chunk_end - chunk_start,
                })

        for idx, work in enumerate(works_config):
            db.execute(text("""
                INSERT INTO submission_works
                (complaint_id, work_index, work_name, platform_code,
                 link_count, batch_count, status)
                VALUES (:sid, :widx, :wname, 'xiaohongshu', :lcount, :bcount, 'pending')
            """), {
                'sid': submission_id,
                'widx': idx,
                'wname': work['work_name'],
                'lcount': len(work.get('links', [])),
                'bcount': math.ceil(len(work.get('links', [])) / MAX_LINKS_PER_BATCH),
            })

        db.commit()
    except Exception as e:
        db.rollback()
        return jsonify({'success': False, 'error': f'数据库写入失败：{e}'}), 500
    finally:
        db.close()

    enqueue_xiaohongshu_task({
        'task_id': task_id,
        'submission_id': submission_id,
        'cookie': cookie,
        'principal_name': principal_name,
        'works_config': works_config,
        'total_batches': total_batches,
    })

    _tasks()[task_id] = {'status': 'queued', 'submitted_at': datetime.now().isoformat()}
    return jsonify({'success': True, 'task_id': task_id, 'submission_id': submission_id})


# ── status list ───────────────────────────────────────────────────────────────

@xiaohongshu_bp.route('/status_list', methods=['GET'])
@login_required
def xiaohongshu_status_list():
    db = get_db_session()
    try:
        rows = db.execute(text("""
            SELECT complaint_id AS submission_id, task_id, collect_account, work_name,
                   total_links, batch_count, submitted_at, estimated_finish_at, status,
                   complaint_numbers_json, error_message, operator
            FROM complaints
            WHERE platform_code = 'xiaohongshu'
            ORDER BY submitted_at DESC
            LIMIT 50
        """)).fetchall()
        status_map = {
            'queued': '等待中', 'running': '执行中', 'completed': '已完成',
            'failed': '失败', 'partial_failed': '部分失败',
        }
        result = []
        for row in rows:
            complaint_numbers = []
            if row.complaint_numbers_json:
                try:
                    complaint_numbers = json.loads(row.complaint_numbers_json)
                except Exception:
                    pass
            result.append({
                'submission_id': row.submission_id,
                'task_id': row.task_id,
                'collect_account': row.collect_account,
                'work_name': row.work_name,
                'total_links': row.total_links,
                'batch_count': row.batch_count,
                'submitted_at': normalize_datetime(row.submitted_at),
                'estimated_finish_at': normalize_datetime(row.estimated_finish_at),
                'status': status_map.get(row.status, row.status or '等待中'),
                'complaint_numbers': complaint_numbers,
                'operator': row.operator or '',
            })
        return jsonify({'success': True, 'data': result})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        db.close()


# ── export excel ──────────────────────────────────────────────────────────────

@xiaohongshu_bp.route('/export_excel/<submission_id>', methods=['GET'])
@login_required
def xiaohongshu_export_excel(submission_id):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    db = get_db_session()
    try:
        sub = db.execute(text("""
            SELECT complaint_id, collect_account, submitted_at, complaint_numbers_json
            FROM complaints WHERE complaint_id = :sid AND platform_code = 'xiaohongshu'
        """), {'sid': submission_id}).fetchone()
        if not sub:
            return jsonify({'success': False, 'error': '记录不存在'}), 404

        works = db.execute(text("""
            SELECT work_name, status, feedback_numbers FROM submission_works
            WHERE complaint_id = :sid ORDER BY work_index
        """), {'sid': submission_id}).fetchall()

        complaint_numbers = []
        if sub.complaint_numbers_json:
            try:
                complaint_numbers = json.loads(sub.complaint_numbers_json)
            except Exception:
                pass

        submitted_at = ''
        if sub.submitted_at:
            submitted_at = sub.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if hasattr(sub.submitted_at, 'strftime') else str(sub.submitted_at)

        wb = Workbook()
        ws = wb.active
        ws.title = '投诉结果'
        ws.append(['采集时间', '采集账号', '作品名称', '投诉单号'])
        for cell in ws[1]:
            cell.font = Font(bold=True)

        has_per_work = any(getattr(w, 'feedback_numbers', None) for w in works)
        number_idx = 0
        for work in works:
            if has_per_work:
                nums = []
                raw = getattr(work, 'feedback_numbers', None)
                if raw:
                    try:
                        nums = json.loads(raw) if isinstance(raw, str) else list(raw)
                    except (TypeError, json.JSONDecodeError):
                        nums = []
                if not nums:
                    nums = ['']
                for num in nums:
                    ws.append([submitted_at, sub.collect_account, work.work_name, str(num)])
            else:
                num = complaint_numbers[number_idx] if number_idx < len(complaint_numbers) else ''
                number_idx += 1
                ws.append([submitted_at, sub.collect_account, work.work_name, str(num)])

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 30

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True,
                         download_name=f'xiaohongshu_{submission_id}.xlsx')
    finally:
        db.close()


# ── task status ────────────────────────────────────────────────────────────────

@xiaohongshu_bp.route('/task/<task_id>', methods=['GET'])
@login_required
def xiaohongshu_task_status(task_id):
    db = get_db_session()
    try:
        row = db.execute(text("""
            SELECT task_id, status, batch_count, completed_batches, failed_batches,
                   complaint_numbers_json, error_message,
                   submitted_at, started_at, completed_at
            FROM complaints WHERE task_id = :tid
        """), {'tid': task_id}).fetchone()
        if not row:
            mem = _tasks().get(task_id)
            if mem:
                return jsonify({'success': True, 'task': mem})
            return jsonify({'success': False, 'error': '任务不存在'}), 404
        return jsonify({
            'success': True,
            'task': {
                'task_id': row.task_id,
                'status': row.status,
                'batch_count': row.batch_count,
                'completed_batches': row.completed_batches,
                'failed_batches': row.failed_batches,
                'complaint_numbers': json.loads(row.complaint_numbers_json) if row.complaint_numbers_json else [],
                'error_message': row.error_message,
                'submitted_at': normalize_datetime(row.submitted_at),
                'started_at': normalize_datetime(row.started_at),
                'completed_at': normalize_datetime(row.completed_at),
            }
        })
    finally:
        db.close()
